from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.http import HttpResponse
import io
from collections import defaultdict
from django.db.models import Count, Sum, Value, Q
from django.db.models.functions import Coalesce
from django.db.utils import OperationalError, ProgrammingError
from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl import Workbook, load_workbook

from .models import (
    Campaign, CampaignQueue, CampaignProgram,
    CampaignRegion, CampaignOrganization,
    QueueStageDeadline, Lead, LeadChecklistValue, LeadChecklistAttachment,
    LeadInteraction,
    LeadActivityLog,
    CampaignSubfunnel,
    LeadSubfunnel,
    LeadSubfunnelChecklistValue,
)
from .task_workflow import TASK_WORKFLOW_STATUS_VALUES
from .db_compat import lead_table_has_quota_split_columns
from apps.funnels.models import StageChecklistItem, FunnelStage, SubfunnelTemplate
from .serializers import (
    CampaignListSerializer, CampaignDetailSerializer,
    CampaignCreateSerializer, CampaignQueueSerializer,
    CampaignProgramSerializer, CampaignRegionSerializer,
    CampaignOrganizationSerializer,
    QueueStageDeadlineSerializer,
    LeadListSerializer, LeadDetailSerializer,
    LeadChecklistValueSerializer, LeadInteractionSerializer,
    CampaignSubfunnelSerializer,
    LeadSubfunnelSerializer,
    LeadSubfunnelChecklistValueSerializer,
)
from apps.funnels.models import SubfunnelTemplateItem, TaskTemplateStage, SubfunnelTemplateBinding


def _parse_bulk_ids(raw_ids):
    if not isinstance(raw_ids, list) or not raw_ids:
        return None, Response(
            {"detail": "Передайте непустой список ids."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        return [int(x) for x in raw_ids], None
    except (TypeError, ValueError):
        return None, Response(
            {"detail": "ids должны быть числами."},
            status=status.HTTP_400_BAD_REQUEST,
        )


def _normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _normalize_header(value):
    return _normalize_cell(value).lower().replace("ё", "е")


def _parse_non_negative_int(raw, *, field_label):
    s = _normalize_cell(raw)
    if not s:
        return None
    try:
        value = int(float(s)) if isinstance(raw, float) else int(s)
    except (TypeError, ValueError):
        raise ValueError(f"«{field_label}»: ожидается целое число")
    if value < 0:
        raise ValueError(f"«{field_label}»: значение не может быть отрицательным")
    return value


def _checklist_value_to_template_text(value):
    if not value:
        return ""
    parts = []
    if value.select_value:
        parts.append(f"выбор: {value.select_value}")
    if value.text_value:
        txt = _normalize_cell(value.text_value)
        if txt:
            parts.append(txt)
    contact_name = value.contact_name or (value.contact.full_name if value.contact else "")
    if contact_name:
        if contact_name:
            parts.append(f"контакт: {contact_name}")
    if value.attachments.exists():
        parts.append("файл")
    elif value.file_value:
        parts.append("файл")
    if value.is_completed:
        parts.append("выполнено")
    return "; ".join(dict.fromkeys(parts))


def _build_leads_demand_import_column_map(header_values):
    aliases = {
        "lead_id": {"lead_id", "lead id", "id лида", "id lead", "лид id"},
        "forecast_demand": {"план", "план (прогноз)", "план прогноз", "forecast_demand", "forecast"},
        "demand_quota_declared": {"заявленная", "заявленная квота", "квота заявленная", "demand_quota_declared"},
        "demand_quota_list": {
            "списочная", "списочная (факт)", "списочная факт", "квота списочная", "demand_quota_list", "факт"
        },
    }
    out = {}
    if not header_values:
        return out
    for idx, header in enumerate(header_values):
        normalized = _normalize_header(header)
        if not normalized:
            continue
        for key, names in aliases.items():
            if normalized in names:
                out[key] = idx
                break
    return out


def _empty_workload_task_stats():
    return {
        "total": 0,
        "backlog": 0,
        "in_progress": 0,
        "paused": 0,
        "rejected": 0,
        "done": 0,
        "overdue": 0,
    }


def _empty_activity_by_period_bucket():
    return {
        "date": None,
        "backlog": 0,
        "in_progress": 0,
        "paused": 0,
        "rejected": 0,
        "done": 0,
    }


def _bump_workload_task_stats(stats, status, is_overdue):
    stats["total"] += 1
    normalized = LeadSubfunnel.normalize_status(status)
    if normalized in stats:
        stats[normalized] += 1
    if is_overdue:
        stats["overdue"] += 1


def _log_lead_activity(lead, user, event_type, summary):
    if not summary:
        return
    try:
        LeadActivityLog.objects.create(
            lead=lead,
            event_type=event_type,
            summary=summary[:500],
            created_by=user if getattr(user, "is_authenticated", False) else None,
        )
    except (OperationalError, ProgrammingError):
        # Таблица ещё не создана (миграции не применены) — не блокируем основное действие
        pass


def _extract_forwarded_from_notes(notes):
    first_line = ((notes or "").splitlines() or [""])[0].strip()
    prefix = "Передано от организации:"
    if not first_line.startswith(prefix):
        return None
    value = first_line[len(prefix):].strip()
    if ". Комментарий:" in value:
        value = value.split(". Комментарий:", 1)[0].strip()
    return value or None


def _append_synthetic_checklist_from_values(lead, items):
    """Добавляет в ленту отметки по чек-листу из completed_at, если в журнале нет той же записи."""
    existing = set()
    for it in items:
        if it.get("kind") in ("checklist", "stage"):
            existing.add((it.get("summary", ""), (it.get("at") or "")[:19]))

    for cv in (
        lead.checklist_values.select_related("checklist_item", "completed_by")
        .filter(completed_at__isnull=False, is_completed=True)
    ):
        text = cv.checklist_item.text
        summary = f"Отмечен пункт «{text}»"
        at_str = cv.completed_at.isoformat()
        sig = (summary, at_str[:19])
        if sig in existing:
            continue
        existing.add(sig)
        u = cv.completed_by
        row = {
            "kind": "checklist",
            "id": 10_000_000 + cv.pk,
            "at": at_str,
            "summary": summary,
            "created_by_name": str(u) if u else None,
        }
        if cv.contact_id:
            row["contact_id"] = cv.contact_id
        items.append(row)


def _filter_timeline_items(items, kinds, contact_id):
    """Фильтр ?kind=interaction,stage&contact=12 — по контакту только события, где он указан."""
    if not kinds and contact_id is None:
        return items
    out = []
    for it in items:
        k = it.get("kind")
        if kinds and k not in kinds:
            continue
        if contact_id is not None:
            if k == "interaction":
                cid = (it.get("data") or {}).get("contact")
                if cid != contact_id:
                    continue
            elif k == "checklist":
                if it.get("contact_id") != contact_id:
                    continue
            else:
                continue
        out.append(it)
    return out


_COMMUNICATION_STATUS_BY_STEP = {
    "email_prepared": Lead.PrimaryContactStatus.EMAIL_PREPARED,
    "email_sent": Lead.PrimaryContactStatus.EMAIL_SENT,
    "response_received": Lead.PrimaryContactStatus.RESPONSE_RECEIVED,
    "result_recorded": Lead.PrimaryContactStatus.RESULT_RECORDED,
}


def _sync_primary_contact_status_from_checklist(value):
    step = (value.checklist_item.communication_step or "").strip()
    if not step:
        return
    lead = value.lead
    next_status = _COMMUNICATION_STATUS_BY_STEP.get(step)
    if not next_status:
        return

    if value.is_completed:
        update_fields = []
        if lead.primary_contact_status != next_status:
            lead.primary_contact_status = next_status
            update_fields.append("primary_contact_status")
        if step == "result_recorded":
            next_result = (
                (value.text_value or "").strip()
                or (value.select_value or "").strip()
                or lead.primary_contact_result
            )
            if next_result != lead.primary_contact_result:
                lead.primary_contact_result = next_result
                update_fields.append("primary_contact_result")
        if (
            value.primary_contact_specialist_id
            and lead.primary_contact_specialist_id != value.primary_contact_specialist_id
        ):
            lead.primary_contact_specialist_id = value.primary_contact_specialist_id
            update_fields.append("primary_contact_specialist")
        if update_fields:
            lead.save(update_fields=update_fields + ["updated_at"])
    elif lead.primary_contact_status == next_status:
        lead.primary_contact_status = Lead.PrimaryContactStatus.NEW
        lead.save(update_fields=["primary_contact_status", "updated_at"])


def _organization_ids_from_xlsx_bytes(raw: bytes, import_kind: str):
    from apps.organizations.views import _xlsx_rows, _find_organization_by_ref, _normalize_cell

    rows = _xlsx_rows(io.BytesIO(raw), import_kind=import_kind)
    ids = []
    seen = set()
    last_org_ref = ""
    for _line_no, row in rows:
        if import_kind == "contacts":
            org_raw = _normalize_cell(row.get("organization")) or last_org_ref
        else:
            org_raw = _normalize_cell(row.get("organization") or row.get("name") or row.get("inn"))
        if not org_raw:
            continue
        last_org_ref = org_raw
        found, _, _ = _find_organization_by_ref(org_raw)
        if found and found.id not in seen:
            seen.add(found.id)
            ids.append(found.id)
    return ids


def _run_registry_import_xlsx(viewset_cls, request, raw: bytes, filename: str, extra_data=None):
    from rest_framework.test import APIRequestFactory, force_authenticate

    bio = io.BytesIO(raw)
    bio.name = filename
    factory = APIRequestFactory()
    data = {"update_existing": "true", "source": "bulk", **(extra_data or {})}
    req = factory.post("/import-xlsx/", data, format="multipart")
    req.FILES["file"] = bio
    req.user = request.user
    force_authenticate(req, user=request.user)
    view = viewset_cls.as_view({"post": "import_xlsx"})
    return view(req)


def _link_orgs_to_collect_campaign(
    campaign,
    organization_ids,
    errors,
    campaign_region_id=None,
    organization_contact_map=None,
    source_lead_id=None,
    source_transfer_comment="",
):
    from apps.organizations.models import Contact, Organization

    from .collect_tasks import get_entry_funnel_stage_for_lead

    campaign_regions_qs = campaign.campaign_regions.select_related("queue", "region")
    if campaign_region_id:
        campaign_regions_qs = campaign_regions_qs.filter(id=campaign_region_id)
        if not campaign_regions_qs.exists():
            errors.append("Выбранный регион задачи не принадлежит кампании.")
            return {"leads_created": 0, "leads_by_region": {}, "linked_ids": []}
    campaign_regions = {cr.region_id: cr for cr in campaign_regions_qs}

    funnel_link = campaign.campaign_funnels.select_related("funnel").first()
    if not funnel_link:
        errors.append("У кампании не выбрана воронка.")
        return {"leads_created": 0, "leads_by_region": {}, "linked_ids": []}

    funnel = funnel_link.funnel
    entry_stage = get_entry_funnel_stage_for_lead(funnel)
    first_queue = campaign.queues.order_by("queue_number").first()
    leads_created = 0
    skipped_wrong_region = 0
    leads_by_region = defaultdict(int)
    linked_ids = set()
    source_org_name = None
    source_lead_defaults = None
    source_transfer_comment = (source_transfer_comment or "").strip()
    if source_lead_id:
        source_lead = (
            Lead.objects.select_related("organization", "queue")
            .filter(id=source_lead_id, campaign=campaign)
            .only(
                "id",
                "organization__name",
                "queue_id",
                "manager_id",
                "primary_contact_specialist_id",
            )
            .first()
        )
        if source_lead and source_lead.organization_id and source_lead.organization:
            source_org_name = source_lead.organization.name
            source_lead_defaults = source_lead
        else:
            errors.append("Лид-источник для передачи не найден в этой кампании.")

    for org_id in organization_ids:
        try:
            org = Organization.objects.get(id=org_id)
        except Organization.DoesNotExist:
            continue
        if campaign_regions and org.region_id not in campaign_regions:
            skipped_wrong_region += 1
            continue
        cr = campaign_regions.get(org.region_id)
        queue = (cr.queue if cr else None) or (source_lead_defaults.queue if source_lead_defaults else None) or first_queue
        lead, created = Lead.objects.get_or_create(
            campaign=campaign,
            organization_id=org.id,
            funnel_id=funnel.id,
            region_id=org.region_id,
            defaults={
                "queue": queue,
                "manager_id": (cr.manager_id if cr else None) or (source_lead_defaults.manager_id if source_lead_defaults else None),
                "primary_contact_specialist_id": (
                    (cr.primary_contact_specialist_id if cr else None)
                    or (source_lead_defaults.primary_contact_specialist_id if source_lead_defaults else None)
                ),
            },
        )
        update_fields = []
        if entry_stage and lead.current_stage_id != entry_stage.id:
            lead.current_stage = entry_stage
            update_fields.append("current_stage")
        if (
            entry_stage
            and entry_stage.primary_contact_specialist_id
            and not lead.primary_contact_specialist_id
        ):
            lead.primary_contact_specialist_id = entry_stage.primary_contact_specialist_id
            update_fields.append("primary_contact_specialist")
        if update_fields:
            lead.save(update_fields=update_fields + ["updated_at"])
        selected_contact_id = None
        if organization_contact_map:
            selected_contact_id = organization_contact_map.get(org.id)
        if selected_contact_id:
            contact = (
                Contact.objects.filter(id=selected_contact_id, organization_id=org.id)
                .only("id")
                .first()
            )
            if contact:
                if lead.primary_contact_id != contact.id:
                    lead.primary_contact_id = contact.id
                    lead.save(update_fields=["primary_contact", "updated_at"])
            else:
                errors.append(
                    f"Контакт {selected_contact_id} не найден в организации {org.name}."
                )
        CampaignOrganization.objects.get_or_create(campaign=campaign, organization_id=org.id)
        if source_org_name:
            transfer_note = f"Передано от организации: {source_org_name}"
            if source_transfer_comment:
                transfer_note = f"{transfer_note}. Комментарий: {source_transfer_comment}"
            current_notes = (lead.notes or "").strip()
            if not current_notes.startswith(transfer_note):
                lead.notes = f"{transfer_note}\n{current_notes}".strip() if current_notes else transfer_note
                lead.save(update_fields=["notes", "updated_at"])
        CampaignCreateSerializer._materialize_lead_subfunnels(lead, source="collect_import")
        if created:
            leads_created += 1
            leads_by_region[org.region_id] += 1
        linked_ids.add(org.id)

    if skipped_wrong_region:
        errors.append(
            f"Организаций вне регионов кампании (лиды не созданы): {skipped_wrong_region}"
        )
    return {
        "leads_created": leads_created,
        "leads_by_region": dict(leads_by_region),
        "linked_ids": list(linked_ids),
    }


class OrganizationListCaptureContactSerializer(serializers.Serializer):
    type = serializers.ChoiceField(choices=["person", "department", "main", "other"], required=False, default="person")
    first_name = serializers.CharField(required=False, allow_blank=True, default="")
    last_name = serializers.CharField(required=False, allow_blank=True, default="")
    middle_name = serializers.CharField(required=False, allow_blank=True, default="")
    department_name = serializers.CharField(required=False, allow_blank=True, default="")
    position = serializers.CharField(required=False, allow_blank=True, default="")
    phone = serializers.CharField(required=False, allow_blank=True, default="")
    phone_extension = serializers.CharField(required=False, allow_blank=True, default="")
    email = serializers.CharField(required=False, allow_blank=True, default="")
    messenger = serializers.CharField(required=False, allow_blank=True, default="")
    is_manager = serializers.BooleanField(required=False, default=False)
    comment = serializers.CharField(required=False, allow_blank=True, default="")


class OrganizationListCaptureOrganizationSerializer(serializers.Serializer):
    name = serializers.CharField(max_length=500)
    short_name = serializers.CharField(required=False, allow_blank=True, default="")
    inn = serializers.CharField(required=False, allow_blank=True, allow_null=True, default="")
    region_id = serializers.IntegerField(required=False, allow_null=True, default=None)
    org_type = serializers.CharField(required=False, allow_blank=True, default="other")
    parent_organization_id = serializers.IntegerField(required=False, allow_null=True, default=None)


class OrganizationListCaptureItemSerializer(serializers.Serializer):
    organization = OrganizationListCaptureOrganizationSerializer()
    contact = OrganizationListCaptureContactSerializer(required=False)


class OrganizationListCaptureRequestSerializer(serializers.Serializer):
    mode = serializers.ChoiceField(choices=["minimal", "full"], required=False, default="minimal")
    campaign_region_id = serializers.IntegerField(required=False, allow_null=True)
    source_lead_id = serializers.IntegerField(required=False, allow_null=True)
    source_transfer_comment = serializers.CharField(required=False, allow_blank=True, default="")
    items = OrganizationListCaptureItemSerializer(many=True, allow_empty=False)


class CampaignViewSet(viewsets.ModelViewSet):
    filterset_fields = ["status", "federal_operator", "project", "acting_organization", "responsible"]
    search_fields = ["name"]

    def get_queryset(self):
        qs = Campaign.objects.select_related(
            "federal_operator", "created_by", "responsible"
        ).prefetch_related(
            "tags",
            "queues__stage_deadlines",
            "campaign_funnels__funnel",
            "campaign_programs__program__profession",
            "campaign_regions__region__federal_district",
            "campaign_regions__queue",
            "subfunnels__template",
            "subfunnels__role",
            "subfunnels__default_assignee",
            "organizations__organization__region",
            "organizations__organization__tags",
            "leads__organization__region",
            "leads__region",
            "leads__funnel",
            "leads__current_stage",
            "leads__queue",
            "leads__manager",
            "leads__primary_contact",
            "leads__tags",
            "leads__organization__tags",
        )
        tag_ids = self.request.query_params.get("tags")
        if tag_ids:
            ids = [int(x) for x in tag_ids.split(",") if x.strip().isdigit()]
            if ids:
                qs = qs.filter(tags__id__in=ids).distinct()
        if self.action == "list":
            qs = qs.order_by("-created_at")
            # Без колонок 0006 annotate даёт SQL error → 500 на проде при пропущенном migrate
            if lead_table_has_quota_split_columns():
                qs = qs.annotate(
                    _d_plan=Coalesce(Sum("leads__forecast_demand"), Value(0)),
                    _d_cd=Coalesce(Sum("leads__demand_collected_declared"), Value(0)),
                    _d_cl=Coalesce(Sum("leads__demand_collected_list"), Value(0)),
                    _d_qd=Coalesce(Sum("leads__demand_quota_declared"), Value(0)),
                    _d_ql=Coalesce(Sum("leads__demand_quota_list"), Value(0)),
                )
        return qs

    def get_serializer_class(self):
        if self.action == "list":
            return CampaignListSerializer
        if self.action in ("create",):
            return CampaignCreateSerializer
        if self.action in ("update", "partial_update"):
            return CampaignCreateSerializer
        return CampaignDetailSerializer

    @action(detail=True, methods=["post"], url_path="programs")
    def add_programs(self, request, pk=None):
        campaign = self.get_object()
        program_ids = request.data.get("program_ids", [])
        created = []
        for pid in program_ids:
            obj, was_created = CampaignProgram.objects.get_or_create(
                campaign=campaign, program_id=pid,
                defaults={"manager_id": request.data.get("manager_id")},
            )
            if was_created:
                created.append(obj)
        return Response(
            CampaignProgramSerializer(created, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="regions")
    def add_regions(self, request, pk=None):
        campaign = self.get_object()
        regions_data = request.data.get("regions", [])
        created = []
        for rd in regions_data:
            obj, was_created = CampaignRegion.objects.get_or_create(
                campaign=campaign,
                region_id=rd["region_id"],
                defaults={
                    "queue_id": rd.get("queue_id"),
                    "manager_id": rd.get("manager_id"),
                },
            )
            if was_created:
                created.append(obj)
        return Response(
            CampaignRegionSerializer(created, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="organizations")
    def add_organizations(self, request, pk=None):
        campaign = self.get_object()
        org_ids = request.data.get("organization_ids", [])
        created = []
        for oid in org_ids:
            obj, was_created = CampaignOrganization.objects.get_or_create(
                campaign=campaign,
                organization_id=oid,
                defaults={"manager_id": request.data.get("manager_id")},
            )
            if was_created:
                created.append(obj)
        return Response(
            CampaignOrganizationSerializer(created, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="leads")
    def add_leads(self, request, pk=None):
        campaign = self.get_object()
        lead_data = request.data.get("leads", [])
        first_queue = campaign.queues.order_by("queue_number").first()
        created = []
        for ld in lead_data:
            queue_id = ld.get("queue_id") or (first_queue.id if first_queue else None)
            region_id = ld.get("region_id")
            if region_id is None:
                from apps.organizations.models import Organization
                region_id = (
                    Organization.objects.filter(id=ld["organization_id"])
                    .values_list("region_id", flat=True)
                    .first()
                )
            obj, was_created = Lead.objects.get_or_create(
                campaign=campaign,
                organization_id=ld["organization_id"],
                funnel_id=ld["funnel_id"],
                region_id=region_id,
                defaults={
                    "queue_id": queue_id,
                    "manager_id": ld.get("manager_id"),
                    "forecast_demand": ld.get("forecast_demand"),
                },
            )
            if was_created:
                created.append(obj)
        if campaign.status == Campaign.Status.ACTIVE and created:
            CampaignCreateSerializer._activate_leads(campaign)
        return Response(
            LeadListSerializer(created, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="collect-stage-import")
    def collect_stage_import(self, request, pk=None):
        """Импорт организаций и контактов в кампанию с нулевой стадией (по регионам отбора)."""
        from apps.organizations.views import ContactViewSet, OrganizationViewSet

        campaign = self.get_object()
        force_task_addition = str(request.data.get("force_task_addition", "")).lower() in {"1", "true", "yes", "on"}
        if (
            campaign.operational_stage != Campaign.OperationalStage.ORGANIZATION_LIST
            and not force_task_addition
        ):
            return Response(
                {"detail": "Импорт доступен только на стадии кампании «Формирование перечня организаций»."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        campaign_region_id = request.data.get("campaign_region_id")
        if campaign_region_id not in (None, ""):
            try:
                campaign_region_id = int(campaign_region_id)
            except (TypeError, ValueError):
                return Response(
                    {"detail": "campaign_region_id должен быть числом."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            campaign_region_id = None
        source_lead_id = request.data.get("source_lead_id")
        if source_lead_id not in (None, ""):
            try:
                source_lead_id = int(source_lead_id)
            except (TypeError, ValueError):
                return Response(
                    {"detail": "source_lead_id должен быть числом."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            source_lead_id = None
        source_transfer_comment = (request.data.get("source_transfer_comment") or "").strip()

        org_file = request.FILES.get("organizations_file") or request.FILES.get("file")
        contacts_file = request.FILES.get("contacts_file")

        if not org_file and not contacts_file:
            return Response(
                {"detail": "Передайте organizations_file и/или contacts_file (.xlsx)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        errors = []
        org_import = None
        contact_import = None
        organization_ids = []

        if org_file:
            if not org_file.name.lower().endswith(".xlsx"):
                return Response(
                    {"detail": "Файл организаций: только .xlsx"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            raw_org = org_file.read()
            org_import = _run_registry_import_xlsx(
                OrganizationViewSet,
                request,
                raw_org,
                org_file.name,
                {
                    "default_org_type": request.data.get("default_org_type") or "other",
                    "tag_ids": request.data.get("organization_tag_ids") or "",
                },
            )
            if org_import.status_code >= 400:
                return org_import
            organization_ids = _organization_ids_from_xlsx_bytes(raw_org, "organizations")

        if contacts_file:
            if not contacts_file.name.lower().endswith(".xlsx"):
                return Response(
                    {"detail": "Файл контактов: только .xlsx"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            raw_contacts = contacts_file.read()
            contact_import = _run_registry_import_xlsx(
                ContactViewSet,
                request,
                raw_contacts,
                contacts_file.name,
                {
                    "default_org_type": request.data.get("default_org_type") or "other",
                    "create_missing_organizations": "true",
                    "organization_tag_ids": request.data.get("organization_tag_ids") or "",
                    "contact_tag_ids": request.data.get("contact_tag_ids") or "",
                },
            )
            if contact_import.status_code >= 400:
                return contact_import
            contact_org_ids = _organization_ids_from_xlsx_bytes(raw_contacts, "contacts")
            organization_ids = list({*organization_ids, *contact_org_ids})

        link_result = _link_orgs_to_collect_campaign(
            campaign,
            organization_ids,
            errors,
            campaign_region_id=campaign_region_id,
            source_lead_id=source_lead_id,
            source_transfer_comment=source_transfer_comment,
        )

        payload = {
            "leads_created": link_result["leads_created"],
            "organizations_linked": len(link_result["linked_ids"]),
            "leads_by_region": link_result["leads_by_region"],
            "errors": errors,
        }
        if org_import is not None:
            payload["organizations_import"] = org_import.data
        if contact_import is not None:
            payload["contacts_import"] = contact_import.data
        return Response(payload)

    @action(detail=True, methods=["post"], url_path="organization-list-select")
    def organization_list_select(self, request, pk=None):
        """Выбор организаций/контактов из базы для региональной задачи."""
        campaign = self.get_object()
        force_task_addition = str(request.data.get("force_task_addition", "")).lower() in {"1", "true", "yes", "on"}
        if (
            campaign.operational_stage != Campaign.OperationalStage.ORGANIZATION_LIST
            and not force_task_addition
        ):
            return Response(
                {"detail": "Добавление доступно только на стадии «Формирование перечня организаций»."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        campaign_region_id = request.data.get("campaign_region_id")
        if campaign_region_id not in (None, ""):
            try:
                campaign_region_id = int(campaign_region_id)
            except (TypeError, ValueError):
                return Response(
                    {"detail": "campaign_region_id должен быть числом."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            campaign_region_id = None
        source_lead_id = request.data.get("source_lead_id")
        if source_lead_id not in (None, ""):
            try:
                source_lead_id = int(source_lead_id)
            except (TypeError, ValueError):
                return Response(
                    {"detail": "source_lead_id должен быть числом."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            source_lead_id = None
        source_transfer_comment = (request.data.get("source_transfer_comment") or "").strip()

        raw_items = request.data.get("items")
        if not isinstance(raw_items, list) or not raw_items:
            return Response(
                {"detail": "Передайте непустой список items."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        organization_ids = []
        organization_contact_map = {}
        for idx, item in enumerate(raw_items, start=1):
            if not isinstance(item, dict):
                return Response(
                    {"detail": f"items[{idx}] должен быть объектом."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            org_id_raw = item.get("organization_id")
            if not org_id_raw or not str(org_id_raw).isdigit():
                return Response(
                    {"detail": f"items[{idx}].organization_id должен быть числом."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            org_id = int(org_id_raw)
            organization_ids.append(org_id)

            contact_id_raw = item.get("contact_id")
            if contact_id_raw in (None, ""):
                continue
            if not str(contact_id_raw).isdigit():
                return Response(
                    {"detail": f"items[{idx}].contact_id должен быть числом."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            organization_contact_map[org_id] = int(contact_id_raw)

        errors = []
        deduped_org_ids = list(dict.fromkeys(organization_ids))
        link_result = _link_orgs_to_collect_campaign(
            campaign,
            deduped_org_ids,
            errors,
            campaign_region_id=campaign_region_id,
            organization_contact_map=organization_contact_map,
            source_lead_id=source_lead_id,
            source_transfer_comment=source_transfer_comment,
        )

        return Response(
            {
                "leads_created": link_result["leads_created"],
                "organizations_linked": len(link_result["linked_ids"]),
                "leads_by_region": link_result["leads_by_region"],
                "errors": errors,
            }
        )

    @action(detail=True, methods=["post"], url_path="organization-list-capture")
    def organization_list_capture(self, request, pk=None):
        from apps.organizations.models import Contact, Organization

        campaign = self.get_object()
        force_task_addition = str(request.data.get("force_task_addition", "")).lower() in {"1", "true", "yes", "on"}
        if (
            campaign.operational_stage != Campaign.OperationalStage.ORGANIZATION_LIST
            and not force_task_addition
        ):
            return Response(
                {"detail": "Добавление доступно только на стадии «Формирование перечня организаций»."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        req = OrganizationListCaptureRequestSerializer(data=request.data)
        req.is_valid(raise_exception=True)
        payload = req.validated_data
        campaign_region_id = payload.get("campaign_region_id")
        source_lead_id = payload.get("source_lead_id")
        source_transfer_comment = (payload.get("source_transfer_comment") or "").strip()

        results = []
        errors = []
        created_count = 0
        skipped_count = 0

        for idx, item in enumerate(payload["items"], start=1):
            org_data = item["organization"]
            contact_data = item.get("contact")

            try:
                org_type = org_data.get("org_type") or "other"
                parent_organization_id = org_data.get("parent_organization_id")
                if org_type == "company_branch" and not parent_organization_id:
                    raise ValueError("Для подразделения укажите головную организацию.")

                inn = (org_data.get("inn") or "").strip()
                if org_type == "company_branch":
                    inn = ""
                region_id = org_data.get("region_id")
                org = None
                if inn:
                    org = Organization.objects.filter(inn=inn).first()
                if org is None:
                    org = Organization.objects.filter(
                        name=org_data["name"].strip(),
                        region_id=region_id,
                    ).first()

                if org is None:
                    create_kwargs = {
                        "name": org_data["name"].strip(),
                        "short_name": (org_data.get("short_name") or org_data["name"]).strip()[:200],
                        "inn": inn or None,
                        "region_id": region_id,
                        "org_type": org_type,
                    }
                    if parent_organization_id:
                        create_kwargs["parent_organization_id"] = parent_organization_id
                    org = Organization.objects.create(**create_kwargs)

                contact = None
                if contact_data:
                    c_type = contact_data.get("type") or Contact.ContactType.PERSON
                    if c_type not in {item.value for item in Contact.ContactType}:
                        c_type = Contact.ContactType.PERSON
                    contact = Contact.objects.create(
                        organization=org,
                        type=c_type,
                        first_name=(contact_data.get("first_name") or "").strip(),
                        last_name=(contact_data.get("last_name") or "").strip(),
                        middle_name=(contact_data.get("middle_name") or "").strip(),
                        department_name=(contact_data.get("department_name") or "").strip(),
                        position=(contact_data.get("position") or "").strip(),
                        phone=(contact_data.get("phone") or "").strip(),
                        phone_extension=(contact_data.get("phone_extension") or "").strip(),
                        email=(contact_data.get("email") or "").strip(),
                        messenger=(contact_data.get("messenger") or "").strip(),
                        is_manager=bool(contact_data.get("is_manager")),
                        comment=(contact_data.get("comment") or "").strip(),
                    )

                link_errors = []
                link_result = _link_orgs_to_collect_campaign(
                    campaign,
                    [org.id],
                    link_errors,
                    campaign_region_id=campaign_region_id,
                    source_lead_id=source_lead_id,
                    source_transfer_comment=source_transfer_comment,
                )
                if link_errors:
                    skipped_count += 1
                    errors.extend([f"#{idx}: {msg}" for msg in link_errors])
                    continue

                created = bool(link_result["leads_created"])
                if created:
                    created_count += 1
                else:
                    skipped_count += 1

                lead = Lead.objects.filter(campaign=campaign, organization_id=org.id).order_by("-id").first()
                results.append(
                    {
                        "organization_id": org.id,
                        "organization_name": org.name,
                        "contact_id": contact.id if contact else None,
                        "lead_id": lead.id if lead else None,
                        "created": created,
                    }
                )
            except Exception as exc:
                skipped_count += 1
                errors.append(f"#{idx}: {exc}")

        return Response(
            {
                "results": results,
                "summary": {
                    "created": created_count,
                    "skipped": skipped_count,
                    "errors": errors,
                },
            }
        )

    @action(detail=True, methods=["get", "post"], url_path="subfunnels")
    def subfunnels(self, request, pk=None):
        campaign = self.get_object()
        if request.method == "GET":
            rows = campaign.subfunnels.select_related(
                "template", "role", "default_assignee", "funnel", "binding"
            )
            return Response(CampaignSubfunnelSerializer(rows, many=True).data)

        serializer = CampaignSubfunnelSerializer(
            data={**request.data, "campaign": campaign.pk}
        )
        serializer.is_valid(raise_exception=True)
        obj = serializer.save(template_version=serializer.validated_data["template"].version)
        return Response(CampaignSubfunnelSerializer(obj).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"], url_path="subfunnel-workspace")
    def subfunnel_workspace(self, request):
        campaign_id = request.query_params.get("campaign")
        template_id = request.query_params.get("template") or request.query_params.get("subfunnel")
        role_id = request.query_params.get("role")
        assignee_id = request.query_params.get("assignee")
        status_filter = request.query_params.get("status")
        overdue = request.query_params.get("overdue")
        view_mode = request.query_params.get("view_mode", "kanban")

        from django.db.models import Q

        rows = LeadSubfunnel.objects.select_related(
            "lead__campaign",
            "lead__organization",
            "lead__current_stage",
            "campaign_region__campaign",
            "campaign_region__region",
            "campaign_subfunnel__template",
            "campaign_subfunnel__role",
            "current_template_stage",
            "assignee",
        ).prefetch_related("checklist_values__template_item")
        cid = None
        if campaign_id and str(campaign_id).isdigit():
            cid = int(campaign_id)
            rows = rows.filter(
                Q(lead__campaign_id=cid) | Q(campaign_region__campaign_id=cid)
            )
        if role_id and str(role_id).isdigit():
            rows = rows.filter(campaign_subfunnel__role_id=int(role_id))
        if assignee_id and str(assignee_id).isdigit():
            rows = rows.filter(assignee_id=int(assignee_id))
        if overdue in ("1", "true", "True"):
            rows = rows.filter(due_at__lt=timezone.now()).exclude(status=LeadSubfunnel.Status.DONE)

        templates = list(
            rows.values(
                "campaign_subfunnel__template_id",
                "campaign_subfunnel__template__name",
            )
            .annotate(count=Count("id"))
            .order_by("campaign_subfunnel__template__name")
        )
        template_counts = {
            row["campaign_subfunnel__template_id"]: row["count"]
            for row in templates
            if row["campaign_subfunnel__template_id"]
        }
        template_names = {
            row["campaign_subfunnel__template_id"]: row["campaign_subfunnel__template__name"]
            for row in templates
            if row["campaign_subfunnel__template_id"]
        }
        template_tabs_map = {}
        configured_templates = CampaignSubfunnel.objects.filter(
            is_active=True,
            template__is_active=True,
        )
        if cid is not None:
            configured_templates = configured_templates.filter(campaign_id=cid)
        if role_id and str(role_id).isdigit():
            configured_templates = configured_templates.filter(role_id=int(role_id))
        for row in configured_templates.values("template_id", "template__name").distinct():
            template_id_value = row["template_id"]
            if not template_id_value:
                continue
            template_tabs_map[template_id_value] = {
                "id": template_id_value,
                "name": row["template__name"] or template_names.get(template_id_value) or f"Шаблон #{template_id_value}",
                "count": template_counts.get(template_id_value, 0),
            }
        for template_id_value, count in template_counts.items():
            if template_id_value in template_tabs_map:
                continue
            template_tabs_map[template_id_value] = {
                "id": template_id_value,
                "name": template_names.get(template_id_value) or f"Шаблон #{template_id_value}",
                "count": count,
            }
        if not template_tabs_map:
            for template in SubfunnelTemplate.objects.filter(is_active=True).values("id", "name").order_by("name"):
                template_tabs_map[template["id"]] = {
                    "id": template["id"],
                    "name": template["name"] or f"Шаблон #{template['id']}",
                    "count": 0,
                }
        template_tabs = sorted(template_tabs_map.values(), key=lambda item: item["name"] or "")

        active_template_id = None
        if template_id and str(template_id).isdigit():
            candidate = int(template_id)
            if any(tab["id"] == candidate for tab in template_tabs):
                active_template_id = candidate
        if active_template_id is None and template_tabs:
            active_template_id = template_tabs[0]["id"]
        if active_template_id is not None:
            rows = rows.filter(campaign_subfunnel__template_id=active_template_id)
        if status_filter:
            status_filter_str = str(status_filter)
            if status_filter_str in TASK_WORKFLOW_STATUS_VALUES:
                rows = rows.filter(status=status_filter_str)
            elif status_filter_str.startswith("stage-"):
                stage_tail = status_filter_str.removeprefix("stage-")
                if stage_tail == "unassigned":
                    rows = rows.filter(current_template_stage_id__isnull=True)
                elif stage_tail.isdigit():
                    rows = rows.filter(current_template_stage_id=int(stage_tail))
            elif status_filter_str.isdigit():
                rows = rows.filter(current_template_stage_id=int(status_filter_str))

        capture_counts_by_region = {}
        region_rows = list(
            rows.filter(
                campaign_region_id__isnull=False,
                campaign_subfunnel__template__auto_create_on_collect_import=True,
            )
            .values(
                "campaign_region_id",
                "campaign_region__campaign_id",
                "campaign_region__region_id",
            )
            .distinct()
        )
        region_pairs = {
            (r["campaign_region__campaign_id"], r["campaign_region__region_id"])
            for r in region_rows
            if r["campaign_region__campaign_id"] and r["campaign_region__region_id"]
        }
        if region_pairs:
            pairs_query = Q()
            for campaign_id_value, region_id_value in region_pairs:
                pairs_query |= Q(campaign_id=campaign_id_value, region_id=region_id_value)
            lead_counts = Lead.objects.filter(pairs_query).values("campaign_id", "region_id").annotate(
                organizations_count=Count("organization_id", distinct=True),
                contacts_count=Count("primary_contact_id", filter=Q(primary_contact_id__isnull=False), distinct=True),
            )
            counts_by_pair = {
                (it["campaign_id"], it["region_id"]): {
                    "organizations": it["organizations_count"],
                    "contacts": it["contacts_count"],
                }
                for it in lead_counts
            }
            capture_counts_by_region = {
                r["campaign_region_id"]: counts_by_pair.get(
                    (r["campaign_region__campaign_id"], r["campaign_region__region_id"]),
                    {"organizations": 0, "contacts": 0},
                )
                for r in region_rows
            }

        stages = []
        if active_template_id is not None:
            stages = list(
                TaskTemplateStage.objects.filter(
                    template_id=active_template_id,
                    is_active=True,
                ).order_by("order", "id")
            )
        columns = [
            {
                "status": f"stage-{stage.id}",
                "stage_id": stage.id,
                "stage_name": stage.name,
                "order": stage.order,
                "is_work_stage": bool(stage.is_work_stage),
            }
            for stage in stages
        ]
        active_stage_ids = {col["stage_id"] for col in columns}

        payload = []
        now = timezone.now()
        for row in rows:
            checklist_values = sorted(
                row.checklist_values.all(),
                key=lambda v: (v.template_item.order, v.id),
            )
            checklist_total = len(checklist_values)
            checklist_completed = sum(1 for v in checklist_values if v.is_completed)
            checklist_summary = [
                {"text": v.template_item.title, "done": v.is_completed}
                for v in checklist_values
            ]
            is_region_task = bool(row.campaign_region_id)
            if is_region_task:
                campaign_obj = row.campaign_region.campaign
                lead_name = f"Регион: {row.campaign_region.region.name}"
                stage_name = (
                    campaign_obj.get_operational_stage_display()
                    if campaign_obj and campaign_obj.operational_stage
                    else None
                )
            else:
                campaign_obj = row.lead.campaign if row.lead else None
                lead_name = (
                    row.lead.organization.name
                    if row.lead and row.lead.organization
                    else (f"Лид {row.lead_id}" if row.lead_id else "—")
                )
                stage_name = row.lead.current_stage.name if row.lead and row.lead.current_stage else None
            show_capture_counts = bool(
                is_region_task
                and row.campaign_subfunnel.template.auto_create_on_collect_import
            )
            payload.append({
                "id": row.id,
                "campaign_id": campaign_obj.id if campaign_obj else None,
                "campaign_name": campaign_obj.name if campaign_obj else None,
                "lead_id": row.lead_id,
                "lead_name": lead_name,
                "forwarded_from": (
                    _extract_forwarded_from_notes(row.lead.notes)
                    if row.lead_id and row.lead
                    else None
                ),
                "is_region_task": is_region_task,
                "campaign_region_id": row.campaign_region_id,
                "region_id": row.campaign_region.region_id if row.campaign_region_id else None,
                "region_name": row.campaign_region.region.name if row.campaign_region_id else None,
                "stage_name": stage_name,
                "template_id": row.campaign_subfunnel.template_id,
                "template_name": row.campaign_subfunnel.template.name,
                "role_id": row.campaign_subfunnel.role_id,
                "role_name": row.campaign_subfunnel.role.name if row.campaign_subfunnel.role else None,
                "assignee_id": row.assignee_id,
                "assignee_name": str(row.assignee) if row.assignee else None,
                "status": LeadSubfunnel.normalize_status(row.status),
                "current_template_stage_id": row.current_template_stage_id,
                "current_template_stage_name": row.current_template_stage.name if row.current_template_stage else None,
                "current_template_stage_order": row.current_template_stage.order if row.current_template_stage else None,
                "board_stage_key": (
                    f"stage-{row.current_template_stage_id}"
                    if row.current_template_stage_id in active_stage_ids
                    else "stage-unassigned"
                ),
                "due_at": row.due_at.isoformat() if row.due_at else None,
                "is_overdue": bool(row.due_at and row.due_at < now and row.status != LeadSubfunnel.Status.DONE),
                "is_available": row.is_available,
                "checklist_progress": {
                    "total": checklist_total,
                    "completed": checklist_completed,
                },
                "checklist_summary": checklist_summary,
                "show_capture_counts": show_capture_counts,
                "capture_counts": (
                    capture_counts_by_region.get(
                        row.campaign_region_id,
                        {"organizations": 0, "contacts": 0},
                    )
                    if show_capture_counts
                    else None
                ),
            })

        kanban_map = defaultdict(list)
        for item in payload:
            workflow_status = LeadSubfunnel.normalize_status(item.get("status"))
            item["status"] = workflow_status
            kanban_map[item["board_stage_key"]].append(item)

        if kanban_map.get("stage-unassigned"):
            columns.append(
                {
                    "status": "stage-unassigned",
                    "stage_id": None,
                    "stage_name": "Без этапа",
                    "order": 10_000,
                    "is_work_stage": False,
                }
            )

        items_by_stage = {status: items for status, items in kanban_map.items()}
        kanban = []
        for col in columns:
            col_status = col["status"]
            items = kanban_map.get(col_status, [])
            kanban.append(
                {
                    "status": col_status,
                    "stage_id": col.get("stage_id"),
                    "stage_name": col["stage_name"],
                    "items": items,
                }
            )

        return Response({
            "view_mode": view_mode,
            "templates": template_tabs,
            "active_template_id": active_template_id,
            "columns": columns,
            "items_by_stage": items_by_stage,
            "kanban": kanban,
            "table": payload,
            "totals": {
                "all": len(payload),
                "overdue": sum(1 for i in payload if i["is_overdue"]),
                "backlog": sum(1 for i in payload if LeadSubfunnel.normalize_status(i["status"]) == LeadSubfunnel.Status.BACKLOG),
                "in_progress": sum(1 for i in payload if i["status"] == LeadSubfunnel.Status.IN_PROGRESS),
                "paused": sum(1 for i in payload if i["status"] == LeadSubfunnel.Status.PAUSED),
                "rejected": sum(1 for i in payload if i["status"] == LeadSubfunnel.Status.REJECTED),
                "done": sum(1 for i in payload if i["status"] == LeadSubfunnel.Status.DONE),
            },
        })

    @action(detail=True, methods=["get"], url_path="leads-demand-import-template")
    def leads_demand_import_template(self, request, pk=None):
        campaign = self.get_object()
        leads = list(
            campaign.leads.select_related(
                "organization", "region", "funnel", "current_stage", "queue"
            ).prefetch_related(
                "checklist_values__checklist_item",
                "checklist_values__attachments",
                "checklist_values__contact",
            ).order_by("id")
        )

        checklist_items = list(
            StageChecklistItem.objects.filter(
                stage__funnel__campaigns=campaign
            ).select_related("stage").order_by("stage__order", "order", "id")
        )
        checklist_columns = [
            (item.id, f"Чек-лист: {item.stage.name} / {item.text}")
            for item in checklist_items
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Потребность по лидам"

        base_headers = [
            "lead_id",
            "Организация",
            "Регион",
            "Воронка",
            "Очередь",
            "Этап",
            "План",
            "Заявленная",
            "Списочная (факт)",
        ]
        ws.append(base_headers + [label for _, label in checklist_columns])

        for lead in leads:
            values_by_item_id = {v.checklist_item_id: v for v in lead.checklist_values.all()}
            row = [
                lead.id,
                lead.organization.name if lead.organization else "",
                (lead.region.name if lead.region else (lead.organization.region.name if lead.organization and lead.organization.region else "")),
                lead.funnel.name if lead.funnel else "",
                lead.queue.name if lead.queue else "",
                lead.current_stage.name if lead.current_stage else "",
                lead.forecast_demand if lead.forecast_demand is not None else "",
                lead.demand_quota_declared,
                lead.demand_quota_list if lead.demand_quota_list is not None else lead.demand_count,
            ]
            for item_id, _ in checklist_columns:
                value = values_by_item_id.get(item_id)
                row.append(_checklist_value_to_template_text(value))
            ws.append(row)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="campaign_{campaign.id}_leads_demand_template.xlsx"'
        )
        return resp

    @action(detail=True, methods=["post"], url_path="leads-demand-import")
    def leads_demand_import(self, request, pk=None):
        campaign = self.get_object()
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Файл не передан"}, status=status.HTTP_400_BAD_REQUEST)
        if not file_obj.name.lower().endswith(".xlsx"):
            return Response({"detail": "Поддерживаются только .xlsx файлы"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            raw = file_obj.read()
            if not raw:
                return Response({"detail": "Файл пустой"}, status=status.HTTP_400_BAD_REQUEST)
            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
            ws = wb.active
        except Exception as exc:
            return Response(
                {"detail": f"Не удалось прочитать Excel (.xlsx): {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        col_map = _build_leads_demand_import_column_map(header_values)
        if "lead_id" not in col_map:
            return Response(
                {"detail": "В файле не найден столбец lead_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        updated = 0
        skipped = 0
        errors = []

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            lead_id_raw = row[col_map["lead_id"]] if col_map["lead_id"] < len(row) else None
            lead_id_text = _normalize_cell(lead_id_raw)
            if not lead_id_text:
                skipped += 1
                continue
            try:
                lead_id = int(float(lead_id_text)) if "." in lead_id_text else int(lead_id_text)
            except ValueError:
                skipped += 1
                errors.append(f"Строка {row_no}: некорректный lead_id «{lead_id_text}»")
                continue

            try:
                lead = campaign.leads.get(id=lead_id)
            except Lead.DoesNotExist:
                skipped += 1
                errors.append(f"Строка {row_no}: лид #{lead_id} не найден в этой кампании")
                continue

            try:
                plan = _parse_non_negative_int(
                    row[col_map["forecast_demand"]] if "forecast_demand" in col_map and col_map["forecast_demand"] < len(row) else None,
                    field_label="План",
                )
                quota_declared = _parse_non_negative_int(
                    row[col_map["demand_quota_declared"]] if "demand_quota_declared" in col_map and col_map["demand_quota_declared"] < len(row) else None,
                    field_label="Заявленная",
                )
                quota_list = _parse_non_negative_int(
                    row[col_map["demand_quota_list"]] if "demand_quota_list" in col_map and col_map["demand_quota_list"] < len(row) else None,
                    field_label="Списочная (факт)",
                )
            except ValueError as exc:
                skipped += 1
                errors.append(f"Строка {row_no}: {exc}")
                continue

            update_fields = []
            if plan is not None and lead.forecast_demand != plan:
                lead.forecast_demand = plan
                update_fields.append("forecast_demand")
            if quota_declared is not None and lead.demand_quota_declared != quota_declared:
                lead.demand_quota_declared = quota_declared
                update_fields.append("demand_quota_declared")
            if quota_list is not None:
                if lead.demand_quota_list != quota_list:
                    lead.demand_quota_list = quota_list
                    update_fields.append("demand_quota_list")
                if lead.demand_count != quota_list:
                    lead.demand_count = quota_list
                    update_fields.append("demand_count")

            if update_fields:
                lead.save(update_fields=list(dict.fromkeys(update_fields + ["updated_at"])))
                updated += 1
            else:
                skipped += 1

        return Response(
            {
                "updated": updated,
                "skipped": skipped,
                "errors": errors[:200],
            }
        )

    @action(detail=True, methods=["post"], url_path="assign-managers")
    def assign_managers(self, request, pk=None):
        campaign = self.get_object()
        assignments = request.data.get("assignments", [])
        updated = 0
        for a in assignments:
            level = a.get("level")
            target_id = a.get("target_id")
            manager_id = a.get("manager_id")

            if level == "program":
                CampaignProgram.objects.filter(
                    campaign=campaign, program_id=target_id
                ).update(manager_id=manager_id)
                updated += 1
            elif level == "region":
                CampaignRegion.objects.filter(
                    campaign=campaign, region_id=target_id
                ).update(manager_id=manager_id)
                updated += 1
            elif level == "organization":
                CampaignOrganization.objects.filter(
                    campaign=campaign, organization_id=target_id
                ).update(manager_id=manager_id)
                updated += 1
            elif level == "lead":
                Lead.objects.filter(
                    campaign=campaign, id=target_id
                ).update(manager_id=manager_id)
                updated += 1
            elif level == "region_specialist":
                CampaignRegion.objects.filter(
                    campaign=campaign, region_id=target_id
                ).update(primary_contact_specialist_id=manager_id)
                updated += 1
            elif level == "lead_specialist":
                Lead.objects.filter(
                    campaign=campaign, id=target_id
                ).update(primary_contact_specialist_id=manager_id)
                updated += 1
            elif level == "stage_specialist":
                FunnelStage.objects.filter(
                    id=target_id,
                    funnel__campaigns=campaign,
                ).update(primary_contact_specialist_id=manager_id)
                updated += 1
            elif level == "checklist_specialist":
                StageChecklistItem.objects.filter(
                    id=target_id,
                    stage__funnel__campaigns=campaign,
                ).update(primary_contact_specialist_id=manager_id)
                updated += 1
            elif level == "lead_checklist_specialist":
                LeadChecklistValue.objects.filter(
                    id=target_id,
                    lead__campaign=campaign,
                ).update(primary_contact_specialist_id=manager_id)
                updated += 1

        return Response({"updated": updated})

    @action(detail=False, methods=["get"], url_path="workload-dashboard")
    def workload_dashboard(self, request):
        role = request.query_params.get("role", "all")
        if role not in {"all", "manager", "specialist"}:
            role = "all"
        campaign_id = request.query_params.get("campaign")
        funnel_id = request.query_params.get("funnel")
        user_id = request.query_params.get("user")
        date_from_raw = request.query_params.get("date_from")
        date_to_raw = request.query_params.get("date_to")

        campaign_id_int = int(campaign_id) if campaign_id and str(campaign_id).isdigit() else None
        funnel_id_int = int(funnel_id) if funnel_id and str(funnel_id).isdigit() else None
        user_id_int = int(user_id) if user_id and str(user_id).isdigit() else None
        date_from = parse_date(date_from_raw) if date_from_raw else None
        date_to = parse_date(date_to_raw) if date_to_raw else None

        def _date_in_range(d):
            if not d:
                return False
            if date_from and d < date_from:
                return False
            if date_to and d > date_to:
                return False
            return True

        def _dt_in_range(dt):
            return _date_in_range(timezone.localtime(dt).date()) if dt else False

        def _is_open_on_period_end(created_at, completed_at=None):
            if not date_to:
                return False
            if timezone.localtime(created_at).date() > date_to:
                return False
            if not completed_at:
                return True
            return timezone.localtime(completed_at).date() > date_to

        has_period = bool(date_from or date_to)
        today = timezone.localdate()
        now = timezone.now()

        def _bump_activity_status(activity_map, date_obj, status):
            if not date_obj:
                return
            if has_period and not _date_in_range(date_obj):
                return
            normalized = LeadSubfunnel.normalize_status(status)
            if normalized not in {
                LeadSubfunnel.Status.BACKLOG,
                LeadSubfunnel.Status.IN_PROGRESS,
                LeadSubfunnel.Status.PAUSED,
                LeadSubfunnel.Status.REJECTED,
                LeadSubfunnel.Status.DONE,
            }:
                return
            key = date_obj.isoformat()
            bucket = activity_map[key]
            bucket["date"] = key
            bucket[normalized] += 1

        bucket = defaultdict(
            lambda: {
                "user_id": None,
                "user_name": None,
                "role": None,
                "active_leads": 0,
                "pending_checklist": 0,
                "overdue_stage": 0,
                "overdue_checklist": 0,
                "tasks_in_progress": 0,
                "tasks_overdue": 0,
            }
        )

        managers_map = defaultdict(
            lambda: {
                "user_id": None,
                "user_name": None,
                "campaigns": defaultdict(
                    lambda: {"campaign_id": None, "campaign_name": None, "leads": []}
                ),
            }
        )
        def _specialist_campaign_bucket():
            return {
                "campaign_id": None,
                "campaign_name": None,
                "stats": _empty_workload_task_stats(),
                "templates": defaultdict(
                    lambda: {
                        "template_id": None,
                        "template_name": None,
                        "stats": _empty_workload_task_stats(),
                    }
                ),
            }

        specialists_map = defaultdict(
            lambda: {
                "user_id": None,
                "user_name": None,
                "campaigns": defaultdict(_specialist_campaign_bucket),
            }
        )
        activity_by_period = defaultdict(_empty_activity_by_period_bucket)
        chart_data = {
            "manager": {
                "by_campaign": defaultdict(
                    lambda: {"campaign_id": None, "campaign_name": None, "in_progress": 0, "overdue": 0, "done_in_period": 0}
                ),
                "by_user": defaultdict(
                    lambda: {"user_id": None, "user_name": None, "in_progress": 0, "overdue": 0}
                ),
                "by_day": defaultdict(lambda: {"date": None, "opened": 0, "completed": 0, "overdue": 0}),
                "status_pie": defaultdict(int),
            },
            "specialist": {
                "by_campaign": defaultdict(
                    lambda: {"campaign_id": None, "campaign_name": None, "in_progress": 0, "overdue": 0, "done_in_period": 0}
                ),
                "by_user": defaultdict(
                    lambda: {"user_id": None, "user_name": None, "in_progress": 0, "overdue": 0}
                ),
                "by_day": defaultdict(lambda: {"date": None, "opened": 0, "completed": 0, "overdue": 0}),
                "status_pie": defaultdict(int),
            },
        }

        def add_row(user_obj, role_key, active=0, pending=0, overdue_stage=0, overdue_checklist=0, tasks_in_progress=0, tasks_overdue=0):
            if not user_obj:
                return
            if user_id_int and user_obj.id != user_id_int:
                return
            key = (role_key, user_obj.id)
            row = bucket[key]
            row["user_id"] = user_obj.id
            row["user_name"] = str(user_obj)
            row["role"] = role_key
            row["active_leads"] += active
            row["pending_checklist"] += pending
            row["overdue_stage"] += overdue_stage
            row["overdue_checklist"] += overdue_checklist
            row["tasks_in_progress"] += tasks_in_progress
            row["tasks_overdue"] += tasks_overdue

        leads_qs = Lead.objects.select_related(
            "campaign",
            "funnel",
            "current_stage",
            "manager",
            "primary_contact_specialist",
            "current_stage__primary_contact_specialist",
        ).prefetch_related("checklist_values__checklist_item")

        if campaign_id_int:
            leads_qs = leads_qs.filter(campaign_id=campaign_id_int)
        if funnel_id_int:
            leads_qs = leads_qs.filter(funnel_id=funnel_id_int)

        leads = [
            lead
            for lead in leads_qs
            if lead.current_stage_id and not lead.current_stage.is_rejection
        ]

        for lead in leads:
            stage_deadline = lead.get_stage_deadline(lead.current_stage)
            stage_overdue = bool(stage_deadline and stage_deadline < today)
            stage_items = [
                v
                for v in lead.checklist_values.all()
                if v.checklist_item.stage_id == lead.current_stage_id
            ]
            pending_items = [v for v in stage_items if not v.is_completed]

            if has_period:
                lead_active = (
                    _dt_in_range(lead.updated_at)
                    or any(_dt_in_range(v.updated_at) or _dt_in_range(v.completed_at) for v in stage_items)
                    or _date_in_range(stage_deadline)
                    or _is_open_on_period_end(lead.created_at)
                )
                if not lead_active:
                    continue

            if role in ("all", "manager"):
                add_row(
                    lead.manager,
                    "manager",
                    active=1,
                    pending=len(pending_items),
                    overdue_stage=1 if stage_overdue else 0,
                    overdue_checklist=len(pending_items) if stage_overdue else 0,
                )
                if lead.manager and (not user_id_int or lead.manager_id == user_id_int):
                    mgr = managers_map[lead.manager_id]
                    mgr["user_id"] = lead.manager_id
                    mgr["user_name"] = str(lead.manager)
                    mgr_campaign = mgr["campaigns"][lead.campaign_id]
                    mgr_campaign["campaign_id"] = lead.campaign_id
                    mgr_campaign["campaign_name"] = lead.campaign.name
                    mgr_campaign["leads"].append(
                        {
                            "lead_id": lead.id,
                            "organization_name": lead.organization.name if lead.organization else f"Лид {lead.id}",
                            "stage_name": lead.current_stage.name if lead.current_stage else None,
                            "stage_deadline": stage_deadline.isoformat() if stage_deadline else None,
                            "stage_overdue": stage_overdue,
                            "pending_checklist": len(pending_items),
                            "overdue_checklist": len(pending_items) if stage_overdue else 0,
                        }
                    )
                    mgr_chart_campaign = chart_data["manager"]["by_campaign"][lead.campaign_id]
                    mgr_chart_campaign["campaign_id"] = lead.campaign_id
                    mgr_chart_campaign["campaign_name"] = lead.campaign.name
                    mgr_chart_campaign["in_progress"] += 1
                    mgr_chart_campaign["overdue"] += 1 if stage_overdue else 0
                    mgr_chart_user = chart_data["manager"]["by_user"][lead.manager_id]
                    mgr_chart_user["user_id"] = lead.manager_id
                    mgr_chart_user["user_name"] = str(lead.manager)
                    mgr_chart_user["in_progress"] += 1
                    mgr_chart_user["overdue"] += 1 if stage_overdue else 0
                    chart_data["manager"]["status_pie"]["overdue" if stage_overdue else "in_progress"] += 1

            if role in ("all", "specialist"):
                lead_specialist = (
                    lead.primary_contact_specialist
                    or lead.current_stage.primary_contact_specialist
                )
                add_row(
                    lead_specialist,
                    "specialist",
                    active=1,
                    overdue_stage=1 if stage_overdue else 0,
                )
                for value in pending_items:
                    specialist = (
                        value.primary_contact_specialist
                        or value.checklist_item.primary_contact_specialist
                        or lead_specialist
                    )
                    add_row(
                        specialist,
                        "specialist",
                        pending=1,
                        overdue_checklist=1 if stage_overdue else 0,
                    )

        task_rows = LeadSubfunnel.objects.select_related(
            "lead__campaign",
            "lead__funnel",
            "lead__organization",
            "campaign_region__campaign",
            "campaign_region__region",
            "campaign_subfunnel__template",
            "campaign_subfunnel__funnel",
            "current_template_stage",
            "assignee",
        )
        if campaign_id_int:
            task_rows = task_rows.filter(
                Q(lead__campaign_id=campaign_id_int) | Q(campaign_region__campaign_id=campaign_id_int)
            )
        if funnel_id_int:
            task_rows = task_rows.filter(campaign_subfunnel__funnel_id=funnel_id_int)
        if user_id_int:
            task_rows = task_rows.filter(assignee_id=user_id_int)

        for task in task_rows:
            if not task.assignee_id:
                continue
            is_done = task.status == LeadSubfunnel.Status.DONE
            is_overdue = bool(task.due_at and task.due_at < now and not is_done)
            campaign_obj = task.lead.campaign if task.lead_id else (task.campaign_region.campaign if task.campaign_region_id else None)
            if not campaign_obj:
                continue

            activity_dt = (
                task.completed_at
                if is_done and task.completed_at
                else (task.updated_at or task.started_at or task.created_at)
            )
            if activity_dt:
                activity_status = LeadSubfunnel.Status.DONE if is_done else task.status
                _bump_activity_status(
                    activity_by_period,
                    timezone.localtime(activity_dt).date(),
                    activity_status,
                )

            if has_period:
                task_active = (
                    _dt_in_range(task.updated_at)
                    or _dt_in_range(task.started_at)
                    or _dt_in_range(task.completed_at)
                    or _is_open_on_period_end(task.created_at, task.completed_at if is_done else None)
                )
                if not task_active:
                    continue

            if role in ("all", "specialist"):
                add_row(
                    task.assignee,
                    "specialist",
                    pending=0 if is_done else 1,
                    overdue_checklist=1 if is_overdue else 0,
                    tasks_in_progress=0 if is_done else 1,
                    tasks_overdue=1 if is_overdue else 0,
                )
                specialist_item = specialists_map[task.assignee_id]
                specialist_item["user_id"] = task.assignee_id
                specialist_item["user_name"] = str(task.assignee)
                sp_campaign = specialist_item["campaigns"][campaign_obj.id]
                sp_campaign["campaign_id"] = campaign_obj.id
                sp_campaign["campaign_name"] = campaign_obj.name
                template = task.campaign_subfunnel.template if task.campaign_subfunnel else None
                template_id = template.id if template else 0
                sp_template = sp_campaign["templates"][template_id]
                sp_template["template_id"] = template_id
                sp_template["template_name"] = template.name if template else "Без шаблона"
                _bump_workload_task_stats(sp_campaign["stats"], task.status, is_overdue)
                _bump_workload_task_stats(sp_template["stats"], task.status, is_overdue)

                sp_chart_campaign = chart_data["specialist"]["by_campaign"][campaign_obj.id]
                sp_chart_campaign["campaign_id"] = campaign_obj.id
                sp_chart_campaign["campaign_name"] = campaign_obj.name
                if is_done:
                    if _dt_in_range(task.completed_at):
                        sp_chart_campaign["done_in_period"] += 1
                else:
                    sp_chart_campaign["in_progress"] += 1
                    sp_chart_campaign["overdue"] += 1 if is_overdue else 0
                sp_chart_user = chart_data["specialist"]["by_user"][task.assignee_id]
                sp_chart_user["user_id"] = task.assignee_id
                sp_chart_user["user_name"] = str(task.assignee)
                if not is_done:
                    sp_chart_user["in_progress"] += 1
                    sp_chart_user["overdue"] += 1 if is_overdue else 0
                chart_data["specialist"]["status_pie"][LeadSubfunnel.normalize_status(task.status)] += 1

        rows = sorted(
            bucket.values(),
            key=lambda x: (
                x["role"] or "",
                -(x["tasks_overdue"] + x["overdue_stage"] + x["overdue_checklist"]),
                -(x["tasks_in_progress"] + x["active_leads"] + x["pending_checklist"]),
                x["user_name"] or "",
            ),
        )
        totals = {
            "active_leads": sum(r["active_leads"] for r in rows),
            "pending_checklist": sum(r["pending_checklist"] for r in rows),
            "overdue_stage": sum(r["overdue_stage"] for r in rows),
            "overdue_checklist": sum(r["overdue_checklist"] for r in rows),
            "tasks_in_progress": sum(r["tasks_in_progress"] for r in rows),
            "tasks_overdue": sum(r["tasks_overdue"] for r in rows),
        }

        managers = []
        for item in managers_map.values():
            campaigns = []
            for campaign_item in item["campaigns"].values():
                campaign_item["leads"] = sorted(
                    campaign_item["leads"],
                    key=lambda lead_row: (
                        0 if lead_row["stage_overdue"] else 1,
                        lead_row["stage_deadline"] or "9999-99-99",
                        lead_row["organization_name"],
                    ),
                )
                campaigns.append(campaign_item)
            campaigns.sort(key=lambda c: c["campaign_name"] or "")
            managers.append(
                {
                    "user_id": item["user_id"],
                    "user_name": item["user_name"],
                    "campaigns": campaigns,
                }
            )
        managers.sort(key=lambda x: x["user_name"] or "")

        specialists = []
        for item in specialists_map.values():
            campaigns = []
            overdue_total = 0
            for campaign_item in item["campaigns"].values():
                templates = sorted(
                    campaign_item["templates"].values(),
                    key=lambda tpl: tpl["template_name"] or "",
                )
                campaign_item["templates"] = templates
                overdue_total += campaign_item["stats"]["overdue"]
                campaigns.append(campaign_item)
            campaigns.sort(key=lambda c: c["campaign_name"] or "")
            specialists.append(
                {
                    "user_id": item["user_id"],
                    "user_name": item["user_name"],
                    "campaigns": campaigns,
                    "overdue_total": overdue_total,
                }
            )
        specialists.sort(key=lambda x: x["user_name"] or "")

        active_chart_scope = "specialist" if role == "specialist" else "manager"
        charts_source = chart_data[active_chart_scope]
        charts = {
            "scope": active_chart_scope,
            "by_campaign": sorted(charts_source["by_campaign"].values(), key=lambda x: x["campaign_name"] or ""),
            "by_user": sorted(charts_source["by_user"].values(), key=lambda x: x["user_name"] or ""),
            "by_day": sorted(activity_by_period.values(), key=lambda x: x["date"] or ""),
            "status_pie": [{"status": status_key, "count": count} for status_key, count in charts_source["status_pie"].items()],
        }

        return Response(
            {
                "rows": rows,
                "totals": totals,
                "managers": managers,
                "specialists": specialists,
                "charts": charts,
                "meta": {
                    "role": role,
                    "campaign": campaign_id_int,
                    "funnel": funnel_id_int,
                    "user": user_id_int,
                    "date_from": date_from.isoformat() if date_from else None,
                    "date_to": date_to.isoformat() if date_to else None,
                    "period_mode": "activity",
                },
            }
        )

    @action(detail=False, methods=["post"], url_path="bulk-update")
    def bulk_update(self, request):
        id_list, err = _parse_bulk_ids(request.data.get("ids"))
        if err:
            return err

        board_column = request.data.get("board_column")
        status_val = request.data.get("status")
        if board_column is None and status_val is None:
            return Response(
                {"detail": "Укажите board_column или status."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        valid_statuses = {choice[0] for choice in Campaign.Status.choices}
        valid_columns = set(valid_statuses) | {Campaign.OperationalStage.ORGANIZATION_LIST}

        update_fields_base = ["updated_at"]
        rows = list(self.get_queryset().filter(id__in=id_list))
        updated = 0
        skipped = []
        from .collect_tasks import activate_collect_campaign_workflow, deactivate_collect_campaign_workflow

        for row in rows:
            fields = list(update_fields_base)
            if board_column is not None:
                col = str(board_column)
                if col not in valid_columns:
                    skipped.append({"id": row.id, "reason": "Неизвестная стадия доски."})
                    continue
                if col == Campaign.OperationalStage.ORGANIZATION_LIST:
                    row.status = Campaign.Status.ACTIVE
                    fields.append("status")
                else:
                    row.status = col
                    fields.append("status")
            elif status_val is not None:
                st = str(status_val)
                if st not in valid_statuses:
                    skipped.append({"id": row.id, "reason": "Неизвестный статус."})
                    continue
                row.status = st
                fields.append("status")
            row.save(update_fields=list(dict.fromkeys(fields)))
            if board_column is not None:
                if str(board_column) == Campaign.OperationalStage.ORGANIZATION_LIST:
                    activate_collect_campaign_workflow(row)
                else:
                    deactivate_collect_campaign_workflow(row)
            updated += 1

        return Response({"updated": updated, "skipped": skipped, "requested": len(id_list)})

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        id_list, err = _parse_bulk_ids(request.data.get("ids"))
        if err:
            return err
        deleted, _ = Campaign.objects.filter(id__in=id_list).delete()
        return Response({"deleted": deleted, "requested": len(id_list)})


class CampaignQueueViewSet(viewsets.ModelViewSet):
    serializer_class = CampaignQueueSerializer
    filterset_fields = ["campaign"]

    def get_queryset(self):
        return CampaignQueue.objects.all()

    @action(detail=True, methods=["get", "post"], url_path="stage-deadlines")
    def stage_deadlines(self, request, pk=None):
        queue = self.get_object()
        if request.method == "GET":
            deadlines = queue.stage_deadlines.select_related("funnel_stage")
            serializer = QueueStageDeadlineSerializer(deadlines, many=True)
            return Response(serializer.data)

        serializer = QueueStageDeadlineSerializer(
            data={**request.data, "queue": queue.pk}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CampaignOrganizationViewSet(viewsets.ModelViewSet):
    serializer_class = CampaignOrganizationSerializer
    filterset_fields = ["campaign", "status", "manager"]

    def get_queryset(self):
        return CampaignOrganization.objects.select_related(
            "organization__region", "manager"
        )


class LeadViewSet(viewsets.ModelViewSet):
    filterset_fields = ["campaign", "funnel", "queue", "manager", "current_stage"]
    search_fields = ["organization__name"]

    def get_queryset(self):
        qs = Lead.objects.select_related(
            "organization__region", "region", "funnel", "current_stage",
            "queue", "manager", "primary_contact",
        ).prefetch_related(
            "tags",
            "organization__tags",
            "checklist_values__checklist_item",
            "checklist_values__attachments",
            "interactions",
            "subfunnels__campaign_subfunnel__template",
            "subfunnels__current_template_stage",
            "subfunnels__checklist_values",
        )
        tag_ids = self.request.query_params.get("tags")
        if tag_ids:
            ids = [int(x) for x in tag_ids.split(",") if x.strip().isdigit()]
            if ids:
                qs = qs.filter(tags__id__in=ids).distinct()
        return qs

    def get_serializer_class(self):
        if self.action == "list":
            return LeadListSerializer
        return LeadDetailSerializer

    def perform_update(self, serializer):
        super().perform_update(serializer)
        lead = serializer.instance
        if lead.current_stage_id:
            if (
                not lead.primary_contact_specialist_id
                and lead.current_stage.primary_contact_specialist_id
            ):
                lead.primary_contact_specialist_id = lead.current_stage.primary_contact_specialist_id
                lead.save(update_fields=["primary_contact_specialist", "updated_at"])
            self._ensure_checklist_values(lead, lead.current_stage)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.current_stage:
            self._ensure_checklist_values(instance, instance.current_stage)
            instance = self.get_queryset().get(pk=instance.pk)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=True, methods=["get", "post"], url_path="checklist")
    def checklist(self, request, pk=None):
        lead = self.get_object()
        if request.method == "GET":
            values = lead.checklist_values.select_related(
                "checklist_item"
            ).prefetch_related("attachments")
            serializer = LeadChecklistValueSerializer(
                values, many=True, context={"request": request}
            )
            return Response(serializer.data)

        data = {**request.data, "lead": lead.pk}
        if request.data.get("is_completed"):
            data["completed_at"] = timezone.now().isoformat()
            data["completed_by"] = request.user.pk
        serializer = LeadChecklistValueSerializer(data=data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"], url_path="subfunnels")
    def subfunnels(self, request, pk=None):
        lead = self.get_object()
        self._refresh_subfunnel_availability(lead)
        rows = lead.subfunnels.select_related(
            "campaign_subfunnel__template",
            "campaign_subfunnel__role",
            "assignee",
        ).prefetch_related("checklist_values__template_item", "checklist_values__assignee")
        return Response(LeadSubfunnelSerializer(rows, many=True).data)

    @action(detail=True, methods=["post"], url_path="checklist/(?P<value_id>[^/.]+)/toggle")
    def toggle_checklist(self, request, pk=None, value_id=None):
        lead = self.get_object()
        try:
            value = lead.checklist_values.select_related("checklist_item").get(pk=value_id)
        except LeadChecklistValue.DoesNotExist:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)

        value.is_completed = not value.is_completed
        if value.is_completed:
            value.completed_at = timezone.now()
            value.completed_by = request.user
            if not value.primary_contact_specialist_id:
                value.primary_contact_specialist = request.user
        else:
            value.completed_at = None
            value.completed_by = None
        value.save()
        _sync_primary_contact_status_from_checklist(value)
        item_text = value.checklist_item.text
        if value.is_completed:
            summary = f"Отмечен пункт «{item_text}»"
        else:
            summary = f"Снята отметка с пункта «{item_text}»"
        _log_lead_activity(lead, request.user, LeadActivityLog.EventType.CHECKLIST, summary)
        return Response(
            LeadChecklistValueSerializer(value, context={"request": request}).data
        )

    @action(detail=True, methods=["patch"], url_path="checklist/(?P<value_id>[^/.]+)/update")
    def update_checklist_value(self, request, pk=None, value_id=None):
        lead = self.get_object()
        try:
            value = lead.checklist_values.select_related("checklist_item").get(pk=value_id)
        except LeadChecklistValue.DoesNotExist:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)

        files_list = list(request.FILES.getlist("files"))
        if not files_list:
            files_list = list(request.FILES.getlist("file"))
        if not files_list:
            single = request.FILES.get("file") or request.FILES.get("file_value")
            if single:
                files_list = [single]
        files_added = False
        for f in files_list:
            order = value.attachments.count()
            LeadChecklistAttachment.objects.create(
                checklist_value=value, file=f, order=order
            )
            files_added = True

        updatable = [
            "text_value", "select_value",
            "contact_name", "contact_position", "contact_phone",
            "contact_email", "contact_messenger",
        ]
        field_labels = {
            "text_value": "текст",
            "select_value": "выбор",
            "contact_name": "ФИО",
            "contact_position": "должность",
            "contact_phone": "телефон",
            "contact_email": "email",
            "contact_messenger": "мессенджер",
        }
        old_snapshot = {f: getattr(value, f) for f in updatable}
        old_contact_id = value.contact_id
        old_specialist_id = value.primary_contact_specialist_id
        for field in updatable:
            if field in request.data:
                setattr(value, field, request.data[field])
        if "contact" in request.data:
            cid = request.data["contact"]
            value.contact_id = int(cid) if cid is not None and cid != "" else None
        if "primary_contact_specialist" in request.data:
            sid = request.data["primary_contact_specialist"]
            value.primary_contact_specialist_id = (
                int(sid) if sid is not None and sid != "" else None
            )
        value.save()
        _sync_primary_contact_status_from_checklist(value)

        changed_labels = []
        if files_added:
            changed_labels.append("файл")
        for field in updatable:
            if field in request.data and old_snapshot[field] != getattr(value, field):
                changed_labels.append(field_labels.get(field, field))
        if "contact" in request.data:
            new_cid = value.contact_id
            if old_contact_id != new_cid:
                changed_labels.append("контакт из справочника")
        if "primary_contact_specialist" in request.data:
            if old_specialist_id != value.primary_contact_specialist_id:
                changed_labels.append("ответственный специалист")
        if changed_labels:
            item_text = value.checklist_item.text
            parts = ", ".join(sorted(set(changed_labels)))
            _log_lead_activity(
                lead,
                request.user,
                LeadActivityLog.EventType.CHECKLIST,
                f"«{item_text}»: {parts}",
            )
        return Response(
            LeadChecklistValueSerializer(value, context={"request": request}).data
        )

    @action(
        detail=True,
        methods=["delete"],
        url_path=r"checklist/(?P<value_id>[^/.]+)/attachments/(?P<attachment_id>[^/.]+)",
    )
    def delete_checklist_attachment(self, request, pk=None, value_id=None, attachment_id=None):
        lead = self.get_object()
        try:
            value = lead.checklist_values.select_related("checklist_item").get(pk=value_id)
        except LeadChecklistValue.DoesNotExist:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)

        try:
            att = value.attachments.get(pk=attachment_id)
        except LeadChecklistAttachment.DoesNotExist:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)

        item_text = value.checklist_item.text
        att.file.delete(save=False)
        att.delete()
        _log_lead_activity(
            lead,
            request.user,
            LeadActivityLog.EventType.CHECKLIST,
            f"«{item_text}»: удалён файл вложения",
        )
        value.refresh_from_db()
        return Response(
            LeadChecklistValueSerializer(value, context={"request": request}).data
        )

    @action(detail=True, methods=["get", "post"], url_path="interactions")
    def interactions(self, request, pk=None):
        lead = self.get_object()
        if request.method == "GET":
            interactions = lead.interactions.select_related("created_by")
            serializer = LeadInteractionSerializer(interactions, many=True)
            return Response(serializer.data)

        serializer = LeadInteractionSerializer(
            data={**request.data, "lead": lead.pk}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(created_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"], url_path="timeline")
    def timeline(self, request, pk=None):
        """Единая лента: взаимодействия + смена стадий + чек-лист (по дате)."""
        lead = self.get_object()
        items = []
        for i in lead.interactions.select_related("created_by").order_by("-date", "-created_at"):
            dt = i.date or i.created_at
            at_str = dt.isoformat() if dt and hasattr(dt, "isoformat") else (str(dt) if dt else "")
            items.append({
                "kind": "interaction",
                "id": i.pk,
                "at": at_str,
                "data": LeadInteractionSerializer(i).data,
            })
        try:
            activity_qs = list(
                lead.activity_logs.select_related("created_by").all()
            )
        except (OperationalError, ProgrammingError):
            activity_qs = []
        for log in activity_qs:
            user = log.created_by
            items.append({
                "kind": log.event_type,
                "id": log.pk,
                "at": log.created_at.isoformat(),
                "summary": log.summary,
                "created_by_name": str(user) if user else None,
            })
        _append_synthetic_checklist_from_values(lead, items)
        items.sort(key=lambda x: x["at"] or "", reverse=True)

        kinds_raw = request.query_params.get("kind", "").strip()
        kinds = None
        if kinds_raw:
            kinds = {x.strip() for x in kinds_raw.split(",") if x.strip()}
        contact_param = request.query_params.get("contact")
        contact_id = None
        if contact_param not in (None, ""):
            try:
                contact_id = int(contact_param)
            except (TypeError, ValueError):
                contact_id = None
        items = _filter_timeline_items(items, kinds, contact_id)
        return Response(items)

    @action(detail=True, methods=["post"], url_path="advance-stage")
    def advance_stage(self, request, pk=None):
        lead = self.get_object()
        old_stage = lead.current_stage
        normal_stages = list(
            lead.funnel.stages.filter(is_rejection=False).order_by("order")
        )
        if not normal_stages:
            return Response({"detail": "No stages in funnel"}, status=status.HTTP_400_BAD_REQUEST)

        if lead.current_stage is None or lead.current_stage.is_rejection:
            lead.current_stage = normal_stages[0]
            if lead.primary_contact_status == Lead.PrimaryContactStatus.REJECTED:
                lead.primary_contact_status = Lead.PrimaryContactStatus.NEW
        else:
            current_idx = next(
                (i for i, s in enumerate(normal_stages) if s.id == lead.current_stage_id),
                -1,
            )
            if current_idx < len(normal_stages) - 1:
                lead.current_stage = normal_stages[current_idx + 1]
            else:
                return Response(
                    {"detail": "Already at last stage"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        lead.save(update_fields=["current_stage", "primary_contact_status", "updated_at"])
        if lead.current_stage.primary_contact_specialist_id:
            lead.primary_contact_specialist_id = lead.current_stage.primary_contact_specialist_id
            lead.save(update_fields=["primary_contact_specialist", "updated_at"])
        self._ensure_checklist_values(lead, lead.current_stage)
        self._refresh_subfunnel_availability(lead)
        old_name = old_stage.name if old_stage else "—"
        new_name = lead.current_stage.name if lead.current_stage else "—"
        _log_lead_activity(
            lead,
            request.user,
            LeadActivityLog.EventType.STAGE,
            f"{old_name} → {new_name}",
        )
        return Response(LeadDetailSerializer(lead).data)

    @action(detail=True, methods=["post"], url_path="retreat-stage")
    def retreat_stage(self, request, pk=None):
        lead = self.get_object()
        normal_stages = list(
            lead.funnel.stages.filter(is_rejection=False).order_by("order")
        )
        if not normal_stages:
            return Response({"detail": "No stages in funnel"}, status=status.HTTP_400_BAD_REQUEST)

        if lead.current_stage is None or lead.current_stage.is_rejection:
            return Response(
                {"detail": "No current stage to retreat from"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_idx = next(
            (i for i, s in enumerate(normal_stages) if s.id == lead.current_stage_id),
            -1,
        )
        if current_idx <= 0:
            return Response(
                {"detail": "Already at first stage"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        old_stage = lead.current_stage
        lead.current_stage = normal_stages[current_idx - 1]
        lead.save(update_fields=["current_stage", "updated_at"])
        self._refresh_subfunnel_availability(lead)
        old_name = old_stage.name if old_stage else "—"
        new_name = lead.current_stage.name if lead.current_stage else "—"
        _log_lead_activity(
            lead,
            request.user,
            LeadActivityLog.EventType.STAGE,
            f"{old_name} → {new_name}",
        )
        return Response(LeadDetailSerializer(lead).data)

    @action(detail=False, methods=["post"], url_path="bulk-update")
    def bulk_update(self, request):
        id_list, err = _parse_bulk_ids(request.data.get("ids"))
        if err:
            return err
        if "current_stage" not in request.data:
            return Response(
                {"detail": "Укажите current_stage."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        stage_raw = request.data.get("current_stage")
        target_stage = None
        if stage_raw not in (None, ""):
            if not str(stage_raw).isdigit():
                return Response({"detail": "Некорректный current_stage."}, status=status.HTTP_400_BAD_REQUEST)
            try:
                target_stage = FunnelStage.objects.get(id=int(stage_raw))
            except FunnelStage.DoesNotExist:
                return Response({"detail": "Стадия не найдена."}, status=status.HTTP_404_NOT_FOUND)

        rows = list(self.get_queryset().filter(id__in=id_list))
        updated = 0
        skipped = []

        for row in rows:
            if target_stage is not None and row.funnel_id != target_stage.funnel_id:
                skipped.append({"id": row.id, "reason": "Стадия не принадлежит воронке лида."})
                continue
            row.current_stage = target_stage
            row.save(update_fields=["current_stage", "updated_at"])
            if target_stage is not None:
                if (
                    not row.primary_contact_specialist_id
                    and target_stage.primary_contact_specialist_id
                ):
                    row.primary_contact_specialist_id = target_stage.primary_contact_specialist_id
                    row.save(update_fields=["primary_contact_specialist", "updated_at"])
                self._ensure_checklist_values(row, target_stage)
                self._refresh_subfunnel_availability(row)
            updated += 1

        return Response({"updated": updated, "skipped": skipped, "requested": len(id_list)})

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        id_list, err = _parse_bulk_ids(request.data.get("ids"))
        if err:
            return err
        deleted, _ = Lead.objects.filter(id__in=id_list).delete()
        return Response({"deleted": deleted, "requested": len(id_list)})

    @action(detail=True, methods=["post"], url_path="reject")
    def reject(self, request, pk=None):
        lead = self.get_object()
        rejection_stage = lead.funnel.stages.filter(is_rejection=True).first()
        if not rejection_stage:
            return Response(
                {"detail": "No rejection stage in funnel"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        old_stage = lead.current_stage
        lead.current_stage = rejection_stage
        lead.primary_contact_status = Lead.PrimaryContactStatus.REJECTED
        lead.save(update_fields=["current_stage", "primary_contact_status", "updated_at"])
        self._ensure_checklist_values(lead, lead.current_stage)
        self._refresh_subfunnel_availability(lead)
        old_name = old_stage.name if old_stage else "—"
        _log_lead_activity(
            lead,
            request.user,
            LeadActivityLog.EventType.STAGE,
            f"{old_name} → Отказ",
        )
        return Response(LeadDetailSerializer(lead).data)

    @staticmethod
    def _ensure_checklist_values(lead, stage):
        """Create LeadChecklistValue for each StageChecklistItem that doesn't have one yet."""
        if not stage:
            return
        for item in stage.checklist_items.all():
            LeadChecklistValue.objects.get_or_create(
                lead=lead,
                checklist_item=item,
                defaults={
                    "primary_contact_specialist_id": (
                        item.primary_contact_specialist_id
                        or lead.primary_contact_specialist_id
                    )
                },
            )

    @staticmethod
    def _refresh_subfunnel_availability(lead):
        if not lead.current_stage_id:
            return
        rows = lead.subfunnels.select_related(
            "campaign_subfunnel__binding__from_stage",
            "campaign_subfunnel__binding__to_stage",
        )
        for row in rows:
            binding = row.campaign_subfunnel.binding
            new_availability = True
            if binding and binding.binding_type == "stage_range_checklist":
                if not binding.from_stage_id or not binding.to_stage_id:
                    new_availability = False
                else:
                    stage_order = lead.current_stage.order
                    new_availability = (
                        binding.from_stage.order <= stage_order <= binding.to_stage.order
                    )
            if row.is_available != new_availability:
                row.is_available = new_availability
                row.save(update_fields=["is_available", "updated_at"])


class LeadSubfunnelViewSet(viewsets.ModelViewSet):
    serializer_class = LeadSubfunnelSerializer
    filterset_fields = ["lead", "campaign_subfunnel", "status", "current_template_stage", "assignee", "is_available"]

    def get_queryset(self):
        return LeadSubfunnel.objects.select_related(
            "lead__campaign",
            "lead__organization",
            "campaign_region__campaign",
            "campaign_region__region",
            "campaign_subfunnel__template",
            "campaign_subfunnel__binding",
            "campaign_subfunnel__role",
            "current_template_stage",
            "assignee",
        ).prefetch_related("checklist_values__template_item")

    def perform_update(self, serializer):
        prev_stage = serializer.instance.current_template_stage
        prev_status = serializer.instance.status
        explicit_status = "status" in serializer.validated_data
        obj = serializer.save()
        obj.status = LeadSubfunnel.normalize_status(obj.status)
        if not explicit_status and obj.current_template_stage_id:
            mapped = LeadSubfunnel.status_from_stage(obj.current_template_stage)
            if mapped != obj.status:
                obj.status = mapped
        update_fields = []
        if obj.status != prev_status or explicit_status:
            update_fields.extend(["status", "updated_at"])
            if obj.status == LeadSubfunnel.Status.DONE:
                if not obj.completed_at:
                    obj.completed_at = timezone.now()
                    update_fields.append("completed_at")
            elif prev_status == LeadSubfunnel.Status.DONE:
                obj.completed_at = None
                update_fields.append("completed_at")
        if update_fields:
            obj.save(update_fields=list(dict.fromkeys(update_fields)))
        self._advance_lead_stage_from_task_transition(
            obj,
            request_user=self.request.user,
            from_stage=prev_stage,
            to_stage=obj.current_template_stage,
        )

    @staticmethod
    def _advance_lead_stage_from_task_transition(subfunnel, request_user=None, from_stage=None, to_stage=None):
        if not subfunnel.lead_id:
            return
        if not from_stage or not to_stage:
            return
        if not to_stage.is_terminal:
            return
        if to_stage.order <= from_stage.order:
            return
        campaign_subfunnel = subfunnel.campaign_subfunnel
        if not campaign_subfunnel:
            return
        binding = campaign_subfunnel.binding
        if not binding or binding.binding_type != SubfunnelTemplateBinding.BindingType.STAGE:
            return
        if not getattr(binding, "advance_lead_on_task_stage_forward", False):
            return
        lead = subfunnel.lead
        if not lead or not lead.funnel_id or not lead.current_stage_id:
            return
        if lead.current_stage.is_rejection:
            return

        normal_stages = list(lead.funnel.stages.filter(is_rejection=False).order_by("order"))
        if not normal_stages:
            return
        current_idx = next((i for i, s in enumerate(normal_stages) if s.id == lead.current_stage_id), -1)
        if current_idx < 0 or current_idx >= len(normal_stages) - 1:
            return

        old_stage = lead.current_stage
        lead.current_stage = normal_stages[current_idx + 1]
        update_fields = ["current_stage", "updated_at"]
        if lead.current_stage.primary_contact_specialist_id:
            lead.primary_contact_specialist_id = lead.current_stage.primary_contact_specialist_id
            update_fields.append("primary_contact_specialist")
        lead.save(update_fields=update_fields)
        LeadViewSet._ensure_checklist_values(lead, lead.current_stage)
        LeadViewSet._refresh_subfunnel_availability(lead)
        old_name = old_stage.name if old_stage else "—"
        new_name = lead.current_stage.name if lead.current_stage else "—"
        _log_lead_activity(
            lead,
            request_user,
            LeadActivityLog.EventType.STAGE,
            f"{old_name} → {new_name} (авто из задачи)",
        )

    @action(detail=True, methods=["post"], url_path="advance-task-stage")
    def advance_task_stage(self, request, pk=None):
        subfunnel = self.get_object()
        stages = list(
            TaskTemplateStage.objects.filter(
                template_id=subfunnel.campaign_subfunnel.template_id,
                is_active=True,
            ).order_by("order", "id")
        )
        if not stages:
            return Response({"detail": "Шаблон задачи не содержит этапов."}, status=status.HTTP_400_BAD_REQUEST)
        current_id = subfunnel.current_template_stage_id or stages[0].id
        idx = next((i for i, s in enumerate(stages) if s.id == current_id), 0)
        if idx >= len(stages) - 1:
            return Response({"detail": "Задача уже на последнем этапе."}, status=status.HTTP_400_BAD_REQUEST)
        prev_stage = stages[idx]
        next_stage = stages[idx + 1]
        subfunnel.current_template_stage = next_stage
        subfunnel.status = LeadSubfunnel.status_from_stage(next_stage)
        if subfunnel.status == LeadSubfunnel.Status.DONE and not subfunnel.completed_at:
            subfunnel.completed_at = timezone.now()
        subfunnel.save(update_fields=["current_template_stage", "status", "completed_at", "updated_at"])
        self._advance_lead_stage_from_task_transition(
            subfunnel,
            request_user=request.user,
            from_stage=prev_stage,
            to_stage=next_stage,
        )
        return Response(LeadSubfunnelSerializer(subfunnel).data)

    @action(detail=True, methods=["post"], url_path="retreat-task-stage")
    def retreat_task_stage(self, request, pk=None):
        subfunnel = self.get_object()
        stages = list(
            TaskTemplateStage.objects.filter(
                template_id=subfunnel.campaign_subfunnel.template_id,
                is_active=True,
            ).order_by("order", "id")
        )
        if not stages:
            return Response({"detail": "Шаблон задачи не содержит этапов."}, status=status.HTTP_400_BAD_REQUEST)
        current_id = subfunnel.current_template_stage_id or stages[0].id
        idx = next((i for i, s in enumerate(stages) if s.id == current_id), 0)
        if idx <= 0:
            return Response({"detail": "Задача уже на первом этапе."}, status=status.HTTP_400_BAD_REQUEST)
        prev_stage = stages[idx - 1]
        subfunnel.current_template_stage = prev_stage
        subfunnel.status = LeadSubfunnel.status_from_stage(prev_stage)
        if subfunnel.status != LeadSubfunnel.Status.DONE and subfunnel.completed_at:
            subfunnel.completed_at = None
        subfunnel.save(update_fields=["current_template_stage", "status", "completed_at", "updated_at"])
        return Response(LeadSubfunnelSerializer(subfunnel).data)

    @action(detail=True, methods=["post"], url_path="set-task-stage")
    def set_task_stage(self, request, pk=None):
        subfunnel = self.get_object()
        prev_stage = subfunnel.current_template_stage
        stage_id = request.data.get("stage_id")
        if not stage_id or not str(stage_id).isdigit():
            return Response({"detail": "Нужно передать корректный stage_id."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            target_stage = TaskTemplateStage.objects.get(id=int(stage_id), is_active=True)
        except TaskTemplateStage.DoesNotExist:
            return Response({"detail": "Этап задачи не найден."}, status=status.HTTP_400_BAD_REQUEST)

        if target_stage.template_id != subfunnel.campaign_subfunnel.template_id:
            return Response(
                {"detail": "Этап не принадлежит шаблону этой задачи."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        subfunnel.current_template_stage = target_stage
        subfunnel.status = LeadSubfunnel.status_from_stage(target_stage)
        if subfunnel.status == LeadSubfunnel.Status.DONE:
            subfunnel.completed_at = subfunnel.completed_at or timezone.now()
        else:
            subfunnel.completed_at = None
        subfunnel.save(update_fields=["current_template_stage", "status", "completed_at", "updated_at"])
        self._advance_lead_stage_from_task_transition(
            subfunnel,
            request_user=request.user,
            from_stage=prev_stage,
            to_stage=target_stage,
        )
        return Response(LeadSubfunnelSerializer(subfunnel).data)

    @action(detail=True, methods=["get"], url_path="region-capture")
    def region_capture(self, request, pk=None):
        subfunnel = self.get_object()
        if not subfunnel.campaign_region_id:
            return Response(
                {"detail": "Доступно только для региональных задач."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        campaign_region = subfunnel.campaign_region
        leads_qs = (
            Lead.objects.filter(
                campaign_id=campaign_region.campaign_id,
                region_id=campaign_region.region_id,
            )
            .select_related("organization", "primary_contact")
            .order_by("-id")
        )

        organizations = []
        for lead in leads_qs:
            primary_contact = lead.primary_contact
            organizations.append(
                {
                    "lead_id": lead.id,
                    "organization_id": lead.organization_id,
                    "organization_name": lead.organization.name if lead.organization else None,
                    "primary_contact": str(primary_contact) if primary_contact else None,
                    "created_at": lead.created_at.isoformat() if lead.created_at else None,
                }
            )

        return Response(
            {
                "campaign_region_id": campaign_region.id,
                "region_id": campaign_region.region_id,
                "region_name": campaign_region.region.name if campaign_region.region else None,
                "demand_quota": campaign_region.demand_quota,
                "leads_count": leads_qs.count(),
                "organizations": organizations,
            }
        )

    @action(detail=False, methods=["post"], url_path="bulk-update")
    def bulk_update(self, request):
        ids = request.data.get("ids")
        if not isinstance(ids, list) or not ids:
            return Response({"detail": "Передайте непустой список ids."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            id_list = [int(x) for x in ids]
        except (TypeError, ValueError):
            return Response({"detail": "ids должны быть числами."}, status=status.HTTP_400_BAD_REQUEST)

        assignee_provided = "assignee" in request.data
        due_at_provided = "due_at" in request.data
        clear_due_at = bool(request.data.get("clear_due_at"))
        stage_id_provided = "stage_id" in request.data
        status_provided = "status" in request.data

        if not any([assignee_provided, due_at_provided, clear_due_at, stage_id_provided, status_provided]):
            return Response(
                {"detail": "Укажите хотя бы одно поле: assignee, due_at, clear_due_at, stage_id или status."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        target_stage = None
        if stage_id_provided:
            stage_raw = request.data.get("stage_id")
            if stage_raw is not None and str(stage_raw).strip() != "":
                if not str(stage_raw).isdigit():
                    return Response({"detail": "Некорректный stage_id."}, status=status.HTTP_400_BAD_REQUEST)
                try:
                    target_stage = TaskTemplateStage.objects.get(id=int(stage_raw), is_active=True)
                except TaskTemplateStage.DoesNotExist:
                    return Response({"detail": "Этап задачи не найден."}, status=status.HTTP_400_BAD_REQUEST)

        parsed_due_at = None
        if due_at_provided and not clear_due_at:
            raw_due = request.data.get("due_at")
            if raw_due:
                parsed_due_at = parse_datetime(str(raw_due))
                if parsed_due_at is None:
                    return Response({"detail": "Некорректный формат due_at."}, status=status.HTTP_400_BAD_REQUEST)
                if timezone.is_naive(parsed_due_at):
                    parsed_due_at = timezone.make_aware(parsed_due_at, timezone.get_current_timezone())

        bulk_status = None
        if status_provided:
            raw_status = request.data.get("status")
            if raw_status not in TASK_WORKFLOW_STATUS_VALUES:
                return Response({"detail": "Некорректный status."}, status=status.HTTP_400_BAD_REQUEST)
            bulk_status = LeadSubfunnel.normalize_status(raw_status)

        rows = list(self.get_queryset().filter(id__in=id_list))
        updated = 0
        skipped = []

        for row in rows:
            update_fields = ["updated_at"]
            prev_stage = row.current_template_stage

            if assignee_provided:
                assignee = request.data.get("assignee")
                row.assignee_id = int(assignee) if assignee not in (None, "") else None
                update_fields.append("assignee")

            if clear_due_at:
                row.due_at = None
                update_fields.append("due_at")
            elif due_at_provided:
                row.due_at = parsed_due_at
                update_fields.append("due_at")

            if status_provided:
                row.status = bulk_status
                update_fields.append("status")
                if row.status == LeadSubfunnel.Status.DONE:
                    row.completed_at = row.completed_at or timezone.now()
                    update_fields.append("completed_at")
                elif row.completed_at:
                    row.completed_at = None
                    update_fields.append("completed_at")

            if stage_id_provided:
                if target_stage is None:
                    row.current_template_stage = None
                    if not status_provided:
                        row.status = LeadSubfunnel.Status.BACKLOG
                        update_fields.append("status")
                    update_fields.append("current_template_stage")
                elif target_stage.template_id != row.campaign_subfunnel.template_id:
                    skipped.append({
                        "id": row.id,
                        "reason": "Этап не принадлежит шаблону задачи.",
                    })
                    continue
                else:
                    row.current_template_stage = target_stage
                    update_fields.append("current_template_stage")
                    if not status_provided:
                        row.status = LeadSubfunnel.status_from_stage(target_stage)
                        update_fields.append("status")
                    if row.status == LeadSubfunnel.Status.DONE:
                        row.completed_at = row.completed_at or timezone.now()
                        update_fields.append("completed_at")
                    elif row.completed_at:
                        row.completed_at = None
                        update_fields.append("completed_at")

            row.save(update_fields=list(dict.fromkeys(update_fields)))
            if stage_id_provided and target_stage is not None:
                self._advance_lead_stage_from_task_transition(
                    row,
                    request_user=request.user,
                    from_stage=prev_stage,
                    to_stage=target_stage,
                )
            updated += 1

        return Response({
            "updated": updated,
            "skipped": skipped,
            "requested": len(id_list),
        })

    @action(detail=False, methods=["post"], url_path="bulk-checklist")
    def bulk_checklist(self, request):
        id_list, err = _parse_bulk_ids(request.data.get("ids"))
        if err:
            return err

        template_item_raw = request.data.get("template_item_id")
        if not template_item_raw or not str(template_item_raw).isdigit():
            return Response(
                {"detail": "Укажите template_item_id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        is_completed_provided = "is_completed" in request.data
        text_value_provided = "text_value" in request.data
        if not is_completed_provided and not text_value_provided:
            return Response(
                {"detail": "Укажите is_completed и/или text_value."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            template_item = SubfunnelTemplateItem.objects.get(id=int(template_item_raw))
        except SubfunnelTemplateItem.DoesNotExist:
            return Response({"detail": "Пункт шаблона не найден."}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(
            self.get_queryset()
            .filter(id__in=id_list)
            .prefetch_related("checklist_values")
        )
        updated_tasks = 0
        updated_values = 0
        skipped = []

        for row in rows:
            if row.campaign_subfunnel.template_id != template_item.template_id:
                skipped.append({
                    "id": row.id,
                    "reason": "Пункт не принадлежит шаблону задачи.",
                })
                continue
            try:
                value = row.checklist_values.get(template_item_id=template_item.id)
            except LeadSubfunnelChecklistValue.DoesNotExist:
                skipped.append({
                    "id": row.id,
                    "reason": "Пункт чек-листа не найден для задачи.",
                })
                continue

            changed = False
            if is_completed_provided:
                new_completed = bool(request.data.get("is_completed"))
                if value.is_completed != new_completed:
                    value.is_completed = new_completed
                    if new_completed:
                        value.completed_at = timezone.now()
                        value.completed_by = request.user
                    else:
                        value.completed_at = None
                        value.completed_by = None
                    changed = True

            if text_value_provided:
                new_text = request.data.get("text_value") or ""
                if value.text_value != new_text:
                    value.text_value = new_text
                    changed = True

            if changed:
                value.save()
                updated_values += 1
                updated_tasks += 1

        return Response({
            "updated_tasks": updated_tasks,
            "updated_values": updated_values,
            "skipped": skipped,
            "requested": len(id_list),
        })

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        id_list, err = _parse_bulk_ids(request.data.get("ids"))
        if err:
            return err
        deleted, _ = LeadSubfunnel.objects.filter(id__in=id_list).delete()
        return Response({"deleted": deleted, "requested": len(id_list)})

    @action(detail=True, methods=["get", "patch"], url_path="checklist")
    def checklist(self, request, pk=None):
        subfunnel = self.get_object()
        if request.method == "GET":
            rows = subfunnel.checklist_values.select_related("template_item", "assignee", "completed_by")
            return Response(LeadSubfunnelChecklistValueSerializer(rows, many=True).data)
        updated = 0
        payload = request.data if isinstance(request.data, list) else [request.data]
        for row in payload:
            value_id = row.get("id")
            if not value_id:
                continue
            try:
                value = subfunnel.checklist_values.get(id=value_id)
            except LeadSubfunnelChecklistValue.DoesNotExist:
                continue
            if "is_completed" in row:
                value.is_completed = bool(row["is_completed"])
                if value.is_completed:
                    value.completed_at = timezone.now()
                    value.completed_by = request.user
                else:
                    value.completed_at = None
                    value.completed_by = None
            if "text_value" in row:
                value.text_value = row.get("text_value") or ""
            if "assignee" in row:
                assignee = row.get("assignee")
                value.assignee_id = assignee or None
            value.save()
            updated += 1
        rows = subfunnel.checklist_values.select_related("template_item", "assignee", "completed_by")
        return Response(
            {
                "updated": updated,
                "rows": LeadSubfunnelChecklistValueSerializer(rows, many=True).data,
            }
        )
