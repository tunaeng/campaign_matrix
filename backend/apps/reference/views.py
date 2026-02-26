import csv
import io
import json
from typing import List

from rest_framework import viewsets, generics, permissions, status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Q
from openpyxl import load_workbook

from .models import (
    FederalDistrict, Region, Profession, ProfessionDemandStatus,
    ProfessionApprovalStatus, Program, FederalOperator, Contract,
    ContractProgram, Quota, DemandImport, DemandImportSnapshot,
)
from .serializers import (
    FederalDistrictSerializer, RegionSerializer, ProfessionSerializer,
    ProfessionDemandStatusSerializer, ProgramSerializer,
    FederalOperatorSerializer, ContractSerializer,
    ContractProgramSerializer, QuotaSerializer,
)


class LargePagination(PageNumberPagination):
    """Allows clients to request up to 500 items per page (for dropdowns)."""
    page_size = 50
    max_page_size = 500
    page_size_query_param = "page_size"


class FederalDistrictViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = FederalDistrict.objects.prefetch_related("regions")
    serializer_class = FederalDistrictSerializer
    pagination_class = LargePagination

    @action(detail=True, methods=["get"])
    def regions(self, request, pk=None):
        district = self.get_object()
        regions = district.regions.all()
        return Response(RegionSerializer(regions, many=True).data)


class RegionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Region.objects.select_related("federal_district")
    serializer_class = RegionSerializer
    filterset_fields = ["federal_district"]
    search_fields = ["name"]
    pagination_class = LargePagination


class ProfessionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Profession.objects.all()
    serializer_class = ProfessionSerializer
    search_fields = ["name", "number"]
    pagination_class = LargePagination

    @action(detail=True, methods=["get"], url_path="demand-map")
    def demand_map(self, request, pk=None):
        profession = self.get_object()
        year = request.query_params.get("year", 2026)
        statuses = ProfessionDemandStatus.objects.filter(
            profession=profession, year=year
        ).select_related("region")
        return Response(
            ProfessionDemandStatusSerializer(statuses, many=True).data
        )


class ProgramViewSet(viewsets.ModelViewSet):
    queryset = Program.objects.select_related("profession").prefetch_related(
        "contract_entries__contract__federal_operator"
    )
    serializer_class = ProgramSerializer
    filterset_fields = ["profession", "is_active"]
    search_fields = ["name", "profession__name"]

    def get_queryset(self):
        qs = super().get_queryset()
        contract_status = self.request.query_params.get("contract_status")
        operator = self.request.query_params.get("operator")

        if contract_status:
            qs = qs.filter(contract_entries__status=contract_status)
        if operator:
            qs = qs.filter(
                contract_entries__contract__federal_operator_id=operator
            )

        demanded_in_region = self.request.query_params.get("demanded_in_region")
        if demanded_in_region:
            qs = qs.filter(
                profession__demand_statuses__region_id=demanded_in_region,
                profession__demand_statuses__is_demanded=True,
            )

        demanded_only = self.request.query_params.get("demanded_only")
        if demanded_only and demanded_only.lower() == "true":
            qs = qs.filter(
                profession__demand_statuses__is_demanded=True,
            )

        return qs.distinct()


class FederalOperatorViewSet(viewsets.ModelViewSet):
    queryset = FederalOperator.objects.all()
    serializer_class = FederalOperatorSerializer
    search_fields = ["name"]
    pagination_class = LargePagination


class ContractViewSet(viewsets.ModelViewSet):
    queryset = Contract.objects.select_related("federal_operator")
    serializer_class = ContractSerializer
    filterset_fields = ["federal_operator", "year", "status"]


class ContractProgramViewSet(viewsets.ModelViewSet):
    queryset = ContractProgram.objects.select_related("contract", "program")
    serializer_class = ContractProgramSerializer
    filterset_fields = ["contract", "program", "status"]


class QuotaViewSet(viewsets.ModelViewSet):
    queryset = Quota.objects.select_related(
        "federal_operator", "program", "region"
    )
    serializer_class = QuotaSerializer
    filterset_fields = ["federal_operator", "program", "region", "year"]


class DemandMatrixImportView(APIView):
    """
    CSV import for profession demand statuses.
    Upsert rules:
      - Profession: add if missing by number/name
      - Demand status: update_or_create by (profession, region, year)
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @staticmethod
    def _to_bool(value):
        if value is None:
            return False
        normalized = str(value).strip().lower()
        if normalized in {"1", "true", "yes", "y", "да", "д", "x", "+"}:
            return True
        if normalized in {"0", "0.0", "false", "no", "нет", "н"}:
            return False
        try:
            return float(value) == 1
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _norm_header(value):
        return str(value).strip().lower().replace(" ", "_")

    @staticmethod
    def _clean_cell(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize(value):
        if value is None:
            return ""
        return str(value).strip().lower()

    @staticmethod
    def _save_import_snapshot(federal_operator, year, user):
        demand_import = DemandImport.objects.create(
            federal_operator=federal_operator,
            year=year,
            imported_by=user if user and user.is_authenticated else None,
        )
        current = ProfessionDemandStatus.objects.filter(
            federal_operator=federal_operator, year=year,
        ).values_list("profession_id", "region_id", "is_demanded")
        snapshots = [
            DemandImportSnapshot(
                demand_import=demand_import,
                profession_id=prof_id,
                region_id=reg_id,
                is_demanded=demanded,
            )
            for prof_id, reg_id, demanded in current
        ]
        BATCH = 1000
        for i in range(0, len(snapshots), BATCH):
            DemandImportSnapshot.objects.bulk_create(snapshots[i:i + BATCH])
        demand_import.snapshot_count = len(snapshots)
        demand_import.save(update_fields=["snapshot_count"])
        return demand_import

    def _read_csv_rows(self, raw: bytes) -> List[List[str]]:
        decoded = None
        for enc in ("utf-8-sig", "cp1251", "utf-8"):
            try:
                decoded = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise ValueError("Cannot decode file. Use UTF-8 or CP1251 CSV.")
        sample = decoded[:2048]
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(io.StringIO(decoded), delimiter=delimiter)
        return [[self._clean_cell(c) for c in row] for row in reader]

    def _read_xlsx_rows(self, raw: bytes) -> List[List[str]]:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([self._clean_cell(c) for c in row])
        return rows

    def _get_or_create_profession(
        self,
        profession_number,
        profession_name,
        profession_by_number,
        profession_by_name,
    ):
        created_profession = False
        updated_profession = False
        profession = None
        if profession_number is not None:
            profession = profession_by_number.get(profession_number)
        if profession is None and profession_name:
            profession = profession_by_name.get(profession_name.lower())

        if profession is None:
            new_number = profession_number
            if new_number is None:
                max_number = (
                    Profession.objects.order_by("-number")
                    .values_list("number", flat=True)
                    .first()
                    or 0
                )
                new_number = max_number + 1
            profession = Profession.objects.create(
                number=new_number,
                name=profession_name or f"Профессия {new_number}",
            )
            created_profession = True
            profession_by_number[profession.number] = profession
            profession_by_name[profession.name.strip().lower()] = profession
        elif profession_name and profession.name != profession_name:
            profession.name = profession_name
            profession.save(update_fields=["name"])
            updated_profession = True
            profession_by_name[profession.name.strip().lower()] = profession

        return profession, created_profession, updated_profession

    def _import_wide_matrix(
        self,
        rows,
        default_year,
        federal_operator,
        region_by_name,
        profession_by_number,
        profession_by_name,
    ):
        if len(rows) < 2:
            raise ValueError("Matrix file is too short.")

        header = rows[0]
        # Сначала пробуем формат с одной колонкой профессий: [название, регион1, регион2, ...]
        region_columns_1 = []
        for idx in range(1, len(header)):
            name = self._clean_cell(header[idx]).lower()
            if not name:
                continue
            region = region_by_name.get(name)
            if region:
                region_columns_1.append((idx, region))
        # Формат с двумя колонками: [№, Профессия, регион1, ...]
        region_columns_2 = []
        for idx in range(2, len(header)):
            name = self._clean_cell(header[idx]).lower()
            if not name:
                continue
            region = region_by_name.get(name)
            if region:
                region_columns_2.append((idx, region))

        # Одна колонка профессий только если в ячейке 1 действительно регион (а не "Профессия")
        if len(region_columns_1) >= 5 and len(header) > 1 and self._clean_cell(header[1]).lower() in region_by_name:
            region_columns = region_columns_1
            single_profession_column = True
            data_start = 1
        elif len(region_columns_2) >= 5:
            region_columns = region_columns_2
            single_profession_column = False
            data_start = 2
        else:
            raise ValueError("No region columns found in matrix header.")

        created_professions = 0
        updated_professions = 0
        skipped_rows = 0
        errors = []

        existing = {
            (s.profession_id, s.region_id): s
            for s in ProfessionDemandStatus.objects.filter(
                federal_operator=federal_operator, year=default_year,
            ).only("id", "profession_id", "region_id", "is_demanded")
        }

        to_create = []
        to_update = []

        for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
            if not row or len(row) < 1:
                skipped_rows += 1
                continue

            if single_profession_column:
                raw_number = ""
                profession_name = self._clean_cell(row[0]) if len(row) > 0 else ""
            else:
                raw_number = self._clean_cell(row[0]) if len(row) > 0 else ""
                profession_name = self._clean_cell(row[1]) if len(row) > 1 else ""

            if not profession_name or profession_name.lower() == "итого":
                skipped_rows += 1
                continue

            profession_number = None
            if raw_number:
                try:
                    profession_number = int(float(raw_number))
                except Exception:
                    profession_number = None

            try:
                profession, was_created, was_updated = self._get_or_create_profession(
                    profession_number,
                    profession_name,
                    profession_by_number,
                    profession_by_name,
                )
                if was_created:
                    created_professions += 1
                if was_updated:
                    updated_professions += 1

                for col_idx, region in region_columns:
                    cell_val = row[col_idx] if col_idx < len(row) else ""
                    is_demanded = self._to_bool(cell_val)
                    key = (profession.id, region.id)
                    obj = existing.get(key)
                    if obj is None:
                        to_create.append(ProfessionDemandStatus(
                            federal_operator=federal_operator,
                            profession=profession,
                            region=region,
                            year=default_year,
                            is_demanded=is_demanded,
                        ))
                    elif obj.is_demanded != is_demanded:
                        obj.is_demanded = is_demanded
                        to_update.append(obj)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Row {row_idx}: {exc}")

        BATCH = 1000
        for i in range(0, len(to_create), BATCH):
            ProfessionDemandStatus.objects.bulk_create(to_create[i:i + BATCH])
        for i in range(0, len(to_update), BATCH):
            ProfessionDemandStatus.objects.bulk_update(to_update[i:i + BATCH], ["is_demanded"])

        return {
            "created_professions": created_professions,
            "updated_professions": updated_professions,
            "created_statuses": len(to_create),
            "updated_statuses": len(to_update),
            "skipped_rows": skipped_rows,
            "errors_count": len(errors),
            "errors": errors[:20],
            "format": "wide_matrix",
        }

    def _import_row_based(
        self,
        rows,
        default_year,
        federal_operator,
        region_by_name,
        region_by_code,
        profession_by_number,
        profession_by_name,
    ):
        if not rows:
            raise ValueError("CSV has no rows.")
        header = rows[0]
        header_map = {self._norm_header(h): idx for idx, h in enumerate(header) if h}

        def pick(*aliases):
            for alias in aliases:
                if alias in header_map:
                    return header_map[alias]
            return None

        profession_number_col = pick("profession_number", "number", "код_профессии")
        profession_name_col = pick("profession_name", "name", "профессия")
        region_col = pick("region_name", "region", "регион")
        region_code_col = pick("region_code", "код_региона")
        demanded_col = pick("is_demanded", "demanded", "востребована")
        year_col = pick("year", "год")

        if demanded_col is None or (
            profession_name_col is None and profession_number_col is None
        ):
            raise ValueError(
                "Required columns: demanded + profession_name or profession_number."
            )

        created_professions = 0
        updated_professions = 0
        skipped_rows = 0
        errors = []

        existing = {
            (s.federal_operator_id, s.profession_id, s.region_id, s.year): s
            for s in ProfessionDemandStatus.objects.filter(
                federal_operator=federal_operator,
            ).only("id", "federal_operator_id", "profession_id", "region_id", "year", "is_demanded")
        }

        to_create = []
        to_update = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                if not row:
                    skipped_rows += 1
                    continue
                profession_number = None
                if profession_number_col is not None and profession_number_col < len(row):
                    raw_number = self._clean_cell(row[profession_number_col])
                    if raw_number:
                        profession_number = int(float(raw_number))

                profession_name = ""
                if profession_name_col is not None and profession_name_col < len(row):
                    profession_name = self._clean_cell(row[profession_name_col])
                if not profession_name and profession_number is None:
                    skipped_rows += 1
                    continue

                region = None
                if region_col is not None and region_col < len(row):
                    region_name = self._clean_cell(row[region_col]).lower()
                    if region_name:
                        region = region_by_name.get(region_name)
                if (
                    region is None
                    and region_code_col is not None
                    and region_code_col < len(row)
                ):
                    region_code = self._clean_cell(row[region_code_col]).lower()
                    if region_code:
                        region = region_by_code.get(region_code)
                if region is None:
                    skipped_rows += 1
                    continue

                demanded_raw = row[demanded_col] if demanded_col < len(row) else ""
                is_demanded = self._to_bool(demanded_raw)
                year = default_year
                if year_col is not None and year_col < len(row):
                    year_raw = self._clean_cell(row[year_col])
                    if year_raw:
                        year = int(float(year_raw))

                profession, was_created, was_updated = self._get_or_create_profession(
                    profession_number,
                    profession_name,
                    profession_by_number,
                    profession_by_name,
                )
                if was_created:
                    created_professions += 1
                if was_updated:
                    updated_professions += 1

                key = (federal_operator.id, profession.id, region.id, year)
                obj = existing.get(key)
                if obj is None:
                    to_create.append(ProfessionDemandStatus(
                        federal_operator=federal_operator,
                        profession=profession,
                        region=region,
                        year=year,
                        is_demanded=is_demanded,
                    ))
                elif obj.is_demanded != is_demanded:
                    obj.is_demanded = is_demanded
                    to_update.append(obj)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Row {row_idx}: {exc}")

        BATCH = 1000
        for i in range(0, len(to_create), BATCH):
            ProfessionDemandStatus.objects.bulk_create(to_create[i:i + BATCH])
        for i in range(0, len(to_update), BATCH):
            ProfessionDemandStatus.objects.bulk_update(to_update[i:i + BATCH], ["is_demanded"])

        return {
            "created_professions": created_professions,
            "updated_professions": updated_professions,
            "created_statuses": len(to_create),
            "updated_statuses": len(to_update),
            "skipped_rows": skipped_rows,
            "errors_count": len(errors),
            "errors": errors[:20],
            "format": "row_based",
        }

    def post(self, request):
        if not getattr(request.user, "is_admin_role", False):
            return Response(
                {"detail": "Only admin can import demand matrix."},
                status=status.HTTP_403_FORBIDDEN,
            )

        upload = request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": "File is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        raw = upload.read()
        try:
            filename = (upload.name or "").lower()
            if filename.endswith(".xlsx"):
                rows = self._read_xlsx_rows(raw)
            else:
                rows = self._read_csv_rows(raw)
        except Exception as exc:  # noqa: BLE001
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        federal_operator_id = request.data.get("federal_operator_id") or request.data.get("federal_operator")
        if not federal_operator_id:
            return Response(
                {"detail": "federal_operator_id is required for import."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            federal_operator = FederalOperator.objects.get(pk=int(federal_operator_id))
        except (ValueError, FederalOperator.DoesNotExist):
            return Response(
                {"detail": "Invalid federal_operator_id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        default_year = int(
            request.data.get("import_year")
            or request.data.get("year")
            or 2026
        )

        region_by_name = {r.name.strip().lower(): r for r in Region.objects.all()}
        region_by_code = {str(r.code).strip().lower(): r for r in Region.objects.all()}
        profession_by_number = {
            p.number: p for p in Profession.objects.all() if p.number is not None
        }
        profession_by_name = {p.name.strip().lower(): p for p in Profession.objects.all()}

        # Detect wide matrix: first row has region names (col>=2 или col>=1)
        first_row = rows[0] if rows else []
        matched_2 = sum(1 for c in first_row[2:] if self._clean_cell(c).lower() in region_by_name)
        matched_1 = sum(1 for c in first_row[1:] if self._clean_cell(c).lower() in region_by_name)
        if matched_2 >= 5 or matched_1 >= 5:
            response = self._import_wide_matrix(
                rows,
                default_year,
                federal_operator,
                region_by_name,
                profession_by_number,
                profession_by_name,
            )
        else:
            response = self._import_row_based(
                rows,
                default_year,
                federal_operator,
                region_by_name,
                region_by_code,
                profession_by_number,
                profession_by_name,
            )

        self._save_import_snapshot(federal_operator, default_year, request.user)

        return Response(response, status=status.HTTP_200_OK)


class DemandMatrixImportPreviewView(DemandMatrixImportView):
    """
    Dry-run import: returns invalid regions, new professions, and a
    change summary without writing anything to the database.
    """

    def _preview_wide_matrix(
        self, rows, region_by_name, profession_by_name,
        federal_operator, default_year,
    ):
        if len(rows) < 2:
            raise ValueError("Matrix file is too short.")

        header = rows[0]
        invalid_regions = []
        invalid_regions_seen = set()
        region_columns_1 = []
        region_columns_2 = []

        for idx in range(1, len(header)):
            raw = self._clean_cell(header[idx])
            norm = self._normalize(raw)
            if not norm:
                continue
            region = region_by_name.get(norm)
            if region:
                region_columns_1.append((idx, region))
            elif norm not in invalid_regions_seen:
                invalid_regions.append({"raw": raw, "normalized": norm})
                invalid_regions_seen.add(norm)
        for idx in range(2, len(header)):
            raw = self._clean_cell(header[idx])
            norm = self._normalize(raw)
            if not norm:
                continue
            region = region_by_name.get(norm)
            if region:
                region_columns_2.append((idx, region))
            elif norm not in invalid_regions_seen:
                invalid_regions.append({"raw": raw, "normalized": norm})
                invalid_regions_seen.add(norm)

        if len(region_columns_1) >= 5 and len(header) > 1 and self._normalize(header[1]) in region_by_name:
            region_columns = region_columns_1
            data_start = 1
            single_profession_column = True
        elif len(region_columns_2) >= 5:
            region_columns = region_columns_2
            data_start = 2
            single_profession_column = False
        else:
            region_columns = region_columns_1 or region_columns_2
            data_start = 1 if (region_columns and region_columns[0][0] == 1) else 2
            single_profession_column = bool(region_columns and region_columns[0][0] == 1)

        new_professions = []
        new_professions_seen = set()

        existing = {
            (s.profession_id, s.region_id): s
            for s in ProfessionDemandStatus.objects.filter(
                federal_operator=federal_operator, year=default_year,
            ).only("id", "profession_id", "region_id", "is_demanded")
        }

        created_professions = 0
        created_statuses = 0
        updated_statuses = 0
        skipped_rows = 0
        errors = []

        for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
            if not row or len(row) < 1:
                skipped_rows += 1
                continue

            if single_profession_column:
                raw_number_str = ""
                profession_name = self._clean_cell(row[0]) if len(row) > 0 else ""
            else:
                raw_number_str = self._clean_cell(row[0]) if len(row) > 0 else ""
                profession_name = self._clean_cell(row[1]) if len(row) > 1 else ""
            name_norm = self._normalize(profession_name)

            if not profession_name or name_norm == "итого":
                skipped_rows += 1
                continue

            profession_number = None
            if raw_number_str:
                try:
                    profession_number = int(float(raw_number_str))
                except Exception:
                    pass

            profession = profession_by_name.get(name_norm)

            if profession is None:
                if name_norm not in new_professions_seen:
                    new_professions.append({
                        "name": profession_name,
                        "number": profession_number,
                        "normalized": name_norm,
                    })
                    new_professions_seen.add(name_norm)
                    created_professions += 1
                created_statuses += len(region_columns)
            else:
                for col_idx, region in region_columns:
                    cell_val = row[col_idx] if col_idx < len(row) else ""
                    is_demanded = self._to_bool(cell_val)
                    key = (profession.id, region.id)
                    obj = existing.get(key)
                    if obj is None:
                        created_statuses += 1
                    elif obj.is_demanded != is_demanded:
                        updated_statuses += 1

        return {
            "invalid_regions": invalid_regions,
            "new_professions": new_professions,
            "preview": {
                "created_professions": created_professions,
                "created_statuses": created_statuses,
                "updated_statuses": updated_statuses,
                "skipped_rows": skipped_rows,
                "errors": errors[:20],
                "format": "wide_matrix",
            },
        }

    def _preview_row_based(
        self, rows, region_by_name, region_by_code, profession_by_name,
        federal_operator, default_year,
    ):
        if not rows:
            raise ValueError("CSV has no rows.")
        header = rows[0]
        header_map = {self._norm_header(h): idx for idx, h in enumerate(header) if h}

        def pick(*aliases):
            for alias in aliases:
                if alias in header_map:
                    return header_map[alias]
            return None

        profession_number_col = pick("profession_number", "number", "код_профессии")
        profession_name_col = pick("profession_name", "name", "профессия")
        region_col = pick("region_name", "region", "регион")
        region_code_col = pick("region_code", "код_региона")
        demanded_col = pick("is_demanded", "demanded", "востребована")
        year_col = pick("year", "год")

        if demanded_col is None or profession_name_col is None:
            raise ValueError(
                "Required columns: demanded + profession_name (or name/профессия)."
            )

        invalid_regions = []
        invalid_regions_seen = set()
        new_professions = []
        new_professions_seen = set()

        existing = {
            (s.federal_operator_id, s.profession_id, s.region_id, s.year): s
            for s in ProfessionDemandStatus.objects.filter(
                federal_operator=federal_operator,
            ).only(
                "id", "federal_operator_id", "profession_id",
                "region_id", "year", "is_demanded",
            )
        }

        created_professions = 0
        created_statuses = 0
        updated_statuses = 0
        skipped_rows = 0
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                if not row:
                    skipped_rows += 1
                    continue

                profession_name = ""
                if profession_name_col is not None and profession_name_col < len(row):
                    profession_name = self._clean_cell(row[profession_name_col])
                if not profession_name:
                    skipped_rows += 1
                    continue

                name_norm = self._normalize(profession_name)

                region = None
                region_raw = ""
                if region_col is not None and region_col < len(row):
                    region_raw = self._clean_cell(row[region_col])
                    region_norm = self._normalize(region_raw)
                    if region_norm:
                        region = region_by_name.get(region_norm)
                if (
                    region is None
                    and region_code_col is not None
                    and region_code_col < len(row)
                ):
                    code_raw = self._clean_cell(row[region_code_col])
                    code_norm = self._normalize(code_raw)
                    if code_norm:
                        region = region_by_code.get(code_norm)
                        if not region_raw:
                            region_raw = code_raw

                if region is None:
                    if region_raw:
                        norm = self._normalize(region_raw)
                        if norm and norm not in invalid_regions_seen:
                            invalid_regions.append({"raw": region_raw, "normalized": norm})
                            invalid_regions_seen.add(norm)
                    skipped_rows += 1
                    continue

                profession = profession_by_name.get(name_norm)

                if profession is None:
                    if name_norm not in new_professions_seen:
                        prof_number = None
                        if (
                            profession_number_col is not None
                            and profession_number_col < len(row)
                        ):
                            raw_num = self._clean_cell(row[profession_number_col])
                            if raw_num:
                                try:
                                    prof_number = int(float(raw_num))
                                except Exception:
                                    pass
                        new_professions.append({
                            "name": profession_name,
                            "number": prof_number,
                            "normalized": name_norm,
                        })
                        new_professions_seen.add(name_norm)
                        created_professions += 1
                    created_statuses += 1
                else:
                    demanded_raw = row[demanded_col] if demanded_col < len(row) else ""
                    is_demanded = self._to_bool(demanded_raw)
                    year = default_year
                    if year_col is not None and year_col < len(row):
                        year_raw = self._clean_cell(row[year_col])
                        if year_raw:
                            year = int(float(year_raw))
                    key = (federal_operator.id, profession.id, region.id, year)
                    obj = existing.get(key)
                    if obj is None:
                        created_statuses += 1
                    elif obj.is_demanded != is_demanded:
                        updated_statuses += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Row {row_idx}: {exc}")

        return {
            "invalid_regions": invalid_regions,
            "new_professions": new_professions,
            "preview": {
                "created_professions": created_professions,
                "created_statuses": created_statuses,
                "updated_statuses": updated_statuses,
                "skipped_rows": skipped_rows,
                "errors": errors[:20],
                "format": "row_based",
            },
        }

    def post(self, request):
        if not getattr(request.user, "is_admin_role", False):
            return Response(
                {"detail": "Only admin can import demand matrix."},
                status=status.HTTP_403_FORBIDDEN,
            )

        upload = request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": "File is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        raw = upload.read()
        try:
            filename = (upload.name or "").lower()
            if filename.endswith(".xlsx"):
                rows = self._read_xlsx_rows(raw)
            else:
                rows = self._read_csv_rows(raw)
        except Exception as exc:  # noqa: BLE001
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        federal_operator_id = (
            request.data.get("federal_operator_id")
            or request.data.get("federal_operator")
        )
        if not federal_operator_id:
            return Response(
                {"detail": "federal_operator_id is required for import."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            federal_operator = FederalOperator.objects.get(pk=int(federal_operator_id))
        except (ValueError, FederalOperator.DoesNotExist):
            return Response(
                {"detail": "Invalid federal_operator_id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        default_year = int(
            request.data.get("import_year")
            or request.data.get("year")
            or 2026
        )

        region_by_name = {self._normalize(r.name): r for r in Region.objects.all()}
        region_by_code = {self._normalize(r.code): r for r in Region.objects.all()}
        profession_by_name = {
            self._normalize(p.name): p for p in Profession.objects.all()
        }

        first_row = rows[0] if rows else []
        matched_2 = sum(1 for c in first_row[2:] if self._normalize(c) in region_by_name)
        matched_1 = sum(1 for c in first_row[1:] if self._normalize(c) in region_by_name)
        try:
            if matched_2 >= 5 or matched_1 >= 5:
                result = self._preview_wide_matrix(
                    rows, region_by_name, profession_by_name,
                    federal_operator, default_year,
                )
            else:
                result = self._preview_row_based(
                    rows, region_by_name, region_by_code, profession_by_name,
                    federal_operator, default_year,
                )
        except Exception as exc:  # noqa: BLE001
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(result, status=status.HTTP_200_OK)


class DemandMatrixImportApplyView(DemandMatrixImportView):
    """
    Apply import with user-provided region and profession mappings.
    Accepts the same file + federal_operator_id + import_year, plus
    region_mapping and profession_mapping JSON fields.
    """

    def _resolve_mappings(self, region_mapping_raw, profession_mapping_raw):
        """Pre-resolve mapping dicts into Region / Profession objects."""
        region_mapping = {}
        if region_mapping_raw:
            region_ids = set()
            for v in region_mapping_raw.values():
                try:
                    region_ids.add(int(v))
                except (ValueError, TypeError):
                    pass
            regions_by_id = {r.id: r for r in Region.objects.filter(id__in=region_ids)}
            for norm_key, rid in region_mapping_raw.items():
                try:
                    r = regions_by_id.get(int(rid))
                    if r:
                        region_mapping[self._normalize(norm_key)] = r
                except (ValueError, TypeError):
                    pass

        profession_mapping = {}
        if profession_mapping_raw:
            prof_ids = set()
            for v in profession_mapping_raw.values():
                if str(v) != "new":
                    try:
                        prof_ids.add(int(v))
                    except (ValueError, TypeError):
                        pass
            profs_by_id = {
                p.id: p for p in Profession.objects.filter(id__in=prof_ids)
            }
            for norm_key, val in profession_mapping_raw.items():
                key = self._normalize(norm_key)
                if str(val) == "new":
                    profession_mapping[key] = "new"
                else:
                    try:
                        p = profs_by_id.get(int(val))
                        if p:
                            profession_mapping[key] = p
                    except (ValueError, TypeError):
                        pass

        return region_mapping, profession_mapping

    def _resolve_profession(
        self, name_norm, profession_name, raw_number_str,
        profession_by_name, profession_mapping, counters,
    ):
        """Resolve a profession from file data using name match or mapping."""
        profession = profession_by_name.get(name_norm)
        if profession is not None:
            return profession

        mapping_val = profession_mapping.get(name_norm)
        if mapping_val is not None and mapping_val != "new":
            profession_by_name[name_norm] = mapping_val
            return mapping_val

        profession_number = None
        if raw_number_str:
            try:
                profession_number = int(float(raw_number_str))
            except Exception:
                pass
        new_number = profession_number
        if new_number is None:
            max_number = (
                Profession.objects.order_by("-number")
                .values_list("number", flat=True)
                .first()
                or 0
            )
            new_number = max_number + 1
        profession = Profession.objects.create(
            number=new_number,
            name=profession_name or f"Профессия {new_number}",
        )
        counters["created_professions"] += 1
        profession_by_name[name_norm] = profession
        return profession

    def _apply_wide_matrix(
        self, rows, default_year, federal_operator,
        region_by_name, profession_by_name,
        region_mapping, profession_mapping,
    ):
        if len(rows) < 2:
            raise ValueError("Matrix file is too short.")

        header = rows[0]
        region_columns_1 = []
        for idx in range(1, len(header)):
            raw = self._clean_cell(header[idx])
            norm = self._normalize(raw)
            if not norm:
                continue
            region = region_by_name.get(norm) or region_mapping.get(norm)
            if region:
                region_columns_1.append((idx, region))
        region_columns_2 = []
        for idx in range(2, len(header)):
            raw = self._clean_cell(header[idx])
            norm = self._normalize(raw)
            if not norm:
                continue
            region = region_by_name.get(norm) or region_mapping.get(norm)
            if region:
                region_columns_2.append((idx, region))

        h1 = self._normalize(header[1]) if len(header) > 1 else ""
        if len(region_columns_1) >= 5 and len(header) > 1 and (h1 in region_by_name or (region_mapping and h1 in region_mapping)):
            region_columns = region_columns_1
            single_profession_column = True
            data_start = 1
        elif len(region_columns_2) >= 5:
            region_columns = region_columns_2
            single_profession_column = False
            data_start = 2
        else:
            raise ValueError("No region columns resolved in matrix header.")

        counters = {"created_professions": 0}
        skipped_rows = 0
        errors = []

        existing = {
            (s.profession_id, s.region_id): s
            for s in ProfessionDemandStatus.objects.filter(
                federal_operator=federal_operator, year=default_year,
            ).only("id", "profession_id", "region_id", "is_demanded")
        }

        to_create = []
        to_update = []

        for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
            if not row or len(row) < 1:
                skipped_rows += 1
                continue

            if single_profession_column:
                raw_number_str = ""
                profession_name = self._clean_cell(row[0]) if len(row) > 0 else ""
            else:
                raw_number_str = self._clean_cell(row[0]) if len(row) > 0 else ""
                profession_name = self._clean_cell(row[1]) if len(row) > 1 else ""
            name_norm = self._normalize(profession_name)

            if not profession_name or name_norm == "итого":
                skipped_rows += 1
                continue

            try:
                profession = self._resolve_profession(
                    name_norm, profession_name, raw_number_str,
                    profession_by_name, profession_mapping, counters,
                )

                for col_idx, region in region_columns:
                    cell_val = row[col_idx] if col_idx < len(row) else ""
                    is_demanded = self._to_bool(cell_val)
                    key = (profession.id, region.id)
                    obj = existing.get(key)
                    if obj is None:
                        to_create.append(ProfessionDemandStatus(
                            federal_operator=federal_operator,
                            profession=profession,
                            region=region,
                            year=default_year,
                            is_demanded=is_demanded,
                        ))
                    elif obj.is_demanded != is_demanded:
                        obj.is_demanded = is_demanded
                        to_update.append(obj)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Row {row_idx}: {exc}")

        BATCH = 1000
        for i in range(0, len(to_create), BATCH):
            ProfessionDemandStatus.objects.bulk_create(to_create[i:i + BATCH])
        for i in range(0, len(to_update), BATCH):
            ProfessionDemandStatus.objects.bulk_update(
                to_update[i:i + BATCH], ["is_demanded"],
            )

        return {
            "created_professions": counters["created_professions"],
            "created_statuses": len(to_create),
            "updated_statuses": len(to_update),
            "skipped_rows": skipped_rows,
            "errors_count": len(errors),
            "errors": errors[:20],
            "format": "wide_matrix",
        }

    def _apply_row_based(
        self, rows, default_year, federal_operator,
        region_by_name, region_by_code, profession_by_name,
        region_mapping, profession_mapping,
    ):
        if not rows:
            raise ValueError("CSV has no rows.")
        header = rows[0]
        header_map = {
            self._norm_header(h): idx for idx, h in enumerate(header) if h
        }

        def pick(*aliases):
            for alias in aliases:
                if alias in header_map:
                    return header_map[alias]
            return None

        profession_number_col = pick("profession_number", "number", "код_профессии")
        profession_name_col = pick("profession_name", "name", "профессия")
        region_col = pick("region_name", "region", "регион")
        region_code_col = pick("region_code", "код_региона")
        demanded_col = pick("is_demanded", "demanded", "востребована")
        year_col = pick("year", "год")

        if demanded_col is None or profession_name_col is None:
            raise ValueError("Required columns: demanded + profession_name.")

        counters = {"created_professions": 0}
        skipped_rows = 0
        errors = []

        existing = {
            (s.federal_operator_id, s.profession_id, s.region_id, s.year): s
            for s in ProfessionDemandStatus.objects.filter(
                federal_operator=federal_operator,
            ).only(
                "id", "federal_operator_id", "profession_id",
                "region_id", "year", "is_demanded",
            )
        }

        to_create = []
        to_update = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                if not row:
                    skipped_rows += 1
                    continue

                profession_name = ""
                if profession_name_col is not None and profession_name_col < len(row):
                    profession_name = self._clean_cell(row[profession_name_col])
                if not profession_name:
                    skipped_rows += 1
                    continue

                name_norm = self._normalize(profession_name)

                region = None
                if region_col is not None and region_col < len(row):
                    region_raw = self._clean_cell(row[region_col])
                    region_norm = self._normalize(region_raw)
                    if region_norm:
                        region = (
                            region_by_name.get(region_norm)
                            or region_mapping.get(region_norm)
                        )
                if (
                    region is None
                    and region_code_col is not None
                    and region_code_col < len(row)
                ):
                    code_raw = self._clean_cell(row[region_code_col])
                    code_norm = self._normalize(code_raw)
                    if code_norm:
                        region = (
                            region_by_code.get(code_norm)
                            or region_mapping.get(code_norm)
                        )
                if region is None:
                    skipped_rows += 1
                    continue

                raw_number_str = ""
                if (
                    profession_number_col is not None
                    and profession_number_col < len(row)
                ):
                    raw_number_str = self._clean_cell(row[profession_number_col])

                profession = self._resolve_profession(
                    name_norm, profession_name, raw_number_str,
                    profession_by_name, profession_mapping, counters,
                )

                demanded_raw = row[demanded_col] if demanded_col < len(row) else ""
                is_demanded = self._to_bool(demanded_raw)
                year = default_year
                if year_col is not None and year_col < len(row):
                    year_raw = self._clean_cell(row[year_col])
                    if year_raw:
                        year = int(float(year_raw))

                key = (federal_operator.id, profession.id, region.id, year)
                obj = existing.get(key)
                if obj is None:
                    to_create.append(ProfessionDemandStatus(
                        federal_operator=federal_operator,
                        profession=profession,
                        region=region,
                        year=year,
                        is_demanded=is_demanded,
                    ))
                elif obj.is_demanded != is_demanded:
                    obj.is_demanded = is_demanded
                    to_update.append(obj)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Row {row_idx}: {exc}")

        BATCH = 1000
        for i in range(0, len(to_create), BATCH):
            ProfessionDemandStatus.objects.bulk_create(to_create[i:i + BATCH])
        for i in range(0, len(to_update), BATCH):
            ProfessionDemandStatus.objects.bulk_update(
                to_update[i:i + BATCH], ["is_demanded"],
            )

        return {
            "created_professions": counters["created_professions"],
            "created_statuses": len(to_create),
            "updated_statuses": len(to_update),
            "skipped_rows": skipped_rows,
            "errors_count": len(errors),
            "errors": errors[:20],
            "format": "row_based",
        }

    def post(self, request):
        if not getattr(request.user, "is_admin_role", False):
            return Response(
                {"detail": "Only admin can import demand matrix."},
                status=status.HTTP_403_FORBIDDEN,
            )

        upload = request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": "File is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        raw = upload.read()
        try:
            filename = (upload.name or "").lower()
            if filename.endswith(".xlsx"):
                rows = self._read_xlsx_rows(raw)
            else:
                rows = self._read_csv_rows(raw)
        except Exception as exc:  # noqa: BLE001
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        federal_operator_id = (
            request.data.get("federal_operator_id")
            or request.data.get("federal_operator")
        )
        if not federal_operator_id:
            return Response(
                {"detail": "federal_operator_id is required for import."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            federal_operator = FederalOperator.objects.get(pk=int(federal_operator_id))
        except (ValueError, FederalOperator.DoesNotExist):
            return Response(
                {"detail": "Invalid federal_operator_id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        default_year = int(
            request.data.get("import_year")
            or request.data.get("year")
            or 2026
        )

        region_mapping_raw = request.data.get("region_mapping", "{}")
        profession_mapping_raw = request.data.get("profession_mapping", "{}")
        try:
            rm = (
                json.loads(region_mapping_raw)
                if isinstance(region_mapping_raw, str)
                else (region_mapping_raw or {})
            )
            pm = (
                json.loads(profession_mapping_raw)
                if isinstance(profession_mapping_raw, str)
                else (profession_mapping_raw or {})
            )
        except json.JSONDecodeError:
            return Response(
                {"detail": "Invalid JSON in region_mapping or profession_mapping."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        region_mapping, profession_mapping = self._resolve_mappings(rm, pm)

        region_by_name = {self._normalize(r.name): r for r in Region.objects.all()}
        region_by_code = {self._normalize(r.code): r for r in Region.objects.all()}
        profession_by_name = {
            self._normalize(p.name): p for p in Profession.objects.all()
        }

        first_row = rows[0] if rows else []
        matched_2 = sum(1 for c in first_row[2:] if (self._normalize(c) in region_by_name or self._normalize(c) in region_mapping))
        matched_1 = sum(1 for c in first_row[1:] if (self._normalize(c) in region_by_name or self._normalize(c) in region_mapping))
        try:
            if matched_2 >= 5 or matched_1 >= 5:
                result = self._apply_wide_matrix(
                    rows, default_year, federal_operator,
                    region_by_name, profession_by_name,
                    region_mapping, profession_mapping,
                )
            else:
                result = self._apply_row_based(
                    rows, default_year, federal_operator,
                    region_by_name, region_by_code, profession_by_name,
                    region_mapping, profession_mapping,
                )
        except Exception as exc:  # noqa: BLE001
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        self._save_import_snapshot(federal_operator, default_year, request.user)

        return Response(result, status=status.HTTP_200_OK)


class DemandMatrixView(generics.ListAPIView):
    """
    Returns profession demand matrix across all regions.
    Query params: year, profession_ids, region_ids, demanded_only, approval_statuses
    """
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        year = int(request.query_params.get("year", 2026))
        profession_ids = request.query_params.get("profession_ids")
        region_ids = request.query_params.get("region_ids")
        demanded_only = request.query_params.get("demanded_only", "").lower() == "true"
        approval_statuses = request.query_params.get("approval_statuses")

        regions = Region.objects.all().order_by("name")
        if region_ids:
            ids = [int(x) for x in region_ids.split(",")]
            regions = regions.filter(id__in=ids)

        professions = Profession.objects.all().order_by("number")
        if profession_ids:
            ids = [int(x) for x in profession_ids.split(",")]
            professions = professions.filter(id__in=ids)
        
        # Filter by approval statuses
        if approval_statuses:
            status_list = approval_statuses.split(",")
            professions = professions.filter(
                approval_statuses__year=year,
                approval_statuses__approval_status__in=status_list
            ).distinct()

        statuses = ProfessionDemandStatus.objects.filter(year=year)
        if region_ids:
            ids = [int(x) for x in region_ids.split(",")]
            statuses = statuses.filter(region_id__in=ids)
        if profession_ids:
            ids = [int(x) for x in profession_ids.split(",")]
            statuses = statuses.filter(profession_id__in=ids)

        # All operators for "missing in" labels
        operators = list(
            FederalOperator.objects.values("id", "short_name", "name").order_by("name")
        )
        operator_display = {
            o["id"]: (o["short_name"] or o["name"]).strip() or o["name"]
            for o in operators
        }
        operator_ids = [o["id"] for o in operators]
        operator_count = len(operator_ids)

        # (profession_id, region_id) -> set(operator_id) where demand is True
        demanded_by_cell = {}
        for profession_id, region_id, operator_id, is_demanded in statuses.values_list(
            "profession_id", "region_id", "federal_operator_id", "is_demanded"
        ):
            if not is_demanded:
                continue
            key = (profession_id, region_id)
            demanded_by_cell.setdefault(key, set()).add(operator_id)

        # Fetch approval statuses
        approvals = ProfessionApprovalStatus.objects.filter(year=year)
        if region_ids:
            ids = [int(x) for x in region_ids.split(",")]
            approvals = approvals.filter(region_id__in=ids)
        if profession_ids:
            ids = [int(x) for x in profession_ids.split(",")]
            approvals = approvals.filter(profession_id__in=ids)

        approval_map = {
            (profession_id, region_id): approval_status
            for profession_id, region_id, approval_status in approvals.values_list(
                "profession_id", "region_id", "approval_status"
            )
        }

        region_list = list(regions.values("id", "name"))
        profession_list = list(professions.values("id", "number", "name"))

        result = []
        for prof in profession_list:
            region_demands = {}
            region_approvals = {}
            region_missing_operators = {}
            has_any = False
            for r in region_list:
                cell_key = (prof["id"], r["id"])
                demanded_ops = demanded_by_cell.get(cell_key, set())
                # demanded if at least one operator has demand
                val = bool(demanded_ops)
                region_demands[str(r["id"])] = val
                region_approvals[str(r["id"])] = approval_map.get(
                    (prof["id"], r["id"]), None
                )
                # Send "missing in operators" only for partial mismatch:
                # at least one operator has demand, but not all.
                if val and len(demanded_ops) < operator_count:
                    missing = [
                        {"id": op_id, "short_name": operator_display[op_id]}
                        for op_id in operator_ids
                        if op_id not in demanded_ops
                    ]
                    if missing:
                        region_missing_operators[str(r["id"])] = missing
                if val:
                    has_any = True
            if demanded_only and not has_any:
                continue
            result.append({
                "profession_id": prof["id"],
                "profession_number": prof["number"],
                "profession_name": prof["name"],
                "regions": region_demands,
                "approvals": region_approvals,
                "region_missing_operators": region_missing_operators,
            })

        return Response({
            "regions": region_list,
            "professions": result,
            "year": year,
            "federal_operators": [
                {"id": o["id"], "short_name": operator_display[o["id"]]}
                for o in operators
            ],
        })
