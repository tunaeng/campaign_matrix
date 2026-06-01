import io
import logging
import json
import re
from datetime import timedelta

import requests as http_requests
from openpyxl import Workbook, load_workbook
from django.conf import settings
from django.http import HttpResponse
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Exists, OuterRef, ProtectedError
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.utils import timezone
from django.utils.text import slugify
from apps.reference.models import Region
from .models import (
    Organization,
    OrganizationInteraction,
    Contact,
    EntityFieldChange,
    ImportBatch,
    ImportBatchRecord,
    OrganizationTag,
    Project,
    ProjectOrganizationMembership,
    UserActingOrganization,
    BitrixOAuthConnection,
)
from .serializers import (
    OrganizationSerializer, OrganizationShortSerializer,
    OrganizationInteractionSerializer, ContactSerializer,
    EntityFieldChangeSerializer,
    ImportBatchSerializer,
    OrganizationTagSerializer,
    ProjectSerializer,
    ProjectOrganizationMembershipSerializer,
    UserActingOrganizationSerializer,
)

logger = logging.getLogger(__name__)

DEFAULT_ACTING_ORGANIZATION_INN = "6321261206"
DEFAULT_ACTING_ORGANIZATION_NAME = (
    "ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "
    "«Союз Энергетиков Поволжья»"
)
DEFAULT_ACTING_ORGANIZATION_SHORT_NAME = "ООО «СЭП»"
DEFAULT_ACTING_ORGANIZATION_NOTES = "ИНН/КПП 6321261206 / 632101001"


BITRIX_ORGANIZATIONS_ENDPOINTS = ("/contacts/api/organization/", "/api/organization/")
BITRIX_CONTACTS_LIST_ENDPOINTS = ("/contacts/api/contacts/", "/api/contact/")
BITRIX_CONTACT_ADD_ENDPOINTS = ("/api/contact/add/", "/contacts/api/contact/add/")
BITRIX_CONTACT_UPDATE_ENDPOINTS = ("/api/contact/update/", "/contacts/api/contact/update/")
BITRIX_CONTACT_HISTORY_ENDPOINTS = ("/api/contact/history/", "/contacts/api/contact/history/")
BITRIX_COMMUNICATION_LIST_ENDPOINTS = ("/api/communication/", "/contacts/api/communication/")
BITRIX_COMMUNICATION_ADD_ENDPOINTS = ("/api/communication/add/", "/contacts/api/communication/add/")
BITRIX_COMMUNICATION_UPDATE_ENDPOINTS = ("/api/communication/update/", "/contacts/api/communication/update/")

_CHANGE_SOURCE_VALUES = {item.value for item in EntityFieldChange.Source}


class RegistryPagination(PageNumberPagination):
    page_size = 50
    max_page_size = 500
    page_size_query_param = "page_size"


_CONTACT_AUDIT_FIELDS = (
    "organization",
    "type",
    "comment",
    "current",
    "first_name",
    "last_name",
    "middle_name",
    "position",
    "phone",
    "phone_extension",
    "email",
    "messenger",
    "is_manager",
    "department_name",
    "tags",
)

_ORGANIZATION_AUDIT_FIELDS = (
    "name",
    "short_name",
    "inn",
    "org_type",
    "region",
    "parent_organization",
    "contact_person",
    "contact_email",
    "contact_phone",
    "contact_phone_extension",
    "is_our_side",
    "description",
    "tags",
)


def _normalize_change_source(raw: str | None) -> str:
    if raw in _CHANGE_SOURCE_VALUES:
        return raw
    return EntityFieldChange.Source.MANUAL


def _value_to_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, tuple, set)):
        return json.dumps(list(value), ensure_ascii=False)
    return str(value)


def _capture_contact_state(contact: Contact) -> dict:
    return {
        "organization": contact.organization_id,
        "type": contact.type or "",
        "comment": contact.comment or "",
        "current": bool(contact.current),
        "first_name": contact.first_name or "",
        "last_name": contact.last_name or "",
        "middle_name": contact.middle_name or "",
        "position": contact.position or "",
        "phone": contact.phone or "",
        "phone_extension": contact.phone_extension or "",
        "email": contact.email or "",
        "messenger": contact.messenger or "",
        "is_manager": bool(contact.is_manager),
        "department_name": contact.department_name or "",
        "tags": list(contact.tags.order_by("id").values_list("id", flat=True)),
    }


def _capture_organization_state(organization: Organization) -> dict:
    return {
        "name": organization.name or "",
        "short_name": organization.short_name or "",
        "inn": organization.inn or "",
        "org_type": organization.org_type or "",
        "region": organization.region_id,
        "parent_organization": organization.parent_organization_id,
        "contact_person": organization.contact_person or "",
        "contact_email": organization.contact_email or "",
        "contact_phone": organization.contact_phone or "",
        "contact_phone_extension": organization.contact_phone_extension or "",
        "is_our_side": bool(organization.is_our_side),
        "description": organization.description or "",
        "tags": list(organization.tags.order_by("id").values_list("id", flat=True)),
    }


def _ensure_default_acting_organization() -> Organization:
    org, _ = Organization.objects.get_or_create(
        inn=DEFAULT_ACTING_ORGANIZATION_INN,
        defaults={
            "name": DEFAULT_ACTING_ORGANIZATION_NAME,
            "short_name": DEFAULT_ACTING_ORGANIZATION_SHORT_NAME,
            "org_type": Organization.OrgType.PRIVATE,
            "notes": DEFAULT_ACTING_ORGANIZATION_NOTES,
            "is_our_side": True,
        },
    )
    return org


def _create_field_change_rows(
    *,
    before: dict | None,
    after: dict,
    fields: tuple[str, ...],
    source: str,
    changed_by,
    organization: Organization | None = None,
    contact: Contact | None = None,
):
    rows = []
    for field in fields:
        old_value = None if before is None else before.get(field)
        new_value = after.get(field)
        if old_value == new_value:
            continue
        # Для создания пропускаем пустые значения, чтобы не засорять журнал.
        if before is None and new_value in ("", None, [], False):
            continue
        rows.append(
            EntityFieldChange(
                organization=organization,
                contact=contact,
                field_name=field,
                old_value=_value_to_text(old_value),
                new_value=_value_to_text(new_value),
                source=source,
                changed_by=changed_by,
            )
        )
    if rows:
        EntityFieldChange.objects.bulk_create(rows)


_IMPORT_HEADER_ALIASES = {
    "organization": {
        "организация",
        "наименование",
        "название организации",
        "org",
        "organization",
    },
    "full_name": {"фио", "контакт", "контактное лицо", "фамилия имя отчество"},
    "position": {"должность", "позиция"},
    "comment": {
        "описание контакта (свободная форма)",
        "описание котнтакта (свободная форма)",
        "описание",
        "комментарий",
    },
    "phone": {"телефон", "моб телефон", "контактный телефон"},
    "phone_extension": {
        "добавочный",
        "добавочный номер",
        "доб",
        "доб.",
        "доб номер",
        "добавочный телефон",
        "ext",
        "extension",
    },
    "email": {"email", "e-mail", "почта"},
    "messenger_link": {
        "cсылка на мессенджер (если в нем идет общение)",
        "cсылка на мессенджер (если в нём идет общение)",
        "ссылка на мессенджер (если в нем идет общение)",
        "ссылка на мессенджер (если в нём идет общение)",
        "ссылка на мессенджер",
    },
    "messenger_type": {"какой мессенджер", "мессенджер"},
    "region": {"регион"},
    "inn": {"инн"},
    "short_name": {
        "краткое наименование",
        "краткое название",
        "краткое",
        "short name",
        "short_name",
    },
    "org_type": {
        "тип организации",
        "тип",
        "org type",
        "org_type",
        "типорганизации",
    },
    "organization_tags": {
        "теги организации",
        "теги организаций",
        "теги",
        "метки организации",
        "organization tags",
        "organization_tags",
    },
    "contact_tags": {
        "теги контакта",
        "теги контактов",
        "contact tags",
        "contact_tags",
    },
    "parent_organization": {
        "головная организация",
        "головная",
        "родительская организация",
        "инн головной",
        "инн головной организации",
        "parent",
        "parent_organization",
        "head organization",
    },
}


def _normalize_header(value):
    s = str(value or "").strip().lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def _normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _trim_for_model_field(model_cls, field_name, value):
    """
    Normalize text and trim to model field max_length when needed.
    """
    normalized = _normalize_cell(value)
    if not normalized:
        return normalized
    try:
        field = model_cls._meta.get_field(field_name)
        max_length = getattr(field, "max_length", None)
        if isinstance(max_length, int) and max_length > 0 and len(normalized) > max_length:
            return normalized[:max_length]
    except Exception:
        # Keep normalized value if field metadata is unavailable for any reason.
        return normalized
    return normalized


def _normalize_inn(value):
    s = re.sub(r"\D", "", _normalize_cell(value))
    if len(s) in (10, 12):
        return s
    return ""


def _split_phone_extension_from_combined(raw):
    """
    Добавочный в той же ячейке, что и телефон (доб., добавочный, ext или запятая перед номером).
    Возвращает (основной телефон, добавочный). Добавочный может содержать цифры и дефисы.
    """
    s = _normalize_cell(raw)
    if not s:
        return "", ""
    patterns = [
        r"(?i)\s+добавочн[а-яё]*\.?\s*[:\-]?\s*([\d\-\s]+)\s*$",
        r"(?i)\s+доб\.?\s*[:\-]?\s*([\d\-\s]+)\s*$",
        r"(?i)\s+ext\.?\s*[:\-]?\s*([\d\-\s]+)\s*$",
        r"(?i)\s*[,;]\s*(?:доб\.?\s*)?([\d\-\s]+)\s*$",
    ]
    for pat in patterns:
        m = re.search(pat, s)
        if m:
            ext_raw = (m.group(1) or "").strip()
            main = s[: m.start()].strip().rstrip(",;")
            if main:
                return main, ext_raw
    return s, ""


def _resolve_org_type_from_cell(raw):
    """Возвращает значение Organization.OrgType или None, если ячейка пустая."""
    s = _normalize_header(raw)
    if not s:
        return None
    valid = {item.value for item in Organization.OrgType}
    if s in valid:
        return s
    for item in Organization.OrgType:
        if s == _normalize_header(str(item.label)):
            return item.value
    synonyms = {
        "подразделение": Organization.OrgType.COMPANY_BRANCH,
        "подразделение компании": Organization.OrgType.COMPANY_BRANCH,
        "подразделение компании без инн": Organization.OrgType.COMPANY_BRANCH,
        "федеральное": Organization.OrgType.FEDERAL,
        "муниципальное": Organization.OrgType.MUNICIPAL,
        "частное": Organization.OrgType.PRIVATE,
        "частная": Organization.OrgType.PRIVATE,
        "коммерческое": Organization.OrgType.PRIVATE,
        "коммерческая": Organization.OrgType.PRIVATE,
    }
    return synonyms.get(s)


def _split_comma_separated_names(raw):
    s = _normalize_cell(raw)
    if not s:
        return []
    return [p.strip() for p in re.split(r"[,;]+", s) if p.strip()]


def _resolve_import_parent_organization(row, line_no, errors, *, existing_org):
    """
    Головная организация для строки импорта подразделения (ИНН или наименование).
    Если колонка пуста и existing_org уже связана с головной — оставляем её.
    Возвращает (Organization | None, успех).
    """
    ref = _normalize_cell(row.get("parent_organization"))
    if not ref:
        if existing_org and existing_org.pk and existing_org.parent_organization_id:
            po = existing_org.parent_organization
            if po and po.inn and str(po.inn).strip():
                return po, True
        errors.append(
            f"Строка {line_no}: для подразделения укажите головную организацию "
            f"(колонка «Головная организация»: ИНН или наименование юрлица)"
        )
        return None, False
    parent, _, _ = _find_organization_by_ref(ref)
    if parent is None:
        errors.append(f"Строка {line_no}: головная организация не найдена («{ref}»)")
        return None, False
    if not parent.inn or not str(parent.inn).strip():
        errors.append(
            f"Строка {line_no}: у головной организации «{ref}» в системе должен быть заполнен ИНН"
        )
        return None, False
    return parent, True


def _resolve_tags_by_names(raw, *, for_organization: bool):
    """
    Сопоставление OrganizationTag по имени или slug (без учёта регистра имени).
    for_organization: True — теги сущности organizations + all; иначе contacts + all.
    """
    names = _split_comma_separated_names(raw)
    if not names:
        return [], []
    allowed = [OrganizationTag.TagType.ALL]
    allowed.append(
        OrganizationTag.TagType.ORGANIZATIONS
        if for_organization
        else OrganizationTag.TagType.CONTACTS
    )
    qs = OrganizationTag.objects.filter(tag_type__in=allowed)
    found = []
    unknown = []
    seen_ids = set()
    for name in names:
        tag = qs.filter(name__iexact=name).first()
        if tag is None:
            tag = qs.filter(slug__iexact=name).first()
        if tag is None:
            unknown.append(name)
        elif tag.id not in seen_ids:
            seen_ids.add(tag.id)
            found.append(tag)
    return found, unknown


def _extract_tag_ids(raw):
    out = set()
    if raw is None:
        return []
    if isinstance(raw, (list, tuple)):
        for item in raw:
            out.update(_extract_tag_ids(item))
        return sorted(out)
    s = str(raw).strip()
    if not s:
        return []
    for p in re.split(r"[,;\s]+", s):
        if p.isdigit():
            out.add(int(p))
    return sorted(out)


def _xlsx_column_map(header_values):
    """header_values: первая строка листа (tuple/list из values_only)."""
    result = {}
    if not header_values:
        return result
    for idx, raw in enumerate(header_values):
        normalized = _normalize_header(raw)
        if not normalized:
            continue
        for key, aliases in _IMPORT_HEADER_ALIASES.items():
            if normalized in aliases:
                result[key] = idx
    if "organization" not in result:
        for idx, raw in enumerate(header_values):
            normalized = _normalize_header(raw)
            if "головн" in normalized or "родительск" in normalized:
                continue
            if "организац" in normalized:
                result["organization"] = idx
                break
    return result


def _organization_import_sheet_score(col_map: dict) -> int:
    """Оценка листа для импорта организаций (несколько листов в одной книге)."""
    s = 0
    if "inn" in col_map:
        s += 4
    if "organization" in col_map:
        s += 2
    if "short_name" in col_map:
        s += 2
    if "region" in col_map:
        s += 1
    if "org_type" in col_map:
        s += 1
    if "organization_tags" in col_map:
        s += 1
    if "parent_organization" in col_map:
        s += 1
    return s


def _contact_import_sheet_score(col_map: dict) -> int:
    s = 0
    if "full_name" in col_map:
        s += 3
    if "organization" in col_map:
        s += 2
    if "inn" in col_map:
        s += 1
    if "phone" in col_map or "email" in col_map or "phone_extension" in col_map:
        s += 1
    if "org_type" in col_map:
        s += 1
    if "organization_tags" in col_map:
        s += 1
    if "contact_tags" in col_map:
        s += 1
    return s


def _pick_import_worksheet(wb, import_kind: str):
    worksheets = list(wb.worksheets)
    if not worksheets:
        raise ValueError("В файле Excel нет листов")
    best_ws = worksheets[0]
    best_score = -1
    for ws in worksheets:
        header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        cm = _xlsx_column_map(header_values)
        if import_kind == "organizations":
            score = _organization_import_sheet_score(cm)
        else:
            score = _contact_import_sheet_score(cm)
        if score > best_score:
            best_score = score
            best_ws = ws
    return best_ws


def _xlsx_rows(uploaded_file, *, import_kind: str = "contacts"):
    """
    Читает весь файл в память: openpyxl read_only + некоторые UploadedFile из Django
    дают сбой при произвольном доступе к ZIP/xlsx.
    """
    if hasattr(uploaded_file, "seek"):
        try:
            uploaded_file.seek(0)
        except (OSError, ValueError, io.UnsupportedOperation):
            pass
    raw = uploaded_file.read()
    if not raw:
        return []
    bio = io.BytesIO(raw)
    wb = load_workbook(bio, data_only=True, read_only=False)
    ws = _pick_import_worksheet(wb, import_kind)
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    col_map = _xlsx_column_map(header_values)
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {}
        for key, col_idx in col_map.items():
            if row and col_idx < len(row):
                item[key] = row[col_idx]
        if any(_normalize_cell(v) for v in item.values()):
            rows.append((idx, item))
    return rows


def _parse_person_name(full_name):
    raw = _normalize_cell(full_name)
    if not raw or raw in {"-", "—"}:
        return "", "", ""
    parts = [p for p in re.split(r"\s+", raw) if p]
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[0], parts[1], ""
    return parts[0], parts[1], " ".join(parts[2:])


def _find_organization_by_ref(org_ref):
    ref = _normalize_cell(org_ref)
    if not ref:
        return None, "", ""
    inn = _normalize_inn(ref)
    if inn:
        return Organization.objects.filter(inn=inn).first(), "inn", inn
    qs = Organization.objects.filter(name__iexact=ref)
    org = qs.first()
    if org:
        return org, "name", ref
    org = Organization.objects.filter(short_name__iexact=ref).first()
    if org:
        return org, "name", ref
    org = Organization.objects.filter(name__icontains=ref).order_by("id").first()
    return org, "name", ref


def _find_organization_for_import(row, org_raw):
    """Поиск организации при импорте: сначала ИНН из колонки, затем org_raw как ИНН/точное имя."""
    inn_from_col = _normalize_inn(row.get("inn"))
    if inn_from_col:
        org = Organization.objects.filter(inn=inn_from_col).first()
        if org:
            return org, "inn", inn_from_col

    ref = _normalize_cell(org_raw)
    if not ref:
        return None, "", ""

    inn = _normalize_inn(ref)
    if inn:
        org = Organization.objects.filter(inn=inn).first()
        if org:
            return org, "inn", inn

    org = Organization.objects.filter(name__iexact=ref).first()
    if org:
        return org, "name", ref
    org = Organization.objects.filter(short_name__iexact=ref).first()
    if org:
        return org, "name", ref
    return None, "", ""


def _restore_organization_state(organization: Organization, snapshot: dict):
    organization.name = snapshot.get("name") or organization.name
    organization.short_name = snapshot.get("short_name") or ""
    inn_val = snapshot.get("inn") or ""
    organization.inn = inn_val or None
    organization.org_type = snapshot.get("org_type") or organization.org_type
    organization.region_id = snapshot.get("region")
    organization.parent_organization_id = snapshot.get("parent_organization")
    organization.contact_person = snapshot.get("contact_person") or ""
    organization.contact_email = snapshot.get("contact_email") or ""
    organization.contact_phone = snapshot.get("contact_phone") or ""
    organization.contact_phone_extension = snapshot.get("contact_phone_extension") or ""
    organization.is_our_side = bool(snapshot.get("is_our_side"))
    organization.description = snapshot.get("description") or ""
    organization.save()
    tag_ids = snapshot.get("tags") or []
    organization.tags.set(tag_ids)


def _restore_contact_state(contact: Contact, snapshot: dict):
    contact.organization_id = snapshot.get("organization") or contact.organization_id
    contact.type = snapshot.get("type") or contact.type
    contact.comment = snapshot.get("comment") or ""
    contact.current = bool(snapshot.get("current", True))
    contact.first_name = snapshot.get("first_name") or ""
    contact.last_name = snapshot.get("last_name") or ""
    contact.middle_name = snapshot.get("middle_name") or ""
    contact.position = snapshot.get("position") or ""
    contact.phone = snapshot.get("phone") or ""
    contact.phone_extension = snapshot.get("phone_extension") or ""
    contact.email = snapshot.get("email") or ""
    contact.is_manager = bool(snapshot.get("is_manager"))
    contact.department_name = snapshot.get("department_name") or ""
    contact.messenger = snapshot.get("messenger") or ""
    contact.save()
    tag_ids = snapshot.get("tags") or []
    contact.tags.set(tag_ids)


def _track_import_record(batch: ImportBatch, *, organization=None, contact=None, action: str, before=None):
    ImportBatchRecord.objects.create(
        batch=batch,
        organization=organization,
        contact=contact,
        action=action,
        snapshot=before or {},
    )


def _rollback_import_batch(batch: ImportBatch, actor):
    if batch.status == ImportBatch.Status.ROLLED_BACK:
        return {"detail": "Импорт уже откатан."}, status.HTTP_400_BAD_REQUEST

    records = list(
        batch.records.select_related("organization", "contact").order_by("-id")
    )
    reverted = 0
    deleted = 0
    skipped = 0
    errors = []

    for record in records:
        try:
            if record.action == ImportBatchRecord.Action.CREATED:
                if record.organization_id and Organization.objects.filter(id=record.organization_id).exists():
                    record.organization.delete()
                    deleted += 1
                elif record.contact_id and Contact.objects.filter(id=record.contact_id).exists():
                    record.contact.delete()
                    deleted += 1
                else:
                    skipped += 1
            elif record.action == ImportBatchRecord.Action.UPDATED:
                if record.organization_id and record.snapshot:
                    org = Organization.objects.filter(id=record.organization_id).first()
                    if org:
                        _restore_organization_state(org, record.snapshot)
                        reverted += 1
                    else:
                        skipped += 1
                elif record.contact_id and record.snapshot:
                    contact = Contact.objects.filter(id=record.contact_id).first()
                    if contact:
                        _restore_contact_state(contact, record.snapshot)
                        reverted += 1
                    else:
                        skipped += 1
                else:
                    skipped += 1
        except ProtectedError:
            target = record.organization or record.contact
            errors.append(f"Не удалось удалить {target}: есть связанные записи.")
            skipped += 1
        except Exception as exc:
            errors.append(str(exc))
            skipped += 1

    batch.status = ImportBatch.Status.ROLLED_BACK
    batch.rolled_back_at = timezone.now()
    batch.rolled_back_by = actor
    batch.save(update_fields=["status", "rolled_back_at", "rolled_back_by"])

    return {
        "deleted": deleted,
        "reverted": reverted,
        "skipped": skipped,
        "errors": errors[:100],
    }, status.HTTP_200_OK


def _resolve_region(region_name):
    name = _normalize_cell(region_name)
    if not name:
        return None
    region = Region.objects.filter(name__iexact=name).first()
    if region:
        return region
    return Region.objects.filter(name__icontains=name).order_by("id").first()


def _extract_bitrix_contact_id(data):
    if isinstance(data, dict):
        raw = data.get("id")
        if raw is not None:
            try:
                return int(raw)
            except (TypeError, ValueError):
                return None
    return None


def _local_organization_by_inn(inn):
    if not inn:
        return None
    inn = str(inn).strip()
    if not inn:
        return None
    return Organization.objects.filter(inn=inn).first()


def _strip_sync_local_from_payload(data):
    """Поля только для Matrix, не отправлять в Bitrix."""
    if not isinstance(data, dict):
        return data
    out = {k: v for k, v in data.items() if k != "sync_local"}
    return out


def _mirror_bitrix_contact_to_local(organization_inn, bitrix_response, request_payload, *, sync_local):
    """
    Дублирует контакт в локальную БД, если есть Organization с таким ИНН.

    Returns (Contact | None, status: str) status in ok | no_local_org | off | not_applicable
    """
    if not sync_local:
        return None, "off"
    org = _local_organization_by_inn(organization_inn)
    if org is None:
        return None, "no_local_org"

    bid = _extract_bitrix_contact_id(bitrix_response)
    merged = dict(request_payload) if isinstance(request_payload, dict) else {}
    if isinstance(request_payload, dict) and "manager" in request_payload:
        merged.setdefault(
            "is_manager",
            _parse_bool(request_payload.get("manager"), False),
        )
    if isinstance(bitrix_response, dict):
        merged.setdefault("type", bitrix_response.get("type", "person"))
        merged.setdefault("first_name", bitrix_response.get("first_name", ""))
        merged.setdefault("last_name", bitrix_response.get("last_name", ""))
        merged.setdefault("middle_name", bitrix_response.get("middle_name", ""))
        merged.setdefault("position", bitrix_response.get("position", ""))
        merged.setdefault("comment", bitrix_response.get("comment", ""))
        merged.setdefault("current", bitrix_response.get("current", True))
        merged.setdefault("department_name", bitrix_response.get("department_name", ""))
        if "manager" in bitrix_response:
            merged["is_manager"] = _parse_bool(bitrix_response.get("manager"), False)
        elif "is_manager" not in merged:
            merged["is_manager"] = False

    c_type = merged.get("type", "person")
    defaults = {
        "comment": merged.get("comment", ""),
        "current": _parse_bool(merged.get("current"), True),
        "first_name": merged.get("first_name", ""),
        "last_name": merged.get("last_name", ""),
        "middle_name": merged.get("middle_name", ""),
        "position": merged.get("position", ""),
        "is_manager": _parse_bool(merged.get("is_manager"), False),
        "department_name": merged.get("department_name", ""),
    }

    if bid:
        contact, _ = Contact.objects.update_or_create(
            bitrix_contact_id=bid,
            defaults={**defaults, "organization": org, "type": c_type},
        )
        return contact, "ok"

    if c_type == "person":
        contact, _ = Contact.objects.update_or_create(
            organization=org,
            type=c_type,
            first_name=defaults["first_name"],
            last_name=defaults["last_name"],
            middle_name=defaults["middle_name"],
            defaults=defaults,
        )
        return contact, "ok"
    if c_type == "department":
        contact, _ = Contact.objects.update_or_create(
            organization=org,
            type=c_type,
            department_name=defaults["department_name"],
            defaults=defaults,
        )
        return contact, "ok"

    contact = Contact.objects.create(organization=org, type=c_type, **defaults)
    return contact, "ok"


class OrganizationViewSet(viewsets.ModelViewSet):
    serializer_class = OrganizationSerializer
    pagination_class = RegistryPagination
    filterset_fields = ["org_type", "region", "parent_organization", "is_our_side"]
    search_fields = ["name", "short_name", "inn"]

    def get_queryset(self):
        qs = Organization.objects.select_related(
            "region", "parent_organization"
        ).prefetch_related("interactions")

        has_history = self.request.query_params.get("has_history")
        if has_history is not None:
            has_history = has_history.lower() == "true"
            qs = qs.annotate(
                _has_history=Exists(
                    OrganizationInteraction.objects.filter(
                        organization=OuterRef("pk")
                    )
                )
            ).filter(_has_history=has_history)

        region_ids = self.request.query_params.get("region_ids")
        if region_ids:
            ids = [int(x) for x in region_ids.split(",")]
            qs = qs.filter(region_id__in=ids)

        tag_ids = self.request.query_params.get("tags")
        if tag_ids:
            ids = [int(x) for x in tag_ids.split(",") if x.strip().isdigit()]
            if ids:
                qs = qs.filter(tags__id__in=ids).distinct()

        project_id = self.request.query_params.get("project")
        if project_id and project_id.isdigit():
            qs = qs.filter(project_memberships__project_id=int(project_id))

        role = self.request.query_params.get("role")
        if role:
            qs = qs.filter(project_memberships__role=role)

        if project_id or role:
            qs = qs.distinct()

        return qs

    @action(detail=True, methods=["get"], url_path="change-log")
    def change_log(self, request, pk=None):
        org = self.get_object()
        qs = (
            EntityFieldChange.objects.filter(organization=org)
            .select_related("changed_by", "contact", "organization")
            .order_by("-changed_at", "-id")
        )
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = EntityFieldChangeSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = EntityFieldChangeSerializer(qs, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        organization = serializer.save()
        after = _capture_organization_state(organization)
        _create_field_change_rows(
            before=None,
            after=after,
            fields=_ORGANIZATION_AUDIT_FIELDS,
            source=_normalize_change_source(self.request.query_params.get("source")),
            changed_by=self.request.user if self.request.user.is_authenticated else None,
            organization=organization,
        )

    def perform_update(self, serializer):
        before = _capture_organization_state(serializer.instance)
        organization = serializer.save()
        after = _capture_organization_state(organization)
        _create_field_change_rows(
            before=before,
            after=after,
            fields=_ORGANIZATION_AUDIT_FIELDS,
            source=_normalize_change_source(self.request.query_params.get("source")),
            changed_by=self.request.user if self.request.user.is_authenticated else None,
            organization=organization,
        )

    @action(detail=False, methods=["post"], url_path="import-xlsx")
    def import_xlsx(self, request):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Файл не передан"}, status=status.HTTP_400_BAD_REQUEST)
        if not file_obj.name.lower().endswith(".xlsx"):
            return Response({"detail": "Поддерживаются только .xlsx файлы"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            source = _normalize_change_source(
                _form_data_scalar(request.data, "source") or EntityFieldChange.Source.BULK
            )
            default_org_type = _form_data_scalar(request.data, "default_org_type") or Organization.OrgType.OTHER
            if default_org_type not in {item.value for item in Organization.OrgType}:
                default_org_type = Organization.OrgType.OTHER
            imported_tag_ids = _extract_tag_ids(
                _form_data_scalar(request.data, "tag_ids") or _form_data_scalar(request.data, "tags")
            )
            imported_tags = list(OrganizationTag.objects.filter(id__in=imported_tag_ids))
            update_existing = _parse_bool(_form_data_scalar(request.data, "update_existing"), True)
        except Exception as exc:
            logger.exception("organizations import_xlsx: параметры запроса")
            return Response(
                {"detail": f"Ошибка разбора формы импорта: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows = _xlsx_rows(file_obj, import_kind="organizations")
        except Exception as exc:
            logger.exception("organizations import_xlsx: не удалось разобрать файл")
            return Response(
                {"detail": f"Не удалось прочитать Excel (.xlsx): {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        skipped = 0
        errors = []
        actor = request.user if request.user.is_authenticated else None
        batch = ImportBatch.objects.create(
            entity_type=ImportBatch.EntityType.ORGANIZATIONS,
            file_name=file_obj.name,
            uploaded_by=actor,
            total_rows=len(rows),
        )

        for line_no, row in rows:
            org_raw = _normalize_cell(
                row.get("organization") or row.get("name") or row.get("inn")
            )
            if not org_raw:
                skipped += 1
                continue

            try:
                with transaction.atomic():
                    inn_from_col = _normalize_inn(row.get("inn"))
                    found, ref_kind, ref_value = _find_organization_for_import(row, org_raw)
                    org_name = org_raw if ref_kind == "name" else _trim_for_model_field(
                        Organization, "name", row.get("organization") or ""
                    )
                    if not org_name and ref_kind == "inn":
                        org_name = f"Организация {ref_value}"
                    inn = inn_from_col or (ref_value if ref_kind == "inn" else "")
                    short_name = _trim_for_model_field(Organization, "short_name", row.get("short_name"))

                    row_ot_raw = row.get("org_type")
                    cell_ot_text = _normalize_cell(row_ot_raw) if row_ot_raw is not None else ""
                    row_resolved_ot = _resolve_org_type_from_cell(row_ot_raw)
                    if cell_ot_text and row_resolved_ot is None:
                        errors.append(
                            f"Строка {line_no}: неизвестный тип организации «{cell_ot_text}»"
                        )
                    effective_org_type = row_resolved_ot if row_resolved_ot else default_org_type

                    row_tag_objs, unk_org_tags = _resolve_tags_by_names(
                        row.get("organization_tags"), for_organization=True
                    )
                    for u in unk_org_tags:
                        errors.append(f"Строка {line_no}: неизвестный тег «{u}»")
                    merged_tag_by_id = {t.id: t for t in imported_tags}
                    for t in row_tag_objs:
                        merged_tag_by_id[t.id] = t
                    merged_tag_objs = list(merged_tag_by_id.values())

                    organization = found
                    before = None
                    is_new = False
                    is_branch = effective_org_type == Organization.OrgType.COMPANY_BRANCH

                    parent_for_branch = None
                    if is_branch:
                        parent_for_branch, p_ok = _resolve_import_parent_organization(
                            row, line_no, errors, existing_org=found
                        )
                        if not p_ok:
                            skipped += 1
                            continue

                    if organization is None:
                        if is_branch:
                            org_name_final = (
                                _trim_for_model_field(
                                    Organization, "name", row.get("organization") or ""
                                )
                                or org_raw
                            )
                            if not org_name_final:
                                skipped += 1
                                errors.append(
                                    f"Строка {line_no}: укажите наименование подразделения"
                                )
                                continue
                            organization = Organization(
                                name=org_name_final,
                                short_name=short_name or "",
                                inn=None,
                                org_type=effective_org_type,
                                parent_organization=parent_for_branch,
                            )
                            is_new = True
                        elif not inn:
                            skipped += 1
                            errors.append(
                                f"Строка {line_no}: не удалось создать организацию без ИНН ({org_raw}); "
                                f"для подразделения задайте тип «подразделение»/«company_branch» "
                                f"и колонку «Головная организация» (ИНН или наименование юрлица)"
                            )
                            continue
                        else:
                            if not org_name:
                                org_name = f"Организация {inn}"
                            organization = Organization(
                                name=org_name,
                                short_name=short_name or "",
                                inn=inn,
                                org_type=effective_org_type,
                            )
                            is_new = True
                    else:
                        before = _capture_organization_state(organization)
                        if not update_existing:
                            skipped += 1
                            continue

                    region = _resolve_region(row.get("region"))
                    description = _normalize_cell(row.get("comment") or row.get("description"))

                    if org_name:
                        organization.name = org_name
                    if inn and not is_branch:
                        organization.inn = inn
                    if is_branch:
                        organization.inn = None
                        organization.parent_organization = parent_for_branch
                    organization.org_type = effective_org_type
                    if region:
                        organization.region = region
                    if description:
                        organization.description = description
                    if short_name:
                        organization.short_name = short_name

                    organization.save()

                    if merged_tag_objs:
                        merged_ids = set(organization.tags.values_list("id", flat=True))
                        merged_ids.update(t.id for t in merged_tag_objs)
                        organization.tags.set(sorted(merged_ids))

                    after = _capture_organization_state(organization)
                    _create_field_change_rows(
                        before=before,
                        after=after,
                        fields=_ORGANIZATION_AUDIT_FIELDS,
                        source=source,
                        changed_by=actor,
                        organization=organization,
                    )

                    if is_new:
                        created += 1
                        _track_import_record(
                            batch,
                            organization=organization,
                            action=ImportBatchRecord.Action.CREATED,
                        )
                    else:
                        updated += 1
                        _track_import_record(
                            batch,
                            organization=organization,
                            action=ImportBatchRecord.Action.UPDATED,
                            before=before,
                        )
            except Exception as exc:
                skipped += 1
                errors.append(f"Строка {line_no}: {exc}")

        batch.created_count = created
        batch.updated_count = updated
        batch.skipped_count = skipped
        batch.save(update_fields=["created_count", "updated_count", "skipped_count"])

        return Response(
            {
                "batch_id": batch.id,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "total_rows": len(rows),
                "errors": errors[:200],
            }
        )

    @action(detail=False, methods=["get"], url_path="import-xlsx-template")
    def import_xlsx_template(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Организации"
        ws.append(
            [
                "Наименование",
                "Краткое наименование",
                "ИНН",
                "Головная организация",
                "Регион",
                "Описание",
                "Тип организации",
                "Теги",
            ]
        )
        ws.append(
            [
                'ГБУЗ "Городская поликлиника № 1" г. Москвы',
                "ГП №1",
                "7701234567",
                "",
                "Москва",
                "Пример строки — удалите или замените своими данными",
                "коммерческая",
                "",
            ]
        )
        ws.append(
            [
                'Филиал ГБУЗ "Городская поликлиника № 1" (пример подразделения)',
                "Филиал пример",
                "",
                "7701234567",
                "Москва",
                "ИНН пустой: строка типа подразделение, головная — по ИНН выше",
                "подразделение",
                "",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="organizations_import_template.xlsx"'
        return resp


class OrganizationInteractionViewSet(viewsets.ModelViewSet):
    serializer_class = OrganizationInteractionSerializer
    filterset_fields = ["organization", "interaction_type", "user", "project", "acting_organization"]

    def get_queryset(self):
        return OrganizationInteraction.objects.select_related(
            "organization", "user"
        )

    def perform_create(self, serializer):
        data = {"user": self.request.user}
        if not serializer.validated_data.get("acting_organization"):
            primary = UserActingOrganization.objects.filter(
                user=self.request.user,
                is_primary=True,
            ).select_related("organization").first()
            if primary:
                data["acting_organization"] = primary.organization
        serializer.save(**data)


class ContactViewSet(viewsets.ModelViewSet):
    serializer_class = ContactSerializer
    pagination_class = RegistryPagination
    filterset_fields = ["organization", "type", "current", "is_manager"]
    search_fields = [
        "first_name",
        "last_name",
        "middle_name",
        "position",
        "department_name",
        "phone",
        "phone_extension",
        "email",
    ]

    def get_queryset(self):
        qs = Contact.objects.select_related("organization").prefetch_related("tags")
        org_name = self.request.query_params.get("organization_name")
        if org_name:
            qs = qs.filter(organization__name__icontains=org_name)

        tag_ids = self.request.query_params.get("tags")
        if tag_ids:
            ids = [int(x) for x in tag_ids.split(",") if x.strip().isdigit()]
            if ids:
                qs = qs.filter(tags__id__in=ids).distinct()

        return qs

    @action(detail=True, methods=["get"], url_path="change-log")
    def change_log(self, request, pk=None):
        contact = self.get_object()
        qs = (
            EntityFieldChange.objects.filter(contact=contact)
            .select_related("changed_by", "contact", "organization")
            .order_by("-changed_at", "-id")
        )
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = EntityFieldChangeSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = EntityFieldChangeSerializer(qs, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        contact = serializer.save()
        after = _capture_contact_state(contact)
        _create_field_change_rows(
            before=None,
            after=after,
            fields=_CONTACT_AUDIT_FIELDS,
            source=_normalize_change_source(self.request.query_params.get("source")),
            changed_by=self.request.user if self.request.user.is_authenticated else None,
            organization=contact.organization,
            contact=contact,
        )

    def perform_update(self, serializer):
        before = _capture_contact_state(serializer.instance)
        contact = serializer.save()
        after = _capture_contact_state(contact)
        _create_field_change_rows(
            before=before,
            after=after,
            fields=_CONTACT_AUDIT_FIELDS,
            source=_normalize_change_source(self.request.query_params.get("source")),
            changed_by=self.request.user if self.request.user.is_authenticated else None,
            organization=contact.organization,
            contact=contact,
        )

    @action(detail=False, methods=["post"], url_path="import-xlsx")
    def import_xlsx(self, request):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Файл не передан"}, status=status.HTTP_400_BAD_REQUEST)
        if not file_obj.name.lower().endswith(".xlsx"):
            return Response({"detail": "Поддерживаются только .xlsx файлы"}, status=status.HTTP_400_BAD_REQUEST)

        source = _normalize_change_source(
            _form_data_scalar(request.data, "source") or EntityFieldChange.Source.BULK
        )
        default_org_type = _form_data_scalar(request.data, "default_org_type") or Organization.OrgType.OTHER
        if default_org_type not in {item.value for item in Organization.OrgType}:
            default_org_type = Organization.OrgType.OTHER
        default_contact_type = (
            _form_data_scalar(request.data, "default_contact_type") or Contact.ContactType.PERSON
        )
        if default_contact_type not in {item.value for item in Contact.ContactType}:
            default_contact_type = Contact.ContactType.PERSON

        create_missing_orgs = _parse_bool(
            _form_data_scalar(request.data, "create_missing_organizations"), True
        )
        org_tag_ids = _extract_tag_ids(_form_data_scalar(request.data, "organization_tag_ids"))
        contact_tag_ids = _extract_tag_ids(_form_data_scalar(request.data, "contact_tag_ids"))
        org_tags = list(OrganizationTag.objects.filter(id__in=org_tag_ids))
        contact_tags = list(OrganizationTag.objects.filter(id__in=contact_tag_ids))

        try:
            rows = _xlsx_rows(file_obj, import_kind="contacts")
        except Exception as exc:
            logger.exception("contacts import_xlsx: не удалось разобрать файл")
            return Response(
                {"detail": f"Не удалось прочитать Excel (.xlsx): {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        actor = request.user if request.user.is_authenticated else None
        batch = ImportBatch.objects.create(
            entity_type=ImportBatch.EntityType.CONTACTS,
            file_name=file_obj.name,
            uploaded_by=actor,
            total_rows=len(rows),
        )

        created = 0
        updated = 0
        skipped = 0
        errors = []
        last_org_ref = ""
        last_contact_by_org = {}

        for line_no, row in rows:
            org_raw = _normalize_cell(row.get("organization"))
            if not org_raw:
                org_raw = last_org_ref
            if not org_raw:
                skipped += 1
                continue
            last_org_ref = org_raw

            try:
                with transaction.atomic():
                    row_ot_raw_early = row.get("org_type")
                    cell_ot_early = (
                        _normalize_cell(row_ot_raw_early) if row_ot_raw_early is not None else ""
                    )
                    row_resolved_ot_early = _resolve_org_type_from_cell(row_ot_raw_early)
                    if cell_ot_early and row_resolved_ot_early is None:
                        errors.append(
                            f"Строка {line_no}: неизвестный тип организации «{cell_ot_early}»"
                        )
                    effective_new_org_type = (
                        row_resolved_ot_early if row_resolved_ot_early else default_org_type
                    )

                    organization, ref_kind, ref_value = _find_organization_for_import(row, org_raw)
                    if organization is None and create_missing_orgs:
                        inn_for_create = _normalize_inn(row.get("inn")) or (
                            ref_value if ref_kind == "inn" else _normalize_inn(org_raw)
                        )
                        if inn_for_create:
                            region = _resolve_region(row.get("region"))
                            short_for_org = _normalize_cell(row.get("short_name"))
                            organization = Organization.objects.create(
                                name=f"Организация {inn_for_create}",
                                short_name=short_for_org,
                                inn=inn_for_create,
                                org_type=effective_new_org_type,
                                region=region,
                            )
                            org_after = _capture_organization_state(organization)
                            _create_field_change_rows(
                                before=None,
                                after=org_after,
                                fields=_ORGANIZATION_AUDIT_FIELDS,
                                source=source,
                                changed_by=actor,
                                organization=organization,
                            )
                            _track_import_record(
                                batch,
                                organization=organization,
                                action=ImportBatchRecord.Action.CREATED,
                            )

                    if organization is None:
                        errors.append(f"Строка {line_no}: организация не найдена ({org_raw})")
                        skipped += 1
                        continue

                    if row_resolved_ot_early:
                        organization.org_type = row_resolved_ot_early
                        organization.save(update_fields=["org_type"])

                    row_org_tags, unk_org_tags = _resolve_tags_by_names(
                        row.get("organization_tags"), for_organization=True
                    )
                    for u in unk_org_tags:
                        errors.append(
                            f"Строка {line_no}: неизвестный тег организации «{u}»"
                        )

                    merged_org_tag_ids = set(organization.tags.values_list("id", flat=True))
                    merged_org_tag_ids.update(t.id for t in org_tags)
                    merged_org_tag_ids.update(t.id for t in row_org_tags)
                    if merged_org_tag_ids:
                        organization.tags.set(sorted(merged_org_tag_ids))

                    row_contact_tags, unk_contact_tags = _resolve_tags_by_names(
                        row.get("contact_tags"), for_organization=False
                    )
                    for u in unk_contact_tags:
                        errors.append(f"Строка {line_no}: неизвестный тег контакта «{u}»")

                    comment = _normalize_cell(row.get("comment"))
                    position = _trim_for_model_field(Contact, "position", row.get("position"))
                    ext_from_col = _trim_for_model_field(
                        Contact, "phone_extension", row.get("phone_extension")
                    )
                    phone_cell = row.get("phone")
                    if ext_from_col:
                        phone = _trim_for_model_field(Contact, "phone", phone_cell)
                        phone_extension = ext_from_col
                    else:
                        phone_main, ext_parsed = _split_phone_extension_from_combined(phone_cell)
                        phone = _trim_for_model_field(Contact, "phone", phone_main)
                        phone_extension = _trim_for_model_field(
                            Contact, "phone_extension", ext_parsed
                        )
                    email = _trim_for_model_field(Contact, "email", row.get("email"))
                    messenger_link = _normalize_cell(row.get("messenger_link"))
                    messenger_type = _normalize_cell(row.get("messenger_type"))
                    messenger_value = ""
                    if messenger_link and messenger_type:
                        messenger_value = f"{messenger_type}: {messenger_link}"
                    elif messenger_link:
                        messenger_value = messenger_link
                    elif messenger_type:
                        messenger_value = messenger_type

                    fio = _normalize_cell(row.get("full_name"))
                    last_name, first_name, middle_name = _parse_person_name(fio)
                    department_name = _trim_for_model_field(
                        Contact, "department_name", row.get("department_name")
                    )
                    contact_type = default_contact_type
                    if not fio and department_name:
                        contact_type = Contact.ContactType.DEPARTMENT

                    base_qs = Contact.objects.filter(organization=organization)
                    existing = None
                    if not fio and not department_name and organization.id in last_contact_by_org:
                        existing = Contact.objects.filter(
                            id=last_contact_by_org[organization.id]
                        ).first()
                    elif contact_type == Contact.ContactType.DEPARTMENT and department_name:
                        existing = base_qs.filter(
                            type=Contact.ContactType.DEPARTMENT,
                            department_name__iexact=department_name,
                        ).first()
                    elif any([last_name, first_name, middle_name]):
                        existing = base_qs.filter(
                            type=contact_type,
                            last_name__iexact=last_name,
                            first_name__iexact=first_name,
                            middle_name__iexact=middle_name,
                        ).first()
                    elif phone:
                        existing = base_qs.filter(phone=phone).first()

                    contact = existing
                    before = _capture_contact_state(existing) if existing else None
                    is_new = existing is None
                    if contact is None:
                        contact = Contact(
                            organization=organization,
                            type=contact_type,
                        )

                    if any([last_name, first_name, middle_name]):
                        contact.last_name = _trim_for_model_field(Contact, "last_name", last_name)
                        contact.first_name = _trim_for_model_field(Contact, "first_name", first_name)
                        contact.middle_name = _trim_for_model_field(Contact, "middle_name", middle_name)
                    if department_name:
                        contact.department_name = department_name
                    if position:
                        contact.position = position
                    if comment:
                        contact.comment = comment
                    if phone:
                        contact.phone = phone
                    if phone_extension:
                        contact.phone_extension = phone_extension
                    if email:
                        contact.email = email
                    if messenger_value:
                        contact.messenger = _trim_for_model_field(Contact, "messenger", messenger_value)
                    contact.type = contact_type

                    if not any(
                        [
                            contact.last_name,
                            contact.first_name,
                            contact.middle_name,
                            contact.department_name,
                            contact.phone,
                            contact.phone_extension,
                            contact.email,
                            contact.messenger,
                        ]
                    ):
                        skipped += 1
                        continue

                    contact.save()

                    merged_contact_tag_objs = list(
                        {t.id: t for t in contact_tags + row_contact_tags}.values()
                    )
                    if merged_contact_tag_objs:
                        merged_contact_tag_ids = set(
                            contact.tags.values_list("id", flat=True)
                        )
                        merged_contact_tag_ids.update(t.id for t in merged_contact_tag_objs)
                        contact.tags.set(sorted(merged_contact_tag_ids))

                    after = _capture_contact_state(contact)
                    _create_field_change_rows(
                        before=before,
                        after=after,
                        fields=_CONTACT_AUDIT_FIELDS,
                        source=source,
                        changed_by=actor,
                        organization=organization,
                        contact=contact,
                    )
                    if is_new:
                        created += 1
                        _track_import_record(
                            batch,
                            contact=contact,
                            action=ImportBatchRecord.Action.CREATED,
                        )
                    else:
                        updated += 1
                        _track_import_record(
                            batch,
                            contact=contact,
                            action=ImportBatchRecord.Action.UPDATED,
                            before=before,
                        )
                    last_contact_by_org[organization.id] = contact.id
            except Exception as exc:
                skipped += 1
                errors.append(f"Строка {line_no}: {exc}")

        batch.created_count = created
        batch.updated_count = updated
        batch.skipped_count = skipped
        batch.save(update_fields=["created_count", "updated_count", "skipped_count"])

        return Response(
            {
                "batch_id": batch.id,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "total_rows": len(rows),
                "errors": errors[:200],
            }
        )

    @action(detail=False, methods=["get"], url_path="import-xlsx-template")
    def import_xlsx_template(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Контакты"
        ws.append(
            [
                "Организация",
                "Краткое наименование",
                "ФИО",
                "Должность",
                "Телефон",
                "Добавочный",
                "Email",
                "Регион",
                "Описание",
                "Тип организации",
                "Теги организации",
                "Теги контакта",
            ]
        )
        ws.append(
            [
                "7701234567",
                "ГП №1",
                "Иванов Иван Иванович",
                "Главный врач",
                "+7 495 000-00-00 доб. 123",
                "",
                "ivanov@example.test",
                "Москва",
                "В одной ячейке с телефоном или отдельной колонкой",
                "муниципальная",
                "",
                "",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="contacts_import_template.xlsx"'
        return resp


class ImportBatchViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ImportBatchSerializer
    pagination_class = RegistryPagination
    filterset_fields = ["entity_type", "status"]

    def get_queryset(self):
        return ImportBatch.objects.select_related("uploaded_by").order_by("-uploaded_at", "-id")

    @action(detail=True, methods=["post"], url_path="rollback")
    def rollback(self, request, pk=None):
        batch = self.get_object()
        payload, code = _rollback_import_batch(
            batch,
            request.user if request.user.is_authenticated else None,
        )
        return Response(payload, status=code)


class OrganizationTagViewSet(viewsets.ModelViewSet):
    serializer_class = OrganizationTagSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["slug", "category"]
    search_fields = ["name", "slug", "category"]

    def get_queryset(self):
        qs = OrganizationTag.objects.all()
        tag_type = self.request.query_params.get("tag_type")
        if tag_type and tag_type != OrganizationTag.TagType.ALL:
            qs = qs.filter(tag_type__in=[tag_type, OrganizationTag.TagType.ALL])
        return qs.order_by("category", "name")

    @staticmethod
    def _build_unique_slug(name: str, *, exclude_id: int | None = None) -> str:
        base = slugify(name or "", allow_unicode=True)
        if not base:
            base = "tag"
        slug = base
        idx = 2
        qs = OrganizationTag.objects.all()
        if exclude_id is not None:
            qs = qs.exclude(id=exclude_id)
        while qs.filter(slug=slug).exists():
            slug = f"{base}-{idx}"
            idx += 1
        return slug

    def perform_create(self, serializer):
        raw_slug = (serializer.validated_data.get("slug") or "").strip()
        slug = raw_slug or self._build_unique_slug(serializer.validated_data.get("name", ""))
        serializer.save(slug=slug)

    def perform_update(self, serializer):
        raw_slug = serializer.validated_data.get("slug")
        if raw_slug is not None and not str(raw_slug).strip():
            slug = self._build_unique_slug(
                serializer.validated_data.get("name", serializer.instance.name),
                exclude_id=serializer.instance.id,
            )
            serializer.save(slug=slug)
            return
        serializer.save()


class ProjectViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["year"]
    search_fields = ["name", "code"]

    def get_queryset(self):
        return Project.objects.prefetch_related("memberships__organization")


class ProjectOrganizationMembershipViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectOrganizationMembershipSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["project", "organization", "role"]

    def get_queryset(self):
        return ProjectOrganizationMembership.objects.select_related("project", "organization")


class UserActingOrganizationViewSet(viewsets.ModelViewSet):
    serializer_class = UserActingOrganizationSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["user", "organization", "is_primary"]

    def get_queryset(self):
        qs = UserActingOrganization.objects.select_related("user", "organization")
        if getattr(self.request.user, "is_admin_role", False):
            return qs
        return qs.filter(user=self.request.user)

    def create(self, request, *args, **kwargs):
        users_raw = request.data.get("users")
        if users_raw is None:
            return super().create(request, *args, **kwargs)

        org_id = request.data.get("organization")
        if not org_id:
            return Response(
                {"organization": "Поле organization обязательно"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if isinstance(users_raw, (list, tuple)):
            user_ids = [int(x) for x in users_raw if str(x).strip().isdigit()]
        else:
            user_ids = [
                int(x)
                for x in re.split(r"[,;\s]+", str(users_raw))
                if str(x).strip().isdigit()
            ]
        if not user_ids:
            return Response(
                {"users": "Передайте хотя бы одного пользователя"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if (
            not getattr(request.user, "is_admin_role", False)
            and any(uid != request.user.id for uid in user_ids)
        ):
            return Response(
                {"detail": "Недостаточно прав для назначения других пользователей"},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            organization = Organization.objects.get(id=int(org_id))
        except (Organization.DoesNotExist, ValueError, TypeError):
            return Response(
                {"organization": "Организация не найдена"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        primary_user_id_raw = request.data.get("primary_user_id")
        primary_user_id = (
            int(primary_user_id_raw)
            if str(primary_user_id_raw or "").isdigit()
            else None
        )
        mark_primary = _parse_bool(request.data.get("is_primary"), False)

        created_rows = []
        with transaction.atomic():
            for user_id in user_ids:
                row, _ = UserActingOrganization.objects.get_or_create(
                    user_id=user_id,
                    organization=organization,
                    defaults={"is_primary": False},
                )
                if mark_primary and (primary_user_id is None or primary_user_id == user_id):
                    UserActingOrganization.objects.filter(
                        user_id=user_id, is_primary=True
                    ).exclude(id=row.id).update(is_primary=False)
                    if not row.is_primary:
                        row.is_primary = True
                        row.save(update_fields=["is_primary"])
                created_rows.append(row)

        serializer = self.get_serializer(created_rows, many=True)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        with transaction.atomic():
            user = serializer.validated_data["user"]
            make_primary = serializer.validated_data.get("is_primary") or not UserActingOrganization.objects.filter(
                user=user
            ).exists()
            if make_primary:
                UserActingOrganization.objects.filter(
                    user=user,
                    is_primary=True,
                ).update(is_primary=False)
            serializer.save(is_primary=make_primary)

    def perform_update(self, serializer):
        with transaction.atomic():
            if serializer.validated_data.get("is_primary"):
                UserActingOrganization.objects.filter(
                    user=serializer.instance.user,
                    is_primary=True,
                ).exclude(id=serializer.instance.id).update(is_primary=False)
            serializer.save()


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def me_acting_organizations(request):
    rows = UserActingOrganization.objects.filter(user=request.user).select_related("organization")
    if not rows.exists():
        org = _ensure_default_acting_organization()
        UserActingOrganization.objects.get_or_create(
            user=request.user,
            organization=org,
            defaults={"is_primary": True},
        )
        rows = UserActingOrganization.objects.filter(user=request.user).select_related(
            "organization"
        )
    return Response(UserActingOrganizationSerializer(rows, many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def communication_history(request):
    """
    Local communications history for autonomous Matrix mode.
    """
    qs = OrganizationInteraction.objects.select_related(
        "organization", "project", "acting_organization", "user"
    )
    project = request.query_params.get("project")
    if project and project.isdigit():
        qs = qs.filter(project_id=int(project))
    organization = request.query_params.get("organization")
    if organization and organization.isdigit():
        qs = qs.filter(organization_id=int(organization))
    acting_organization = request.query_params.get("acting_organization")
    if acting_organization and acting_organization.isdigit():
        qs = qs.filter(acting_organization_id=int(acting_organization))
    date_from = request.query_params.get("date_from")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    date_to = request.query_params.get("date_to")
    if date_to:
        qs = qs.filter(date__lte=date_to)

    payload = [
        {
            "id": row.id,
            "source": "matrix",
            "occurred_at": row.date,
            "type": row.interaction_type,
            "type_display": row.get_interaction_type_display(),
            "organization_id": row.organization_id,
            "organization_name": row.organization.name if row.organization else None,
            "acting_organization_id": row.acting_organization_id,
            "acting_organization_name": row.acting_organization.name if row.acting_organization else None,
            "project_id": row.project_id,
            "project_name": row.project.name if row.project else None,
            "summary": row.notes or "",
            "manager_name": str(row.user) if row.user else None,
            "created_at": row.created_at,
        }
        for row in qs.order_by("-date", "-created_at")[:1000]
    ]
    return Response(payload)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_contacts(request):
    """Proxy to Bitrix contacts API with filtering."""
    params = {}
    for key in ("organization", "organization__contains", "type",
                "department", "department__contains", "manager", "current"):
        val = request.query_params.get(key)
        if val:
            params[key] = val
    try:
        data, _ = _bitrix_request_first_available(
            BITRIX_CONTACTS_LIST_ENDPOINTS,
            params=params,
        )
        return Response(data if isinstance(data, list) else [])
    except Exception as exc:
        return _bitrix_error_response(exc, "Ошибка при обращении к внешнему API контактов")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def external_contact_add(request):
    """Create contact in external Bitrix API; опционально дублирует в локальный Contact (sync_local=true)."""
    sync_local = _parse_bool(request.data.get("sync_local"), True)
    payload = _strip_sync_local_from_payload(request.data)
    try:
        data, upstream_status = _bitrix_request_first_available(
            BITRIX_CONTACT_ADD_ENDPOINTS,
            method="POST",
            payload=payload,
        )
        out = data if isinstance(data, dict) else {"result": data}
        org_inn = payload.get("organization") if isinstance(payload, dict) else None
        if (
            sync_local
            and upstream_status in (status.HTTP_200_OK, status.HTTP_201_CREATED)
            and org_inn
            and isinstance(data, dict)
        ):
            lc, st = _mirror_bitrix_contact_to_local(
                org_inn, data, payload, sync_local=True
            )
            out = dict(out)
            out["local_contact"] = ContactSerializer(lc).data if lc else None
            out["local_sync"] = st
        return Response(out, status=upstream_status)
    except Exception as exc:
        return _bitrix_error_response(exc, "Ошибка при создании контакта во внешнем API")


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def external_contact_update(request):
    """Update contact in external Bitrix API; опционально обновляет локальный Contact (sync_local=true)."""
    sync_local = _parse_bool(request.data.get("sync_local"), True)
    payload = _strip_sync_local_from_payload(request.data)
    try:
        data, upstream_status = _bitrix_request_first_available(
            BITRIX_CONTACT_UPDATE_ENDPOINTS,
            method="PATCH",
            payload=payload,
        )
        out = dict(data) if isinstance(data, dict) else {"result": data}
        org_inn = None
        if isinstance(data, dict):
            org_inn = data.get("organization")
        if not org_inn and isinstance(payload, dict):
            org_inn = payload.get("organization")
        if (
            sync_local
            and upstream_status == status.HTTP_200_OK
            and org_inn
            and isinstance(data, dict)
        ):
            lc, st = _mirror_bitrix_contact_to_local(
                str(org_inn), data, payload, sync_local=True
            )
            out["local_contact"] = ContactSerializer(lc).data if lc else None
            out["local_sync"] = st
        return Response(out, status=upstream_status)
    except Exception as exc:
        return _bitrix_error_response(exc, "Ошибка при изменении контакта во внешнем API")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_contact_history(request):
    """Get contact history from external Bitrix API."""
    contact_id = request.query_params.get("id")
    if not contact_id:
        return Response({"id": "Параметр id обязателен"}, status=status.HTTP_400_BAD_REQUEST)
    params = {"id": contact_id}
    for key in ("page", "page_size"):
        val = request.query_params.get(key)
        if val:
            params[key] = val
    try:
        data, _ = _bitrix_request_first_available(
            BITRIX_CONTACT_HISTORY_ENDPOINTS,
            params=params,
        )
        return Response(data)
    except Exception as exc:
        return _bitrix_error_response(exc, "Ошибка при получении истории контакта из внешнего API")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_communications(request):
    """Proxy to Bitrix communication interactions list endpoint."""
    params = {}
    for key in (
        "organization",
        "counterparty_inn",
        "contact_id",
        "our_organization",
        "project",
        "channel",
        "occurred_after",
        "occurred_before",
        "page",
        "page_size",
    ):
        val = request.query_params.get(key)
        if val:
            params[key] = val
    try:
        data, _ = _bitrix_request_first_available(
            BITRIX_COMMUNICATION_LIST_ENDPOINTS,
            params=params,
        )
        return Response(data)
    except Exception as exc:
        return _bitrix_error_response(exc, "Ошибка при получении журнала коммуникаций из внешнего API")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def external_communication_add(request):
    """Create communication interaction in external Bitrix API."""
    try:
        data, upstream_status = _bitrix_request_first_available(
            BITRIX_COMMUNICATION_ADD_ENDPOINTS,
            method="POST",
            payload=request.data,
        )
        return Response(data, status=upstream_status)
    except Exception as exc:
        return _bitrix_error_response(exc, "Ошибка при создании коммуникации во внешнем API")


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def external_communication_update(request):
    """Update communication interaction in external Bitrix API."""
    try:
        data, upstream_status = _bitrix_request_first_available(
            BITRIX_COMMUNICATION_UPDATE_ENDPOINTS,
            method="PATCH",
            payload=request.data,
        )
        return Response(data, status=upstream_status)
    except Exception as exc:
        return _bitrix_error_response(exc, "Ошибка при изменении коммуникации во внешнем API")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def sync_user_as_our_side_contact(request):
    """
    Create Bitrix contact(s) for Matrix user in one or more our-side organizations.

    В Bitrix один контакт привязан к одной организации; один сотрудник «от лица»
    нескольких юрлиц создаётся как несколько контактов.

    Body:
    - user_id (optional, default = current user)
    - organization_inns: список ИНН или строка "inn1,inn2" (предпочтительно при выборе)
    - organization_inn: один ИНН (обратная совместимость)
    - manager, position, comment (optional)
    - sync_local (optional, default true) — дублировать в локальный Contact при наличии Organization с тем же ИНН
    Если организации не переданы — подставляются ИНН из Bitrix API (организации с is_our_side=true),
    затем при необходимости BITRIX_OUR_ORGANIZATION_INNS / BITRIX_OUR_ORGANIZATION_INN.
    """
    user_id = request.data.get("user_id") or request.user.id
    sync_local = _parse_bool(request.data.get("sync_local"), True)

    inns, err = _collect_organization_inns_for_sync(request)
    if err:
        return Response({"detail": err}, status=status.HTTP_400_BAD_REQUEST)

    UserModel = get_user_model()
    user = UserModel.objects.filter(id=user_id, is_active=True).first()
    if user is None:
        return Response({"user_id": "Пользователь не найден"}, status=status.HTTP_404_NOT_FOUND)

    base_comment = request.data.get(
        "comment",
        f"Синхронизировано из Matrix: user_id={user.id}, username={user.username}",
    )

    results = []
    for organization_inn in inns:
        payload = {
            "organization": organization_inn,
            "type": "person",
            "first_name": user.first_name or "",
            "last_name": user.last_name or user.username,
            "middle_name": getattr(user, "patronymic", "") or "",
            "position": request.data.get("position", ""),
            "manager": _parse_bool(request.data.get("manager"), False),
            "comment": base_comment,
            "current": True,
        }
        try:
            data, upstream_status = _bitrix_request_first_available(
                BITRIX_CONTACT_ADD_ENDPOINTS,
                method="POST",
                payload=payload,
            )
            row = {
                "organization_inn": organization_inn,
                "ok": True,
                "status_code": upstream_status,
                "external_contact": data,
            }
            if sync_local and upstream_status in (
                status.HTTP_200_OK,
                status.HTTP_201_CREATED,
            ):
                lc, st = _mirror_bitrix_contact_to_local(
                    organization_inn,
                    data,
                    {
                        **payload,
                        "is_manager": payload.get("manager", False)
                        if isinstance(payload.get("manager"), bool)
                        else _parse_bool(payload.get("manager"), False),
                    },
                    sync_local=True,
                )
                row["local_contact"] = (
                    ContactSerializer(lc).data if lc else None
                )
                row["local_sync"] = st
            elif not sync_local:
                row["local_sync"] = "off"
            results.append(row)
        except Exception as exc:  # noqa: BLE001
            err_body = _bitrix_error_dict(exc)
            results.append(
                {
                    "organization_inn": organization_inn,
                    "ok": False,
                    **err_body,
                }
            )

    all_ok = all(r.get("ok") for r in results)
    any_ok = any(r.get("ok") for r in results)
    # 207 — частичный успех (не все ИНН прошли); literal для совместимости со старыми DRF
    response_status = status.HTTP_200_OK if all_ok else (207 if any_ok else status.HTTP_502_BAD_GATEWAY)

    return Response(
        {
            "user_id": user.id,
            "organization_inns": inns,
            "results": results,
        },
        status=response_status,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def sync_external_contacts(request):
    """Import contacts from external API response into local DB."""
    contacts_list = request.data.get("contacts", [])
    org_name = request.data.get("organization_name", "")
    if not org_name:
        return Response(
            {"detail": "organization_name обязателен"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    org = Organization.objects.filter(name__icontains=org_name).first()
    if not org:
        return Response(
            {"detail": f"Организация '{org_name}' не найдена"},
            status=status.HTTP_404_NOT_FOUND,
        )

    created = 0
    for ext in contacts_list:
        c_type = ext.get("type", "other")
        defaults = {
            "comment": ext.get("comment", ""),
            "current": ext.get("current", True),
            "first_name": ext.get("first_name", ""),
            "last_name": ext.get("last_name", ""),
            "middle_name": ext.get("middle_name", ""),
            "position": ext.get("position", ""),
            "is_manager": ext.get("manager", False),
            "department_name": ext.get("department_name", ""),
        }
        if c_type == "person":
            Contact.objects.get_or_create(
                organization=org,
                type=c_type,
                first_name=defaults["first_name"],
                last_name=defaults["last_name"],
                middle_name=defaults["middle_name"],
                defaults=defaults,
            )
        elif c_type == "department":
            Contact.objects.get_or_create(
                organization=org,
                type=c_type,
                department_name=defaults["department_name"],
                defaults=defaults,
            )
        else:
            Contact.objects.create(organization=org, type=c_type, **defaults)
        created += 1

    return Response({"synced": created})


def _bitrix_request(endpoint, params=None, method="GET", payload=None, allow_404=False):
    """Helper: make authenticated request to Bitrix API."""
    base_url = f"{settings.BITRIX_API_BASE_URL}{endpoint}"
    auth_header = f"Token {settings.BITRIX_API_TOKEN}"
    conn = BitrixOAuthConnection.objects.filter(is_active=True).order_by("-updated_at").first()
    if conn and conn.access_token:
        if conn.expires_at and conn.expires_at <= timezone.now():
            refreshed = _refresh_bitrix_oauth_token(conn)
            if refreshed:
                conn = refreshed
        if conn.access_token:
            auth_header = f"Bearer {conn.access_token}"
    headers = {
        "Authorization": auth_header,
        "Accept": "application/json",
    }
    try:
        resp = http_requests.request(
            method=method,
            url=base_url,
            headers=headers,
            params=params,
            json=payload,
            timeout=30,
        )
        if resp.status_code == 404:
            logger.warning("Bitrix 404 for %s — body: %s", base_url, resp.text[:200])
            if allow_404:
                return None, resp.status_code
        resp.raise_for_status()
        return _safe_json(resp), resp.status_code
    except http_requests.RequestException as exc:
        logger.error("Bitrix API error: %s", exc)
        raise


def _refresh_bitrix_oauth_token(conn: BitrixOAuthConnection):
    refresh_url = getattr(settings, "BITRIX_OAUTH_TOKEN_URL", "") or ""
    if not (refresh_url and conn.refresh_token):
        return conn
    client_id = conn.client_id or getattr(settings, "BITRIX_CLIENT_ID", "")
    client_secret = conn.client_secret or getattr(settings, "BITRIX_CLIENT_SECRET", "")
    if not (client_id and client_secret):
        return conn
    try:
        resp = http_requests.post(
            refresh_url,
            data={
                "grant_type": "refresh_token",
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": conn.refresh_token,
            },
            timeout=20,
        )
        resp.raise_for_status()
        data = resp.json()
        conn.access_token = data.get("access_token", conn.access_token)
        conn.refresh_token = data.get("refresh_token", conn.refresh_token)
        expires_in = data.get("expires_in")
        if expires_in:
            conn.expires_at = timezone.now() + timedelta(seconds=int(expires_in))
        conn.save(
            update_fields=["access_token", "refresh_token", "expires_at", "updated_at"]
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("Bitrix OAuth refresh failed: %s", exc)
    return conn


def _bitrix_request_first_available(endpoints, method="GET", params=None, payload=None):
    """
    Try API endpoints one-by-one to support both legacy and new path prefixes.
    Returns tuple(data, status_code) from first non-404 response.
    """
    for endpoint in endpoints:
        try:
            data, status_code = _bitrix_request(
                endpoint,
                params=params,
                method=method,
                payload=payload,
                allow_404=True,
            )
            if status_code == 404:
                continue
            return data, status_code
        except http_requests.HTTPError:
            # 4xx/5xx other than 404 should be returned to the client as-is.
            raise
        except Exception:
            # Try next path variant (network issue on one route/proxy).
            continue
    raise RuntimeError("Не удалось найти рабочий эндпоинт внешнего API")


def _safe_json(resp):
    try:
        return resp.json()
    except ValueError:
        return {"detail": resp.text}


def _parse_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _form_data_scalar(data, key, default=None):
    """Multipart / QueryDict иногда отдаёт list по одному ключу."""
    if not hasattr(data, "get"):
        return default
    raw = data.get(key, default)
    if isinstance(raw, list):
        return raw[0] if raw else default
    return raw


def _parse_inn_list_value(raw):
    """Normalize organization_inns from JSON list/tuple or comma-separated string."""
    if raw is None:
        return []
    if isinstance(raw, (list, tuple)):
        return [str(x).strip() for x in raw if str(x).strip()]
    if isinstance(raw, str):
        return [p.strip() for p in raw.split(",") if p.strip()]
    return [str(raw).strip()] if str(raw).strip() else []


def _unique_inn_list(inns):
    seen = set()
    out = []
    for inn in inns:
        if inn not in seen:
            seen.add(inn)
            out.append(inn)
    return out


def _normalize_bitrix_organization_list_payload(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        if "results" in data and isinstance(data["results"], list):
            return data["results"]
        if "data" in data and isinstance(data["data"], list):
            return data["data"]
    return []


def _bitrix_list_organizations_paginated(extra_params=None):
    """
    Собирает все страницы списка организаций Bitrix (DRF pagination или плоский list).
    """
    extra_params = dict(extra_params or {})
    page_size = int(extra_params.pop("page_size", 200))
    page_size = max(1, min(page_size, 1000))
    page = 1
    all_rows = []
    while True:
        params = dict(extra_params)
        params["page"] = page
        params["page_size"] = page_size
        data, _ = _bitrix_request_first_available(
            BITRIX_ORGANIZATIONS_ENDPOINTS,
            params=params,
        )
        if isinstance(data, dict) and "results" in data:
            chunk = data.get("results") or []
            if not isinstance(chunk, list):
                chunk = []
            all_rows.extend(chunk)
            if data.get("next") and chunk:
                page += 1
                continue
            break
        chunk = _normalize_bitrix_organization_list_payload(data)
        all_rows.extend(chunk)
        break
    return all_rows


def _filter_our_side_org_rows(rows):
    """
    Если в JSON есть is_our_side — оставляем только true.

    Если поля нет: доверяем только коротким ответам (вероятно отфильтрованный список на сервере).
    Длинный список без поля is_our_side отбрасываем — иначе можно принять все организации за «наши».
    """
    if not rows:
        return rows
    if any(isinstance(r, dict) and "is_our_side" in r for r in rows):
        return [r for r in rows if isinstance(r, dict) and r.get("is_our_side") is True]
    if len(rows) > 100:
        logger.warning(
            "Bitrix: много организаций без поля is_our_side в ответе — "
            "список не используется. Добавьте is_our_side в OrganizationSerializer и "
            "BooleanFilter is_our_side в OrganizationFilter (BitrixDashboard)."
        )
        return []
    return rows


def _organization_inns_from_rows(rows):
    out = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        inn = r.get("inn")
        if inn:
            out.append(str(inn).strip())
    return _unique_inn_list(out)


def _fetch_our_side_organization_inns_from_bitrix():
    """ИНН организаций с is_our_side=true из Bitrix API."""
    try:
        rows = _bitrix_list_organizations_paginated(
            {"is_our_side": "true", "page_size": 500}
        )
        rows = _filter_our_side_org_rows(rows)
        return _organization_inns_from_rows(rows)
    except Exception as exc:
        logger.warning("Не удалось получить наши организации из Bitrix API: %s", exc)
        return []


def _collect_organization_inns_for_sync(request):
    """
    Returns (list[str] | None, error_message | None).

    Priority:
    1) organization_inns (обязательно непустой, если ключ передан)
    2) organization_inn (один ИНН)
    3) Bitrix API: список организаций с is_our_side=true
    4) BITRIX_OUR_ORGANIZATION_INNS / BITRIX_OUR_ORGANIZATION_INN (fallback)
    """
    data = request.data
    if "organization_inns" in data:
        inns = _unique_inn_list(_parse_inn_list_value(data.get("organization_inns")))
        if not inns:
            return None, "organization_inns не может быть пустым"
        return inns, None

    if "organization_inn" in data:
        single = (data.get("organization_inn") or "").strip()
        if not single:
            return None, "organization_inn не может быть пустым"
        return [single], None

    api_inns = _fetch_our_side_organization_inns_from_bitrix()
    if api_inns:
        return api_inns, None

    env_multi = (getattr(settings, "BITRIX_OUR_ORGANIZATION_INNS", "") or "").strip()
    if env_multi:
        inns = _unique_inn_list(_parse_inn_list_value(env_multi))
        if inns:
            return inns, None

    single_env = (getattr(settings, "BITRIX_OUR_ORGANIZATION_INN", "") or "").strip()
    if single_env:
        return [single_env], None

    return None, (
        "Не заданы организации: передайте organization_inns / organization_inn, "
        "либо настройте Bitrix (GET .../organization/?is_our_side=true и поле is_our_side в ответе), "
        "либо задайте BITRIX_OUR_ORGANIZATION_INNS / BITRIX_OUR_ORGANIZATION_INN"
    )


def _bitrix_error_dict(exc):
    """Structured error for per-organization results (no Response)."""
    if isinstance(exc, http_requests.HTTPError) and exc.response is not None:
        upstream = exc.response
        try:
            upstream_payload = upstream.json()
        except ValueError:
            upstream_payload = upstream.text[:1000]
        return {
            "status_code": upstream.status_code,
            "upstream_payload": upstream_payload,
        }
    return {
        "status_code": None,
        "detail": str(exc),
    }


def _bitrix_error_response(exc, detail):
    """Build consistent error response for upstream Bitrix failures."""
    if isinstance(exc, http_requests.HTTPError) and exc.response is not None:
        upstream = exc.response
        try:
            upstream_payload = upstream.json()
        except ValueError:
            upstream_payload = upstream.text[:1000]
        upstream_status = upstream.status_code
        status_code = (
            upstream_status
            if upstream_status in {400, 401, 403, 404, 409, 422}
            else status.HTTP_502_BAD_GATEWAY
        )
        return Response(
            {
                "detail": detail,
                "upstream_status": upstream_status,
                "upstream_payload": upstream_payload,
            },
            status=status_code,
        )
    return Response({"detail": detail}, status=status.HTTP_502_BAD_GATEWAY)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_organizations(request):
    """Proxy to Bitrix organizations API with filtering.

    Supports multi-value `regions` and `fed_districts` (comma-separated):
    makes one Bitrix request per value and merges results by name.
    """
    base_params = {}
    single_param_map = {
        "type": "type",
        "region": "region",
        "region__contains": "region__contains",
        "fed_district": "fed_district",
        "fed_district__contains": "fed_district__contains",
        "prof_activity": "prof_activity",
        "prof_activity__contains": "prof_activity__contains",
        "federal": "federal",
        "is_active": "is_active",
        "is_our_side": "is_our_side",
        "project": "project",
        "date": "date",
    }
    for frontend_key, api_key in single_param_map.items():
        val = request.query_params.get(frontend_key)
        if val:
            base_params[api_key] = val

    # Build list of (param_key, value) pairs for multi-value filters
    multi_requests = []
    regions_raw = request.query_params.get("regions", "")
    fed_districts_raw = request.query_params.get("fed_districts", "")
    prof_activities_raw = request.query_params.get("prof_activities", "")
    regions_list = [r.strip() for r in regions_raw.split(",") if r.strip()]
    districts_list = [d.strip() for d in fed_districts_raw.split(",") if d.strip()]
    prof_activities_list = [p.strip() for p in prof_activities_raw.split(",") if p.strip()]

    # Build multi-dimensional combinations: (region/district) × prof_activity
    geo_filters = []
    if regions_list or districts_list:
        for r in regions_list:
            geo_filters.append(("region", r))
        for d in districts_list:
            geo_filters.append(("fed_district", d))
    else:
        geo_filters = [None]

    if prof_activities_list:
        combos = []
        for geo in geo_filters:
            for pa in prof_activities_list:
                combos.append((geo, pa))
        multi_requests = combos
    else:
        multi_requests = [(geo, None) for geo in geo_filters]

    try:
        seen_names: set = set()
        merged: list = []
        for geo_extra, pa_extra in multi_requests:
            params = dict(base_params)
            if geo_extra is not None:
                params[geo_extra[0]] = geo_extra[1]
            if pa_extra is not None:
                params["prof_activity__contains"] = pa_extra
            chunk, _ = _bitrix_request_first_available(
                BITRIX_ORGANIZATIONS_ENDPOINTS,
                params=params,
            )
            if isinstance(chunk, list):
                for org in chunk:
                    name = org.get("name") or org.get("full_name", "")
                    if name not in seen_names:
                        seen_names.add(name)
                        merged.append(org)
        return Response(merged)
    except Exception:
        return Response(
            {"detail": "Ошибка при обращении к внешнему API"},
            status=status.HTTP_502_BAD_GATEWAY,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_organizations_our_side(request):
    """
    Список организаций Bitrix с is_our_side=true (для выбора ИНН без .env).

    Проксирует запрос с параметром is_our_side=true и собирает страницы ответа.
    """
    try:
        ps = request.query_params.get("page_size", "500")
        try:
            page_size = max(1, min(int(ps), 1000))
        except ValueError:
            page_size = 500
        rows = _bitrix_list_organizations_paginated(
            {"is_our_side": "true", "page_size": page_size}
        )
        rows = _filter_our_side_org_rows(rows)
        return Response(rows)
    except Exception as exc:
        return _bitrix_error_response(
            exc,
            "Ошибка при получении списка наших организаций из Bitrix",
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_fed_districts(request):
    """Proxy to Bitrix federal districts API."""
    try:
        data, _ = _bitrix_request_first_available(
            ("/contacts/api/get_all/fed_district/", "/api/get_all/fed_district/"),
        )
        return Response(data)
    except Exception:
        return Response(
            {"detail": "Ошибка при обращении к внешнему API"},
            status=status.HTTP_502_BAD_GATEWAY,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_regions(request):
    """Proxy to Bitrix regions API."""
    try:
        data, _ = _bitrix_request_first_available(
            ("/contacts/api/get_all/region/", "/api/get_all/region/"),
        )
        return Response(data)
    except Exception:
        return Response(
            {"detail": "Ошибка при обращении к внешнему API"},
            status=status.HTTP_502_BAD_GATEWAY,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_org_types(request):
    """Proxy to Bitrix organization types API."""
    try:
        data, _ = _bitrix_request_first_available(
            ("/contacts/api/get_all/organization_type/", "/api/get_all/organization_type/"),
        )
        return Response(data)
    except Exception:
        return Response(
            {"detail": "Ошибка при обращении к внешнему API"},
            status=status.HTTP_502_BAD_GATEWAY,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def external_prof_activities(request):
    """Proxy to Bitrix professional activities API."""
    try:
        data, _ = _bitrix_request_first_available(
            ("/contacts/api/get_all/prof_activity/", "/api/get_all/prof_activity/"),
        )
        return Response(data)
    except Exception:
        return Response(
            {"detail": "Ошибка при обращении к внешнему API"},
            status=status.HTTP_502_BAD_GATEWAY,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def sync_external_organizations(request):
    """Create/update local Organization records from external org data."""
    org_list = request.data.get("organizations", [])
    created_count = 0
    updated_count = 0
    from apps.reference.models import Region

    region_cache = {}

    for ext_org in org_list:
        name = ext_org.get("name", "").strip()
        if not name:
            continue
        inn = (ext_org.get("inn") or "").strip()
        if not inn:
            # Подразделения без ИНН не синхронизируются из внешнего справочника (создаются вручную).
            continue

        region_name = ext_org.get("region", "")
        region_obj = None
        if region_name:
            if region_name not in region_cache:
                region_cache[region_name] = Region.objects.filter(
                    name__iexact=region_name
                ).first()
            region_obj = region_cache[region_name]

        defaults = {
            "name": ext_org.get("full_name") or name,
            "short_name": ext_org.get("name", "")[:200],
            "region": region_obj,
        }

        org, was_created = Organization.objects.update_or_create(
            inn=inn,
            defaults=defaults,
        )
        if was_created:
            created_count += 1
        else:
            updated_count += 1

    return Response({
        "created": created_count,
        "updated": updated_count,
        "total": created_count + updated_count,
    })
